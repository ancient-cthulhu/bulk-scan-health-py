#!/usr/bin/env python3
"""
Veracode Tenant Scan Health -- Executive Dashboard layer.

This module is a pure presentation/scoring layer.
"""

from __future__ import annotations

import math
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================================
# Configuration
# ==========================================================================

DASHBOARD_CONFIG: dict[str, Any] = {
    # Scan recency buckets, in days since the published scan.
    "scan_age": {
        "healthy_days": 30,
        "warning_days": 60,
        "critical_days": 90,
    },

    # Count of failing scan-health checks on an application.
    "health_issues": {
        "green_max": 0,     # <= this is GREEN
        "yellow_max": 2,    # <= this is YELLOW, above is RED
    },

    # Count of HIGH-severity failing checks on an application.
    "high_severity_issues": {
        "green_max": 0,
        "yellow_max": 1,
    },

    # Security-findings thresholds. Applied to OPEN flaws (not fixed, not
    # mitigated). "very_high" is Veracode severity 5, "high" is severity 4.
    "flaws": {
        # Which signal drives the findings axis:
        #   "auto"           - use Veracode's own score when the scan reports
        #                      one, otherwise fall back to density (default)
        #   "density"        - always use the local severity-weighted density
        #   "veracode_score" - always use Veracode's score; GRAY when absent
        "findings_model": "auto",

        # --- Absolute severity floor. Never normalized by app size, because a
        # single critical is a single critical whether the app ships 5 modules
        # or 500. Size must not be able to dilute severity. Applies under every
        # findings model, so a clean score can never mask an open critical.
        "very_high_orange": 1,   # >= this many open Very High -> at least ORANGE
        "very_high_red": 3,      # >= this many open Very High -> RED
        "high_orange": 1,        # >= this many open High      -> at least ORANGE
        "high_red": 5,           # >= this many open High      -> RED

        # --- Veracode score thresholds (0-100, higher is better).
        "score_green": 80,       # >= this -> GREEN
        "score_yellow": 60,      # >= this -> YELLOW, below -> RED

        # --- Size-normalized volume, used when no Veracode score is available.
        # Flaws are severity-weighted first so 100 lows never equal 20 highs.
        "severity_weights": {"very_high": 10.0, "high": 5.0, "medium": 1.0, "low": 0.1},

        # Denominator: "analysis_mb" | "selected_modules" | "none".
        # analysis_mb is the default. It comes from the detailedreport's
        # analysis_size_bytes, so it already covers only what Veracode actually
        # analyzed, and it scales continuously with code volume.
        # selected_modules is also correctly scoped (analyzed entry points) but
        # far coarser: a single fat JAR divides by 1.
        # Raw total module count is deliberately NOT offered. It counts
        # dependencies, and it would let an app failing check #29 (excessive
        # modules) dilute its own flaw density and look safer on this axis
        # while failing a scan-health check for the same reason.
        "density_basis": "analysis_mb",
        "min_basis": 1.0,        # floor on the denominator; stops tiny apps exploding

        # Weighted flaws per unit of basis. Calibrate against your own tenant:
        # aggregate_tenant() reports the median and p90 density of each run.
        "density_yellow": 2.0,
        "density_red": 8.0,
    },

    # Attention Score weights. Any component whose input is unavailable is
    # dropped from BOTH numerator and denominator, so the score stays 0-100
    # and missing data never silently reads as "healthy".
    "attention_score": {
        "scan_health_weight": 30,
        "scan_age_weight": 20,
        "flaw_weight": 25,
        "high_severity_issue_weight": 15,
        "trend_weight": 10,
        "data_quality_weight": 10,
        # Ceiling applied to profiles with no published scan results. Without
        # it, every never-scanned profile scores 100 and crowds genuinely
        # measured high-risk applications out of the Top 10. An unmeasured
        # application is a real gap, but it is not evidence of immediate risk.
        # Set to 100 to disable the cap.
        "never_scanned_cap": 70,
    },

    # Attention Score bands: (min, max, label, risk_level)
    "attention_bands": [
        (0, 19, "Normal", "GREEN"),
        (20, 39, "Monitor", "GREEN"),
        (40, 59, "Attention Required", "YELLOW"),
        (60, 79, "High Priority", "ORANGE"),
        (80, 100, "Immediate Attention", "RED"),
    ],

    # Organisational issue prevalence, as a share of the tenant.
    "issue_prevalence": {
        "widespread_pct": 20.0,
        "common_pct": 5.0,
    },

    "top_n_attention": 10,
    "top_n_issues": 10,

    "colors": {
        "GREEN": "2E7D32",
        "YELLOW": "F9A825",
        "ORANGE": "EF6C00",
        "RED": "C62828",
        "GRAY": "9E9E9E",
    },
    # Softer fills for large table bodies; text stays readable.
    "fills": {
        "GREEN": "D6EAD7",
        "YELLOW": "FDF0CC",
        "ORANGE": "FBDFC8",
        "RED": "F4D3D3",
        "GRAY": "E8E8E8",
    },
}

DISCLAIMER = (
    "Scan health indicates the quality and currency of security analysis. "
    "A healthy scan does not mean an application has no security vulnerabilities."
)

RISK_ORDER = {"RED": 0, "ORANGE": 1, "YELLOW": 2, "GREEN": 3, "GRAY": 4}

# Human-readable labels for the 31 checks, and the highest severity each check
# is capable of emitting. Used so the organisational heatmap can list every
# check, including the ones that nothing triggered.
CHECK_CATALOG: dict[int, tuple[str, str]] = {
    1:  ("Unnecessary / junk files uploaded", "medium"),
    2:  ("Third-party component selected as entry point", "medium"),
    3:  ("Implausible flaw count (zero or excessive)", "medium"),
    4:  ("Fatal scan errors", "high"),
    5:  ("Unscannable Java modules", "high"),
    6:  ("Unwanted file types uploaded", "medium"),
    7:  ("Nested archives", "high"),
    8:  ("Missing precompiled .NET views", "high"),
    9:  ("Missing SCA results", "medium"),
    10: ("JavaScript modules not selected", "medium"),
    11: ("Unexpected source code uploaded", "high"),
    12: ("Missing supporting files", "medium"),
    13: ("Missing debug symbols (PDB)", "medium"),
    14: ("Unsupported platform or compiler", "high"),
    15: ("Gradle wrapper selected for analysis", "high"),
    16: ("Sensitive files uploaded", "high"),
    17: ("Source repository uploaded", "medium"),
    18: ("node_modules uploaded", "medium"),
    19: ("Testing artefacts uploaded or selected", "high"),
    20: ("Excessive uploaded file count", "medium"),
    21: ("Excess Microsoft runtime components", "medium"),
    22: ("Loose Java class files", "medium"),
    23: ("Go workspace detected", "medium"),
    24: ("First-party modules not selected", "medium"),
    25: ("Over-scanning (dependencies of selected modules)", "medium"),
    26: ("Dependencies selected as entry points", "medium"),
    27: ("Duplicate files uploaded", "high"),
    28: ("Minified JavaScript uploaded", "medium"),
    29: ("Excessive module count", "medium"),
    30: ("Application not scanned recently", "medium"),
    31: ("Analysis or upload size over threshold", "medium"),
}

DEFAULT_CHECK_CATEGORIES: dict[int, str] = {
    1: "Packaging", 2: "Module Selection", 3: "Flaw Analysis",
    4: "Fatal Errors", 5: "Fatal Errors", 6: "Packaging",
    7: "Packaging", 8: "Packaging", 9: "SCA",
    10: "Module Selection", 11: "Packaging", 12: "Module Quality",
    13: "Module Quality", 14: "Fatal Errors", 15: "Module Selection",
    16: "Security Risk", 17: "Security Risk", 18: "Packaging",
    19: "Packaging", 20: "Packaging", 21: "Packaging",
    22: "Packaging", 23: "Packaging", 24: "Module Selection",
    25: "Module Selection", 26: "Module Selection", 27: "Packaging",
    28: "Packaging", 29: "Module Selection", 30: "Scan Recency",
    31: "Packaging",
}


# ==========================================================================
# Small helpers
# ==========================================================================

def _i(v: Any, d: int = 0) -> int:
    try:
        if v is None or v == "":
            return d
        return int(v)
    except (ValueError, TypeError):
        return d


def _f(v: Any, d: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return d
        return float(v)
    except (ValueError, TypeError):
        return d


def _pct(num: float, den: float, nd: int = 1) -> float:
    """Percentage that never divides by zero."""
    if not den:
        return 0.0
    return round(100.0 * num / den, nd)


def _ctx(num: int, den: int, nd: int = 1) -> str:
    """'43 / 250 (17.2%)' -- never hide the denominator."""
    return f"{num:,} / {den:,} ({_pct(num, den, nd)}%)"


def _clamp(v: float, lo: float = 0.0, hi: float = 1.0) -> float:
    return max(lo, min(hi, v))


def _worst(*levels: str) -> str:
    """Worst (most severe) risk level of those supplied, ignoring GRAY unless
    it is the only input."""
    known = [l for l in levels if l and l != "GRAY"]
    if not known:
        return "GRAY" if levels else "GRAY"
    return min(known, key=lambda l: RISK_ORDER.get(l, 99))


# ==========================================================================
# Intermediate model
# ==========================================================================

@dataclass
class HealthIssue:
    """One failing health check on one application profile."""
    app_id: str = ""
    app_name: str = ""
    sandbox: str = ""
    business_unit: str = ""
    check_number: int = 0
    check_name: str = ""
    check_label: str = ""
    category: str = ""
    severity: str = ""          # high | medium | low
    status: str = "FAILED"
    description: str = ""
    recommendation: str = ""


@dataclass
class ApplicationHealth:
    """Normalized, dashboard-ready view of one application profile."""
    app_id: str = ""
    app_name: str = ""
    sandbox: str = ""
    business_unit: str = ""
    policy: str = ""
    build_id: str = ""

    scan_status: str = ""
    health_status: str = "Unknown"      # Good | Fair | Poor | Unknown
    ever_scanned: bool = False

    last_scan_date: str = ""
    scan_age_days: int | None = None
    scan_age_bucket: str = "Never scanned"

    # Security findings. None means "not available", which is NOT zero.
    critical_flaws: int | None = None    # Veracode Very High (severity 5)
    high_flaws: int | None = None
    medium_flaws: int | None = None
    low_flaws: int | None = None
    total_flaws: int = 0
    open_policy_flaws: int = 0
    flaw_severity_available: bool = False

    veracode_score: int | None = None      # Veracode's own 0-100 scan score
    veracode_rating: str = ""
    policy_status: str = ""
    findings_basis: str = ""               # which signal drove the findings axis

    weighted_flaw_load: float = 0.0
    flaw_density: float | None = None      # None when the basis is unavailable
    density_basis_value: float = 0.0
    density_basis_label: str = ""

    sca_component_count: int = 0
    sca_status: str = "Unknown"          # Present | Missing | Unknown
    analysis_size_mb: float = 0.0
    total_upload_mb: float = 0.0
    files_uploaded: int = 0
    total_modules: int = 0
    selected_modules: int = 0
    fatal_errors: int = 0

    health_issue_count: int = 0
    high_severity_issue_count: int = 0
    medium_severity_issue_count: int = 0
    failed_checks: list[int] = field(default_factory=list)
    recommendations: list[str] = field(default_factory=list)
    top_issue: str = ""

    previous_health: str = ""
    health_trend: str = "Unknown"        # Improving | Stable | Declining | New | Unknown

    attention_score: int = 0
    attention_band: str = ""
    attention_level: str = "GRAY"
    attention_drivers: list[str] = field(default_factory=list)

    # Two independent dimensions, deliberately never merged.
    scan_health_risk: str = "GRAY"
    security_findings_risk: str = "GRAY"
    quadrant: str = ""

    review_url: str = ""
    triage_url: str = ""

    @property
    def key(self) -> str:
        return f"{self.app_id}|{self.sandbox}"

    @property
    def display_name(self) -> str:
        return f"{self.app_name} [{self.sandbox}]" if self.sandbox else self.app_name


# ==========================================================================
# Normalization
# ==========================================================================

_TREND_MAP = {
    "Improved": "Improving",
    "Degraded": "Declining",
    "Unchanged": "Stable",
    "New": "New",
    "": "Unknown",
}


def _age_bucket(days: int | None, cfg: dict) -> str:
    if days is None:
        return "Never scanned"
    a = cfg["scan_age"]
    if days <= a["healthy_days"]:
        return f"<={a['healthy_days']}d"
    if days <= a["warning_days"]:
        return f"{a['healthy_days'] + 1}-{a['warning_days']}d"
    if days <= a["critical_days"]:
        return f"{a['warning_days'] + 1}-{a['critical_days']}d"
    return f">{a['critical_days']}d"


def _scan_age_risk(days: int | None, cfg: dict) -> str:
    if days is None:
        return "GRAY"
    a = cfg["scan_age"]
    if days <= a["healthy_days"]:
        return "GREEN"
    if days <= a["warning_days"]:
        return "YELLOW"
    if days <= a["critical_days"]:
        return "ORANGE"
    return "RED"


def _health_risk(health: str) -> str:
    return {"Good": "GREEN", "Fair": "YELLOW", "Poor": "RED"}.get(health, "GRAY")


def _issue_count_risk(n: int, bounds: dict) -> str:
    if n <= bounds["green_max"]:
        return "GREEN"
    if n <= bounds["yellow_max"]:
        return "YELLOW"
    return "RED"


def flaw_weighted_load(app: ApplicationHealth, cfg: dict) -> float:
    """Severity-weighted count of open flaws.

    Weighting happens before any normalization so that a large pile of low
    findings can never add up to a handful of high ones.
    """
    w = cfg["flaws"]["severity_weights"]
    return (w["very_high"] * (app.critical_flaws or 0)
            + w["high"] * (app.high_flaws or 0)
            + w["medium"] * (app.medium_flaws or 0)
            + w["low"] * (app.low_flaws or 0))


def flaw_density_basis(app: ApplicationHealth, cfg: dict) -> tuple[float, str]:
    """Denominator used to normalize flaw volume for application size."""
    basis = cfg["flaws"]["density_basis"]
    if basis == "analysis_mb":
        return app.analysis_size_mb, "MB analyzed"
    if basis == "selected_modules":
        return float(app.selected_modules), "selected module"
    return 1.0, "application"


def _join_basis(existing: str, addition: str) -> str:
    return f"{existing}; {addition}" if existing else addition


def _severity_floor(app: ApplicationHealth, cfg: dict) -> str:
    """Risk implied by severity alone, before any size normalization.

    This is a floor, not a ratio. Application size can never pull it down, and
    no favourable score can mask it.
    """
    t = cfg["flaws"]
    vh, hi = app.critical_flaws or 0, app.high_flaws or 0
    if vh >= t["very_high_red"] or hi >= t["high_red"]:
        return "RED"
    if vh >= t["very_high_orange"] or hi >= t["high_orange"]:
        return "ORANGE"
    return "GREEN"


def _flaw_risk(app: ApplicationHealth, cfg: dict) -> str:
    """Security-findings risk.

    Severity floor first, then a volume signal that accounts for application
    size. The volume signal is Veracode's own scan score where the scan reports
    one, since that is already severity-weighted and size-adjusted, and a local
    severity-weighted density otherwise.

    GRAY when severity data is unavailable, because a flaw total without
    severities cannot be classified honestly.
    """
    if not app.ever_scanned or not app.flaw_severity_available:
        return "GRAY"
    t = cfg["flaws"]
    floor = _severity_floor(app, cfg)
    vh, hi = app.critical_flaws or 0, app.high_flaws or 0
    if floor != "GREEN":
        parts = []
        if vh:
            parts.append(f"{vh} open very-high")
        if hi:
            parts.append(f"{hi} open high")
        app.findings_basis = "Severity floor: " + ", ".join(parts)
    if floor == "RED":
        return "RED"

    model = t.get("findings_model", "auto")
    use_score = (app.veracode_score is not None
                 and model in ("auto", "veracode_score"))

    if use_score:
        app.findings_basis = _join_basis(app.findings_basis, f"Veracode score {app.veracode_score}")
        sc = app.veracode_score
        if sc < t["score_yellow"]:
            return "RED"
        if sc < t["score_green"]:
            return _worst(floor, "YELLOW")
        return floor

    if model == "veracode_score":
        # Explicitly requested but unavailable. Do not silently substitute.
        app.findings_basis = _join_basis(app.findings_basis, "Veracode score unavailable")
        return floor if floor != "GREEN" else "GRAY"

    if app.flaw_density is None:
        # Cannot normalize. Report unknown rather than assuming a large flaw
        # count is fine, but never discard a severity signal already in hand.
        app.findings_basis = _join_basis(app.findings_basis, "Size unknown; volume not normalized")
        if floor != "GREEN":
            return floor
        return "GRAY" if app.weighted_flaw_load > 0 else "GREEN"

    app.findings_basis = _join_basis(app.findings_basis, f"{app.flaw_density} weighted flaws per {app.density_basis_label}")
    if app.flaw_density >= t["density_red"]:
        return "RED"
    if app.flaw_density >= t["density_yellow"]:
        return _worst(floor, "YELLOW")
    return floor


def _sca_risk(status: str) -> str:
    return {"Present": "GREEN", "Missing": "RED", "Unknown": "GRAY"}.get(status, "GRAY")


def _trend_risk(trend: str) -> str:
    return {"Improving": "GREEN", "Stable": "GRAY",
            "Declining": "RED", "New": "GRAY", "Unknown": "GRAY"}.get(trend, "GRAY")


def _quadrant(app: ApplicationHealth) -> str:
    """Priority matrix placement. X axis is scan-health risk, Y axis is
    security-findings risk. Built only from indicators the scanner actually
    produces; no business criticality is invented."""
    sh_bad = app.scan_health_risk in ("RED", "ORANGE")
    sf_bad = app.security_findings_risk in ("RED", "ORANGE")
    if app.scan_health_risk == "GRAY" and app.security_findings_risk == "GRAY":
        return "Insufficient Data"
    if sf_bad and sh_bad:
        return "Immediate"
    if sf_bad and not sh_bad:
        return "Review"
    if sh_bad and not sf_bad:
        return "Attention"
    return "Monitor"


def build_application_health(
    health_rows: Sequence[dict],
    issue_records: Sequence[dict] | None = None,
    prev_rows: dict | None = None,
    cfg: dict | None = None,
    check_categories: dict[int, str] | None = None,
) -> tuple[list[ApplicationHealth], list[HealthIssue]]:
    """Normalize scanner output into the dashboard model.

    Matching against the previous report is done on App ID first so that
    renamed applications are still tracked, falling back to name.
    """
    cfg = cfg or DASHBOARD_CONFIG
    cats = check_categories or DEFAULT_CHECK_CATEGORIES
    issue_records = issue_records or []
    prev_rows = prev_rows or {}

    # ---- index the previous report by (app_id, sandbox) and (name, sandbox)
    prev_by_id: dict[tuple[str, str], dict] = {}
    prev_by_name: dict[tuple[str, str], dict] = {}
    for row in prev_rows.values():
        sbx = str(row.get("Sandbox") or "")
        aid = str(row.get("App ID") or "")
        if aid:
            prev_by_id[(aid, sbx)] = row
        prev_by_name[(str(row.get("App Name") or ""), sbx)] = row

    # ---- index structured issues by (app_name, sandbox); AggIssue carries no
    #      app_id, so this is the only join key available.
    issues_by_app: dict[tuple[str, str], list[dict]] = {}
    for rec in issue_records:
        k = (str(rec.get("app_name", "")), str(rec.get("sandbox", "") or ""))
        issues_by_app.setdefault(k, []).append(rec)

    apps: list[ApplicationHealth] = []
    all_issues: list[HealthIssue] = []

    for row in health_rows:
        name = str(row.get("App Name", "") or "")
        sbx = str(row.get("Sandbox", "") or "")
        app_id = str(row.get("App ID", "") or "")
        scan_status = str(row.get("Scan Status", "") or "")
        published = str(row.get("Published", "") or "")

        days_raw = row.get("Days Since Scan")
        days = days_raw if isinstance(days_raw, int) else None
        if isinstance(days_raw, str) and days_raw.isdigit():
            days = int(days_raw)

        ever_scanned = bool(published) and scan_status not in ("No Scan", "No Results")

        raw_health = str(row.get("Health", "") or "")
        # The scanner assigns Poor to never-scanned profiles. For the dashboard
        # that is unknown, not measured-and-bad.
        health = raw_health if ever_scanned else "Unknown"

        sev_flag = str(row.get("Flaw Severity Data", "") or "")
        sev_available = ever_scanned and sev_flag == "Available"

        sca_count = _i(row.get("SCA Components"))
        # SCA presence is authoritative from check #9, not from the count of
        # vulnerable components (zero vulnerable components is a good outcome).
        app_issue_recs = issues_by_app.get((name, sbx), [])
        failed_nums = sorted({_i(r.get("check_num")) for r in app_issue_recs})
        if not ever_scanned:
            sca_status = "Unknown"
        elif 9 in failed_nums:
            sca_status = "Missing"
        else:
            sca_status = "Present"

        app = ApplicationHealth(
            app_id=app_id,
            app_name=name,
            sandbox=sbx,
            business_unit=str(row.get("Business Unit", "") or ""),
            policy=str(row.get("Policy", "") or ""),
            build_id=str(row.get("Build ID", "") or ""),
            scan_status=scan_status,
            health_status=health,
            ever_scanned=ever_scanned,
            last_scan_date=published,
            scan_age_days=days,
            scan_age_bucket=_age_bucket(days, cfg),
            total_flaws=_i(row.get("Total Flaws")),
            open_policy_flaws=_i(row.get("Open Affecting Policy")),
            flaw_severity_available=sev_available,
            veracode_score=(_i(row.get("Veracode Score")) if str(row.get("Veracode Score", "")).strip() != "" else None),
            veracode_rating=str(row.get("Veracode Rating", "") or ""),
            policy_status=str(row.get("Policy Status", "") or ""),
            sca_component_count=sca_count,
            sca_status=sca_status,
            analysis_size_mb=_f(row.get("Analysis Size (MB)")),
            total_upload_mb=_f(row.get("Total Upload Size (MB)")),
            files_uploaded=_i(row.get("Files Uploaded")),
            total_modules=_i(row.get("Total Modules")),
            selected_modules=_i(row.get("Selected Modules")),
            fatal_errors=_i(row.get("Fatal Errors")),
            health_issue_count=_i(row.get("Total Issues")),
            high_severity_issue_count=_i(row.get("High Issues")),
            medium_severity_issue_count=_i(row.get("Medium Issues")),
            failed_checks=failed_nums,
            review_url=str(row.get("Review Modules URL", "") or ""),
            triage_url=str(row.get("Triage Flaws URL", "") or ""),
        )

        if sev_available:
            app.critical_flaws = _i(row.get("Open Very High Flaws"))
            app.high_flaws = _i(row.get("Open High Flaws"))
            app.medium_flaws = _i(row.get("Open Medium Flaws"))
            app.low_flaws = _i(row.get("Open Low Flaws"))

        recs = str(row.get("Recommendations", "") or "")
        app.recommendations = [r.strip() for r in recs.split("; ") if r.strip() and r != "None"]

        # ---- structured issues
        for rec in app_issue_recs:
            num = _i(rec.get("check_num"))
            label = CHECK_CATALOG.get(num, ("Check %d" % num, "medium"))[0]
            hi = HealthIssue(
                app_id=app_id, app_name=name, sandbox=sbx,
                business_unit=app.business_unit,
                check_number=num,
                check_name=str(rec.get("check_name", "") or ""),
                check_label=label,
                category=str(rec.get("category", "") or cats.get(num, "Other")),
                severity=str(rec.get("severity", "") or "medium"),
                description=str(rec.get("description", "") or ""),
                recommendation=str(rec.get("recommendation", "") or ""),
            )
            all_issues.append(hi)

        # Top issue: worst severity, then the check with the highest severity rank.
        highs = [i for i in app_issue_recs if str(i.get("severity")) == "high"]
        pool = highs or app_issue_recs
        if pool:
            n = _i(pool[0].get("check_num"))
            app.top_issue = CHECK_CATALOG.get(n, (str(pool[0].get("description", ""))[:80], ""))[0]
        elif not ever_scanned:
            app.top_issue = "No published scan results"

        # ---- trend, matched on App ID first
        prow = prev_by_id.get((app_id, sbx)) if app_id else None
        if prow is None:
            prow = prev_by_name.get((name, sbx))
        if prev_rows:
            if prow is None:
                app.health_trend = "New"
                app.previous_health = ""
            else:
                app.previous_health = str(prow.get("Health", "") or "")
                app.health_trend = _classify_trend(app.previous_health, health)
        else:
            app.health_trend = "Unknown"

        # ---- risk dimensions
        app.scan_health_risk = _worst(
            _health_risk(app.health_status),
            _issue_count_risk(app.high_severity_issue_count, cfg["high_severity_issues"]),
        ) if ever_scanned else "GRAY"
        if app.flaw_severity_available:
            app.weighted_flaw_load = round(flaw_weighted_load(app, cfg), 2)
            basis_val, basis_lbl = flaw_density_basis(app, cfg)
            app.density_basis_value, app.density_basis_label = basis_val, basis_lbl
            if basis_val > 0:
                app.flaw_density = round(
                    app.weighted_flaw_load / max(basis_val, cfg["flaws"]["min_basis"]), 2)
        app.security_findings_risk = _flaw_risk(app, cfg)
        app.quadrant = _quadrant(app)

        apps.append(app)

    # ---- scoring
    for app in apps:
        score, band, level, drivers = compute_attention_score(app, cfg)
        app.attention_score = score
        app.attention_band = band
        app.attention_level = level
        app.attention_drivers = drivers

    apps.sort(key=lambda a: (-a.attention_score, a.app_name.lower(), a.sandbox.lower()))
    return apps, all_issues


_HEALTH_RANK = {"Good": 0, "Fair": 1, "Poor": 2, "Unknown": 3, "": 3}


def _classify_trend(previous: str, current: str) -> str:
    """Direction of travel between two health classifications.

    Unknown on either side means the direction is genuinely unknown; it is not
    reported as stable.
    """
    if not previous or previous == "Unknown" or current == "Unknown":
        return "Unknown"
    pr, cr = _HEALTH_RANK.get(previous, 3), _HEALTH_RANK.get(current, 3)
    if pr == 3 or cr == 3:
        return "Unknown"
    if cr < pr:
        return "Improving"
    if cr > pr:
        return "Declining"
    return "Stable"


# ==========================================================================
# Attention Score
# ==========================================================================

def compute_attention_score(app: ApplicationHealth,
                            cfg: dict | None = None) -> tuple[int, str, str, list[str]]:
    """Return (score 0-100, band label, risk level, explanation drivers).

    Each component contributes ``weight * risk_fraction``. Components whose
    input is unavailable are excluded from the denominator as well as the
    numerator, so an application is never rewarded for missing data and never
    penalised twice for it (the data-quality component covers that).
    """
    cfg = cfg or DASHBOARD_CONFIG
    w = cfg["attention_score"]
    parts: list[tuple[str, float, float, str]] = []   # (name, weight, fraction, reason)

    # 1. Scan health classification
    frac = {"Good": 0.0, "Fair": 0.5, "Poor": 1.0}.get(app.health_status)
    if frac is None:
        parts.append(("Scan health", w["scan_health_weight"], 1.0,
                      "Scan health is unknown (no published scan results)"))
    else:
        reason = f"Scan health is {app.health_status.upper()}"
        parts.append(("Scan health", w["scan_health_weight"], frac, reason))

    # 2. Scan recency
    a = cfg["scan_age"]
    if app.scan_age_days is None:
        parts.append(("Scan recency", w["scan_age_weight"], 1.0, "Application has never been scanned"))
    else:
        d = app.scan_age_days
        if d <= a["healthy_days"]:
            frac = 0.0
        elif d <= a["warning_days"]:
            frac = 0.4
        elif d <= a["critical_days"]:
            frac = 0.7
        else:
            frac = 1.0
        parts.append(("Scan recency", w["scan_age_weight"], frac, f"Scan is {d} days old"))

    # 3. Security findings
    t = cfg["flaws"]
    if not app.ever_scanned:
        pass  # covered by data quality; no findings dimension exists yet
    elif not app.flaw_severity_available:
        parts.append(("Security findings", w["flaw_weight"], 0.5,
                      "Flaw severity breakdown unavailable for this scan"))
    else:
        vh, hi = app.critical_flaws or 0, app.high_flaws or 0
        frac = 0.0
        bits: list[str] = []
        # Severity contributes absolutely; size never dilutes it.
        if vh:
            frac = max(frac, _clamp(0.7 + 0.3 * (vh / max(3 * t["very_high_red"], 1))))
            bits.append(f"{vh} open very-high flaw(s)")
        if hi:
            frac = max(frac, _clamp(0.4 + 0.4 * (hi / max(t["high_red"], 1))))
            bits.append(f"{hi} open high flaw(s)")
        # Volume contributes relative to application size. Veracode's own score
        # already accounts for both severity and size, so prefer it.
        if app.veracode_score is not None:
            if app.veracode_score < t["score_green"]:
                frac = max(frac, _clamp((t["score_green"] - app.veracode_score)
                                        / max(t["score_green"], 1)))
                bits.append(f"Veracode score {app.veracode_score}"
                            + (f" (rating {app.veracode_rating})" if app.veracode_rating else ""))
        elif app.flaw_density is not None:
            if app.flaw_density >= t["density_yellow"]:
                frac = max(frac, _clamp(app.flaw_density / max(t["density_red"], 0.01)))
                bits.append(f"{app.flaw_density} weighted flaws per "
                            f"{app.density_basis_label} ({app.weighted_flaw_load} weighted "
                            f"across {app.density_basis_value:g})")
        elif app.weighted_flaw_load > 0:
            frac = max(frac, 0.5)
            bits.append("Flaw volume cannot be normalized; application size is unknown")
        if _policy_level(app.policy_status) == "RED":
            frac = max(frac, 0.6)
            bits.append(f"Policy status: {app.policy_status}")
        reason = ", ".join(bits) if bits else "No open very-high or high flaws, low flaw volume"
        parts.append(("Security findings", w["flaw_weight"], frac, reason))

    # 4. High-severity scan-health checks
    if app.ever_scanned:
        n = app.high_severity_issue_count
        frac = {0: 0.0, 1: 0.5, 2: 0.75}.get(n, 1.0)
        reason = (f"{n} high-severity scan-health issue(s)" if n
                  else "No high-severity scan-health issues")
        parts.append(("High-severity issues", w["high_severity_issue_weight"], frac, reason))

    # 5. Trend
    if app.health_trend in ("Improving", "Stable", "Declining"):
        frac = {"Improving": 0.0, "Stable": 0.25, "Declining": 1.0}[app.health_trend]
        if app.health_trend == "Declining":
            reason = f"Health declined from {app.previous_health.upper()} to {app.health_status.upper()}"
        elif app.health_trend == "Improving":
            reason = f"Health improved from {app.previous_health.upper()} to {app.health_status.upper()}"
        else:
            reason = f"Health unchanged at {app.health_status.upper()}"
        parts.append(("Trend", w["trend_weight"], frac, reason))

    # 6. Data quality
    gaps: list[str] = []
    if not app.ever_scanned:
        gaps.append("no published scan results")
    else:
        if not app.flaw_severity_available:
            gaps.append("no flaw severity data")
        if app.total_modules == 0:
            gaps.append("no module data")
        if app.sca_status == "Unknown":
            gaps.append("SCA status unknown")
    dq = _clamp(len(gaps) / 2.0) if gaps else 0.0
    if not app.ever_scanned:
        dq = 1.0
    parts.append(("Data quality", w["data_quality_weight"], dq,
                  "Data gaps: " + ", ".join(gaps) if gaps else "Complete data"))

    total_w = sum(p[1] for p in parts)
    if total_w <= 0:
        return 0, "Normal", "GREEN", ["No scoring inputs available"]

    raw = sum(p[1] * p[2] for p in parts)
    score = int(round(100.0 * raw / total_w))
    score = max(0, min(100, score))

    capped = False
    cap = w.get("never_scanned_cap", 100)
    if not app.ever_scanned and score > cap:
        score = cap
        capped = True

    band, level = "Normal", "GREEN"
    for lo, hi, lbl, lvl in cfg["attention_bands"]:
        if lo <= score <= hi:
            band, level = lbl, lvl
            break

    # Drivers: only components that actually pushed the score up, biggest first.
    contribs = sorted(
        [(p[0], p[1] * p[2], p[3]) for p in parts if p[2] > 0],
        key=lambda x: -x[1],
    )
    drivers = [f"{reason} (+{int(round(100.0 * c / total_w))})" for _, c, reason in contribs]
    if capped:
        drivers.append(f"Score capped at {cap}: risk is unmeasured, not confirmed")
    if not drivers:
        drivers = ["No attention drivers; all measured dimensions are healthy"]
    return score, band, level, drivers


# ==========================================================================
# Aggregation
# ==========================================================================

@dataclass
class TenantMetrics:
    generated: str = ""
    total_apps: int = 0
    scanned_apps: int = 0
    never_scanned: int = 0
    good: int = 0
    fair: int = 0
    poor: int = 0
    unknown: int = 0
    stale: int = 0
    fresh: int = 0
    with_high_severity_issues: int = 0
    with_critical_or_high_flaws: int = 0
    flaw_severity_unavailable: int = 0
    total_flaws: int = 0
    open_policy_flaws: int = 0
    critical_flaws: int = 0
    high_flaws: int = 0
    avg_flaws_per_app: float = 0.0
    avg_scan_age: float | None = None
    median_scan_age: float | None = None
    policy_pass: int = 0
    policy_conditional: int = 0
    policy_fail: int = 0
    policy_unknown: int = 0
    with_veracode_score: int = 0
    median_flaw_density: float | None = None
    p90_flaw_density: float | None = None
    density_basis_label: str = ""
    total_sca_components: int = 0
    total_upload_mb: float = 0.0
    total_analysis_mb: float = 0.0
    missing_sca: int = 0
    attention_bands: dict[str, int] = field(default_factory=dict)
    quadrants: dict[str, int] = field(default_factory=dict)
    trends: dict[str, int] = field(default_factory=dict)

    @property
    def scan_compliance_pct(self) -> float:
        return _pct(self.fresh, self.total_apps)

    @property
    def policy_pass_pct(self) -> float:
        """Share of profiles with a known policy outcome that passed.
        Unknown outcomes are excluded from the denominator rather than counted
        as failures."""
        known = self.policy_pass + self.policy_conditional + self.policy_fail
        return _pct(self.policy_pass, known)

    @property
    def healthy_pct(self) -> float:
        return _pct(self.good, self.total_apps)


def _percentile(sorted_vals: list[float], p: float) -> float:
    """Nearest-rank percentile. Deterministic, no numpy dependency."""
    if not sorted_vals:
        return 0.0
    k = max(0, min(len(sorted_vals) - 1, int(math.ceil(p / 100.0 * len(sorted_vals))) - 1))
    return round(sorted_vals[k], 2)


def aggregate_tenant(apps: Sequence[ApplicationHealth], cfg: dict | None = None) -> TenantMetrics:
    cfg = cfg or DASHBOARD_CONFIG
    m = TenantMetrics(
        generated=datetime.now().strftime("%Y-%m-%d %H:%M"),
        total_apps=len(apps),
    )
    ages: list[int] = []
    for a in apps:
        if a.ever_scanned:
            m.scanned_apps += 1
        else:
            m.never_scanned += 1
        m.good += a.health_status == "Good"
        m.fair += a.health_status == "Fair"
        m.poor += a.health_status == "Poor"
        m.unknown += a.health_status == "Unknown"

        if a.scan_age_days is None:
            pass
        else:
            ages.append(a.scan_age_days)
            if a.scan_age_days <= cfg["scan_age"]["healthy_days"]:
                m.fresh += 1
            else:
                m.stale += 1

        pl = _policy_level(a.policy_status)
        m.policy_pass += pl == "GREEN"
        m.policy_conditional += pl == "YELLOW"
        m.policy_fail += pl == "RED"
        m.policy_unknown += pl == "GRAY"
        m.with_veracode_score += a.veracode_score is not None

        m.with_high_severity_issues += a.high_severity_issue_count > 0
        if a.flaw_severity_available:
            m.critical_flaws += a.critical_flaws or 0
            m.high_flaws += a.high_flaws or 0
            if (a.critical_flaws or 0) or (a.high_flaws or 0):
                m.with_critical_or_high_flaws += 1
        elif a.ever_scanned:
            m.flaw_severity_unavailable += 1

        m.total_flaws += a.total_flaws
        m.open_policy_flaws += a.open_policy_flaws
        m.total_sca_components += a.sca_component_count
        m.total_upload_mb += a.total_upload_mb
        m.total_analysis_mb += a.analysis_size_mb
        m.missing_sca += a.sca_status == "Missing"

        m.attention_bands[a.attention_level] = m.attention_bands.get(a.attention_level, 0) + 1
        m.quadrants[a.quadrant] = m.quadrants.get(a.quadrant, 0) + 1
        m.trends[a.health_trend] = m.trends.get(a.health_trend, 0) + 1

    dens = sorted(a.flaw_density for a in apps if a.flaw_density is not None)
    if dens:
        m.median_flaw_density = _percentile(dens, 50)
        m.p90_flaw_density = _percentile(dens, 90)
        m.density_basis_label = next(a.density_basis_label for a in apps
                                     if a.flaw_density is not None)
    m.avg_flaws_per_app = round(m.total_flaws / len(apps), 1) if apps else 0.0
    if ages:
        m.avg_scan_age = round(sum(ages) / len(ages), 1)
        s = sorted(ages)
        mid = len(s) // 2
        m.median_scan_age = float(s[mid]) if len(s) % 2 else round((s[mid - 1] + s[mid]) / 2, 1)
    m.total_upload_mb = round(m.total_upload_mb, 1)
    m.total_analysis_mb = round(m.total_analysis_mb, 1)
    return m


def build_issue_heatmap(issues: Sequence[HealthIssue], total_apps: int,
                        cfg: dict | None = None,
                        prev_issue_counts: dict[int, int] | None = None) -> list[dict]:
    """One row per health check, including checks that nothing triggered.

    Severity-aware: colour combines the check's severity with how widespread it
    is, so 'a few serious issues' stays visually distinct from 'many small
    issues'.
    """
    cfg = cfg or DASHBOARD_CONFIG
    prev_issue_counts = prev_issue_counts or {}
    by_check: dict[int, list[HealthIssue]] = {}
    for i in issues:
        by_check.setdefault(i.check_number, []).append(i)

    p = cfg["issue_prevalence"]
    rows: list[dict] = []
    for num in sorted(CHECK_CATALOG):
        label, max_sev = CHECK_CATALOG[num]
        group = by_check.get(num, [])
        apps_hit = sorted({(g.app_name, g.sandbox) for g in group})
        n = len(apps_hit)
        pct = _pct(n, total_apps)

        if group:
            sev = "HIGH" if any(g.severity == "high" for g in group) else "MEDIUM"
        else:
            sev = max_sev.upper()

        if n == 0:
            level = "GREEN"
        elif sev == "HIGH":
            level = "RED" if pct >= p["common_pct"] else "ORANGE"
        else:
            level = "ORANGE" if pct >= p["widespread_pct"] else "YELLOW"

        prev_n = prev_issue_counts.get(num)
        if prev_n is None:
            trend = "Unknown"
        elif n < prev_n:
            trend = "Improving"
        elif n > prev_n:
            trend = "Declining"
        else:
            trend = "Stable"

        bus = sorted({g.business_unit for g in group if g.business_unit})
        rows.append({
            "Check #": num,
            "Issue": label,
            "Category": group[0].category if group else DEFAULT_CHECK_CATEGORIES.get(num, "Other"),
            "Severity": sev,
            "Apps Affected": n,
            "Total Apps": total_apps,
            "% of Tenant": pct,
            "Occurrences": len(group),
            "Business Units Affected": len(bus),
            "Previous Apps Affected": prev_n if prev_n is not None else "N/A",
            "Trend": trend,
            "_level": level,
            "_apps": [f"{a}{' [' + s + ']' if s else ''}" for a, s in apps_hit],
            "Top Recommendation": next((g.recommendation for g in group if g.recommendation), ""),
        })

    sev_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (-r["Apps Affected"], sev_rank.get(r["Severity"], 3)))
    return rows


def build_tenant_trend(current: TenantMetrics,
                       prev_rows: dict | None,
                       cfg: dict | None = None) -> list[dict]:
    """Tenant-level previous-vs-current comparison.

    Direction is semantic: more healthy applications is an improvement, more
    total flaws is a deterioration. The arrow never simply follows the sign of
    the delta.
    """
    if not prev_rows:
        return []
    cfg = cfg or DASHBOARD_CONFIG

    rows = list(prev_rows.values())
    p_total = len(rows)
    p_good = sum(1 for r in rows if r.get("Health") == "Good")
    p_fair = sum(1 for r in rows if r.get("Health") == "Fair")
    p_poor = sum(1 for r in rows if r.get("Health") == "Poor")
    p_flaws = sum(_i(r.get("Total Flaws")) for r in rows)
    p_open = sum(_i(r.get("Open Affecting Policy")) for r in rows)
    p_stale = 0
    for r in rows:
        d = r.get("Days Since Scan")
        if isinstance(d, int) and d > cfg["scan_age"]["healthy_days"]:
            p_stale += 1
        elif isinstance(d, str) and d.isdigit() and int(d) > cfg["scan_age"]["healthy_days"]:
            p_stale += 1

    # (metric, previous, current, higher_is_better)
    spec = [
        ("Total applications", p_total, current.total_apps, None),
        ("Healthy (Good)", p_good, current.good, True),
        ("Fair", p_fair, current.fair, False),
        ("Poor", p_poor, current.poor, False),
        ("Stale scans", p_stale, current.stale, False),
        ("Total flaws", p_flaws, current.total_flaws, False),
        ("Open policy-affecting flaws", p_open, current.open_policy_flaws, False),
    ]

    out: list[dict] = []
    for name, prev, curr, higher_better in spec:
        delta = curr - prev
        if delta == 0:
            arrow, level = "=", "GRAY"
        elif higher_better is None:
            arrow, level = ("+" if delta > 0 else "-"), "GRAY"
        else:
            improved = (delta > 0) == bool(higher_better)
            arrow = "UP" if delta > 0 else "DOWN"
            level = "GREEN" if improved else "RED"
        out.append({
            "Metric": name,
            "Previous": prev,
            "Current": curr,
            "Change": f"{delta:+d}" if delta else "0",
            "Direction": arrow,
            "_level": level,
        })
    return out


def build_heatmap_rows(apps: Sequence[ApplicationHealth],
                       cfg: dict | None = None) -> list[dict]:
    """Application heatmap. '_lv_<Column>' keys carry the colour for each cell
    and are hidden from the rendered table."""
    cfg = cfg or DASHBOARD_CONFIG
    rows: list[dict] = []
    for a in apps:
        crit = a.critical_flaws if a.flaw_severity_available else "N/A"
        high = a.high_flaws if a.flaw_severity_available else "N/A"
        med = a.medium_flaws if a.flaw_severity_available else "N/A"
        low = a.low_flaws if a.flaw_severity_available else "N/A"

        row = {
            "Application": a.app_name,
            "Sandbox": a.sandbox or "(policy)",
            "Business Unit": a.business_unit or "Unknown",
            "Attention Score": a.attention_score,
            "Priority": a.attention_band,
            "Scan Health": a.health_status,
            "Scan Age": a.scan_age_days if a.scan_age_days is not None else "Never",
            "Scan Age Bucket": a.scan_age_bucket,
            "Findings Risk": a.security_findings_risk if a.flaw_severity_available else "Unknown",
            "Very High Flaws": crit,
            "High Flaws": high,
            "Medium Flaws": med,
            "Low Flaws": low,
            "Veracode Score": a.veracode_score if a.veracode_score is not None else "N/A",
            "Veracode Rating": a.veracode_rating or "N/A",
            "Policy Status": a.policy_status or "Unknown",
            "Findings Basis": a.findings_basis or "N/A",
            "Weighted Flaws": a.weighted_flaw_load if a.flaw_severity_available else "N/A",
            "Flaw Density": (a.flaw_density if a.flaw_density is not None
                             else ("N/A" if a.flaw_severity_available else "Unknown")),
            "Density Basis": (f"{a.density_basis_value:g} {a.density_basis_label}"
                              if a.flaw_severity_available else "Unknown"),
            "Total Flaws": a.total_flaws if a.ever_scanned else "N/A",
            "Open Policy Flaws": a.open_policy_flaws if a.ever_scanned else "N/A",
            "SCA": a.sca_status,
            "SCA Components": a.sca_component_count,
            "Health Issues": a.health_issue_count if a.ever_scanned else "N/A",
            "High-Severity Issues": a.high_severity_issue_count if a.ever_scanned else "N/A",
            "Upload Size (MB)": a.total_upload_mb,
            "Last Scan": a.last_scan_date or "Never",
            "Previous Health": a.previous_health or "N/A",
            "Change": {"Improving": "UP", "Declining": "DOWN",
                       "Stable": "=", "New": "NEW"}.get(a.health_trend, "N/A"),
            "Trend": a.health_trend,
            "Quadrant": a.quadrant,
            "Top Issue": a.top_issue or "None",
            # hidden colour helpers
            "_lv_Attention Score": a.attention_level,
            "_lv_Priority": a.attention_level,
            "_lv_Scan Health": _health_risk(a.health_status),
            "_lv_Scan Age": _scan_age_risk(a.scan_age_days, cfg),
            "_lv_Scan Age Bucket": _scan_age_risk(a.scan_age_days, cfg),
            "_lv_Findings Risk": a.security_findings_risk,
            "_lv_Flaw Density": _density_level(a, cfg),
            "_lv_Veracode Score": _score_level(a.veracode_score, cfg),
            "_lv_Policy Status": _policy_level(a.policy_status),
            "_lv_Very High Flaws": _sev_cell_level(crit, 1, 1),
            "_lv_High Flaws": _sev_cell_level(high, 1, cfg["flaws"]["high_red"]),
            "_lv_SCA": _sca_risk(a.sca_status),
            "_lv_Health Issues": (_issue_count_risk(a.health_issue_count, cfg["health_issues"])
                                  if a.ever_scanned else "GRAY"),
            "_lv_High-Severity Issues": (_issue_count_risk(a.high_severity_issue_count,
                                                           cfg["high_severity_issues"])
                                         if a.ever_scanned else "GRAY"),
            "_lv_Previous Health": _health_risk(a.previous_health) if a.previous_health else "GRAY",
            "_lv_Change": _trend_risk(a.health_trend),
            "_lv_Trend": _trend_risk(a.health_trend),
            "_key": a.key,
        }
        rows.append(row)
    return rows


def _score_level(score: int | None, cfg: dict) -> str:
    if score is None:
        return "GRAY"
    t = cfg["flaws"]
    if score >= t["score_green"]:
        return "GREEN"
    if score >= t["score_yellow"]:
        return "YELLOW"
    return "RED"


def _policy_level(status: str) -> str:
    s = (status or "").strip().lower()
    if not s:
        return "GRAY"
    if s.startswith("pass"):
        return "GREEN"
    if "conditional" in s:
        return "YELLOW"
    if "did not pass" in s or s.startswith("fail"):
        return "RED"
    return "GRAY"


def _density_level(app: ApplicationHealth, cfg: dict) -> str:
    if app.flaw_density is None:
        return "GRAY"
    t = cfg["flaws"]
    if app.flaw_density >= t["density_red"]:
        return "RED"
    if app.flaw_density >= t["density_yellow"]:
        return "YELLOW"
    return "GREEN"


def _sev_cell_level(v: Any, yellow_at: int, red_at: int) -> str:
    if not isinstance(v, int):
        return "GRAY"
    if v >= red_at:
        return "RED"
    if v >= yellow_at:
        return "ORANGE"
    return "GREEN"


def build_priority_list(apps: Sequence[ApplicationHealth],
                        n: int | None = None,
                        cfg: dict | None = None) -> list[dict]:
    cfg = cfg or DASHBOARD_CONFIG
    n = n or cfg["top_n_attention"]
    out: list[dict] = []
    for rank, a in enumerate(apps[:n], 1):
        ch = (f"{a.critical_flaws} / {a.high_flaws}"
              if a.flaw_severity_available else "Not available")
        out.append({
            "Priority": rank,
            "Application": a.display_name,
            "Business Unit": a.business_unit or "Unknown",
            "Attention Score": a.attention_score,
            "Band": a.attention_band,
            "Scan Health": a.health_status,
            "Very High / High Flaws": ch,
            "Scan Age (days)": a.scan_age_days if a.scan_age_days is not None else "Never scanned",
            "Top Issue": a.top_issue or "None",
            "Trend": a.health_trend,
            "Why": "; ".join(a.attention_drivers[:2]),
            "_level": a.attention_level,
        })
    return out


# ==========================================================================
# Excel rendering
# ==========================================================================

_FONT = "Arial"
_TITLE_FILL = PatternFill("solid", fgColor="14213D")
_BAND_FILL = PatternFill("solid", fgColor="1F4E79")
_SUB_FILL = PatternFill("solid", fgColor="E8EDF3")
_KPI_FILL = PatternFill("solid", fgColor="F5F7FA")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _lvl_fill(level: str, cfg: dict, strong: bool = False) -> PatternFill:
    key = "colors" if strong else "fills"
    return PatternFill("solid", fgColor=cfg[key].get(level, cfg[key]["GRAY"]))


def _lvl_font(level: str, cfg: dict, bold: bool = False, size: int = 10) -> Font:
    return Font(name=_FONT, size=size, bold=bold, color=cfg["colors"].get(level, "000000"))


def _write_title(ws, row: int, text: str, width: int, fill=_BAND_FILL,
                 size: int = 11, color: str = "FFFFFF") -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = Font(name=_FONT, size=size, bold=True, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22 if size <= 12 else 30
    return row + 1


def _write_table(ws, row: int, headers: list[str], data: list[list],
                 levels: list[list[str]] | None, cfg: dict,
                 spans: list[int] | None = None,
                 left_cols: Iterable[int] = (),
                 row_height: int | None = None) -> int:
    """Write a table on the fixed dashboard grid.

    `spans` gives the number of grid columns each logical column occupies, so
    long-text columns can breathe without any block resizing the shared grid.
    Column widths are owned by the sheet, never by individual tables.
    """
    spans = spans or [1] * len(headers)
    left = set(left_cols)

    def _cells(r: int, values: list, styler) -> None:
        col = 1
        for idx, v in enumerate(values):
            span = spans[idx]
            if span > 1:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + span - 1)
            c = ws.cell(row=r, column=col, value=v)
            styler(c, idx)
            for cc in range(col, col + span):
                ws.cell(row=r, column=cc).border = _BORDER
            col += span

    def _hdr_style(c, idx):
        c.fill = _SUB_FILL
        c.font = Font(name=_FONT, size=9, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="left" if idx in left else "center",
                                vertical="center", wrap_text=True, indent=1 if idx in left else 0)

    _cells(row, headers, _hdr_style)
    row += 1

    for ri, rowvals in enumerate(data):
        def _body_style(c, idx, ri=ri):
            c.font = Font(name=_FONT, size=10)
            c.alignment = Alignment(horizontal="left" if idx in left else "center",
                                    vertical="center", wrap_text=True,
                                    indent=1 if idx in left else 0)
            if levels:
                lv = levels[ri][idx]
                if lv:
                    c.fill = _lvl_fill(lv, cfg)
                    if lv in ("RED", "ORANGE"):
                        c.font = Font(name=_FONT, size=10, bold=True, color=cfg["colors"][lv])
        _cells(row + ri, rowvals, _body_style)
        if row_height:
            ws.row_dimensions[row + ri].height = row_height
    return row + len(data)


def write_dashboard_sheets(wb: Workbook,
                           apps: Sequence[ApplicationHealth],
                           issues: Sequence[HealthIssue],
                           metrics: TenantMetrics,
                           tenant_trend: list[dict],
                           issue_rows: list[dict],
                           cfg: dict | None = None) -> None:
    """Insert 'Executive Dashboard', 'App Heatmap' and 'Issue Heatmap' as the
    first three sheets of an existing workbook. Existing sheets are untouched.

    The dashboard sheet uses a mixed block layout, which cannot carry an Excel
    autofilter. The two heatmap sheets carry the filterable/sortable tables.
    """
    cfg = cfg or DASHBOARD_CONFIG
    _write_exec_sheet(wb, apps, metrics, tenant_trend, issue_rows, cfg)
    _write_app_heatmap_sheet(wb, apps, cfg)
    _write_issue_heatmap_sheet(wb, issue_rows, metrics, cfg)
    # Order: dashboard, app heatmap, issue heatmap, then everything pre-existing.
    order = ["Executive Dashboard", "App Heatmap", "Issue Heatmap"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else len(order))


# Fixed 12-column grid for the dashboard sheet. Set once, never overridden.
_EXEC_GRID = 12
_EXEC_WIDTHS = {1: 26}
for _c in range(2, _EXEC_GRID + 1):
    _EXEC_WIDTHS[_c] = 13


def _write_exec_sheet(wb: Workbook, apps: Sequence[ApplicationHealth],
                      m: TenantMetrics, tenant_trend: list[dict],
                      issue_rows: list[dict], cfg: dict) -> None:
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    W = _EXEC_GRID
    for ci, wd in _EXEC_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = wd

    def band(r: int, text: str) -> int:
        return _write_title(ws, r, text, W)

    def note(r: int, text: str) -> int:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=W)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=_FONT, size=9, italic=True, color="555555")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        return r + 1

    r = 1
    # ---- title
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=W)
    c = ws.cell(row=r, column=1, value="VERACODE TENANT SCAN HEALTH")
    c.fill = _TITLE_FILL
    c.font = Font(name=_FONT, size=18, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 26
    ws.row_dimensions[r + 1].height = 16
    r += 2
    r = note(r, f"Generated {m.generated}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=W)
    c = ws.cell(row=r, column=1, value=DISCLAIMER)
    c.font = Font(name=_FONT, size=9, italic=True, color="8A6D00")
    c.fill = PatternFill("solid", fgColor="FDF6E3")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _BORDER
    ws.row_dimensions[r].height = 20
    r += 2

    # ---- KPI cards
    comp_lv = ("GREEN" if m.scan_compliance_pct >= 80
               else "YELLOW" if m.scan_compliance_pct >= 60 else "RED")
    r = _kpi_row(ws, r, [
        ("TOTAL APPS", f"{m.total_apps:,}", "profiles evaluated", "GRAY"),
        ("HEALTHY", f"{m.good:,}", f"{_pct(m.good, m.total_apps)}% of tenant", "GREEN"),
        ("NEEDS ATTENTION", f"{m.fair:,}", f"{_pct(m.fair, m.total_apps)}% of tenant", "YELLOW"),
        ("POOR", f"{m.poor:,}", f"{_pct(m.poor, m.total_apps)}% of tenant", "RED"),
        ("UNKNOWN", f"{m.unknown:,}", f"{_pct(m.unknown, m.total_apps)}% never scanned", "GRAY"),
    ], cfg, span=W)

    r = _kpi_row(ws, r, [
        ("SCAN COMPLIANCE", f"{m.fresh:,} / {m.total_apps:,}",
         f"{m.scan_compliance_pct}% scanned within {cfg['scan_age']['healthy_days']}d", comp_lv),
        ("STALE SCANS", f"{m.stale:,} / {m.total_apps:,}",
         f"{_pct(m.stale, m.total_apps)}% of tenant", "RED" if m.stale else "GREEN"),
        ("HIGH-SEV ISSUES", f"{m.with_high_severity_issues:,} / {m.total_apps:,}",
         f"{_pct(m.with_high_severity_issues, m.total_apps)}% of tenant",
         "RED" if m.with_high_severity_issues else "GREEN"),
        ("POLICY FAILING", f"{m.policy_fail:,} / {m.policy_pass + m.policy_conditional + m.policy_fail:,}",
         (f"{m.policy_pass_pct}% of evaluated apps pass"
          if (m.policy_pass + m.policy_conditional + m.policy_fail)
          else "no policy outcome reported"),
         "RED" if m.policy_fail else ("GRAY" if not m.policy_pass else "GREEN")),
        ("VERY HIGH / HIGH", f"{m.critical_flaws:,} / {m.high_flaws:,}",
         (f"{m.flaw_severity_unavailable} app(s) lack severity data"
          if m.flaw_severity_unavailable else "open flaws across tenant"),
         "RED" if (m.critical_flaws or m.high_flaws) else "GREEN"),
    ], cfg, span=W)

    # ---- tenant health distribution
    r = band(r, "TENANT SCAN HEALTH DISTRIBUTION")
    for label, n, lv in (("Good", m.good, "GREEN"), ("Fair", m.fair, "YELLOW"),
                         ("Poor", m.poor, "RED"), ("Unknown", m.unknown, "GRAY")):
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name=_FONT, size=10, bold=True, color=cfg["colors"][lv])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=_ctx(n, m.total_apps))
        c.font = Font(name=_FONT, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=W)
        c = ws.cell(row=r, column=4, value="\u2588" * int(round(50 * _pct(n, m.total_apps) / 100.0)))
        c.font = Font(name=_FONT, size=10, color=cfg["colors"][lv])
        c.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    r = note(r, "Unknown covers profiles with no published scan results. "
                "They are never counted as healthy.")
    r += 1

    # ---- attention distribution
    r = band(r, "ATTENTION SCORE DISTRIBUTION")
    meaning = {"RED": "Immediate attention (80-100)", "ORANGE": "High priority (60-79)",
               "YELLOW": "Attention required (40-59)", "GREEN": "Normal or monitor (0-39)"}
    data, levels = [], []
    for lv in ("RED", "ORANGE", "YELLOW", "GREEN"):
        n = m.attention_bands.get(lv, 0)
        data.append([lv, _ctx(n, m.total_apps), meaning[lv]])
        levels.append([lv, "", ""])
    r = _write_table(ws, r, ["Band", "Applications", "Meaning"], data, levels, cfg,
                     spans=[2, 3, 7], left_cols=(2,))
    r += 1

    # ---- priority matrix
    r = band(r, "PRIORITY MATRIX   scan health vs security findings")
    r = note(r, "Two independent dimensions. Built only from indicators the scanner produces; "
                "business criticality is not available from the Veracode data and is not inferred.")
    q = m.quadrants
    grid = [
        ["", "LOW scan-health risk", "HIGH scan-health risk"],
        ["HIGH findings risk", f"REVIEW\n{q.get('Review', 0)} apps", f"IMMEDIATE\n{q.get('Immediate', 0)} apps"],
        ["LOW findings risk", f"MONITOR\n{q.get('Monitor', 0)} apps", f"ATTENTION\n{q.get('Attention', 0)} apps"],
    ]
    glevels = [["", "", ""], ["", "ORANGE", "RED"], ["", "GREEN", "YELLOW"]]
    for ri, gr in enumerate(grid):
        col = 1
        for ci, v in enumerate(gr):
            span = 1 if ci == 0 else 5
            if span > 1:
                ws.merge_cells(start_row=r + ri, start_column=col, end_row=r + ri,
                               end_column=col + span - 1)
            cell = ws.cell(row=r + ri, column=col, value=v)
            cell.font = Font(name=_FONT, size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            lv = glevels[ri][ci]
            for cc in range(col, col + span):
                ws.cell(row=r + ri, column=cc).border = _BORDER
                if lv:
                    ws.cell(row=r + ri, column=cc).fill = _lvl_fill(lv, cfg)
            col += span
        ws.row_dimensions[r + ri].height = 38 if ri else 18
    r += 3
    r = note(r, f"Insufficient data: {q.get('Insufficient Data', 0)} application profile(s).")
    r += 1

    # ---- top organisational issues
    r = band(r, "TOP ORGANISATIONAL ISSUES")
    data, levels = [], []
    for row_ in [x for x in issue_rows if x["Apps Affected"] > 0][:cfg["top_n_issues"]]:
        data.append([row_["Issue"], row_["Severity"],
                     _ctx(row_["Apps Affected"], row_["Total Apps"]),
                     row_["Category"], row_["Trend"]])
        levels.append(["", row_["_level"], row_["_level"], "", _trend_risk(row_["Trend"])])
    if not data:
        data = [["No failing health checks across the tenant", "", "", "", ""]]
        levels = [["GREEN", "", "", "", ""]]
    r = _write_table(ws, r, ["Issue", "Severity", "Apps Affected", "Category", "Trend"],
                     data, levels, cfg, spans=[4, 1, 3, 2, 2], left_cols=(0,))
    r = note(r, "Full list of all 31 health checks is on the 'Issue Heatmap' sheet.")
    r += 1

    # ---- tenant trend
    r = band(r, "TENANT TREND vs PREVIOUS REPORT")
    if tenant_trend:
        arrows = {"UP": "\u2191", "DOWN": "\u2193", "=": "\u2192", "+": "\u2192", "-": "\u2192"}
        data = [[t["Metric"], t["Previous"], t["Current"], t["Change"],
                 arrows.get(t["Direction"], t["Direction"])] for t in tenant_trend]
        levels = [["", "", "", t["_level"], t["_level"]] for t in tenant_trend]
        r = _write_table(ws, r, ["Metric", "Previous", "Current", "Change", "Direction"],
                         data, levels, cfg, spans=[4, 2, 2, 2, 2], left_cols=(0,))
        r = note(r, "Direction is semantic: fewer poor applications is green, more total flaws "
                    "is red. Arrows do not simply follow the sign of the change.")
    else:
        r = note(r, "No previous report supplied. Re-run with --previous-report to enable "
                    "trend analysis.")
    r += 1

    # ---- applications requiring attention
    r = band(r, "APPLICATIONS REQUIRING ATTENTION")
    r = note(r, "Ranked by the transparent Attention Score. This is a scan-health and findings "
                "prioritisation ranking, not a 'most vulnerable applications' list.")
    plist = build_priority_list(apps, cfg=cfg)
    data, levels = [], []
    for p in plist:
        data.append([p["Priority"], p["Application"], p["Attention Score"], p["Scan Health"],
                     p["Very High / High Flaws"], p["Scan Age (days)"], p["Top Issue"],
                     p["Trend"], p["Why"]])
        levels.append(["", "", p["_level"], _health_risk(p["Scan Health"]),
                       "", "", "", _trend_risk(p["Trend"]), ""])
    if not data:
        data, levels = [["", "No applications scored", "", "", "", "", "", "", ""]], None
    r = _write_table(ws, r,
                     ["#", "Application", "Score", "Health", "V.High / High",
                      "Scan age", "Top issue", "Trend", "Why this score"],
                     data, levels, cfg,
                     spans=[1, 2, 1, 1, 1, 1, 2, 1, 2], left_cols=(1, 6, 8),
                     row_height=60)
    r += 1
    r = note(r, "Drill down on the 'App Heatmap' and 'Issue Heatmap' sheets, both filterable "
                "and sortable. Raw evidence remains on the original detail sheets.")
    ws.freeze_panes = "A5"


def _kpi_row(ws, r: int, kpis: list[tuple[str, str, str, str]], cfg: dict, span: int) -> int:
    """Render KPI cards spread evenly across the full `span` of grid columns."""
    n = len(kpis)
    base, rem = divmod(span, n)
    col = 1
    for idx, (label, value, sub, lv) in enumerate(kpis):
        width = base + (1 if idx < rem else 0)
        end = col + width - 1
        for rr, text, font in (
            (r, label, Font(name=_FONT, size=9, bold=True, color="555555")),
            (r + 1, value, Font(name=_FONT, size=18, bold=True, color=cfg["colors"][lv])),
            (r + 2, sub, Font(name=_FONT, size=8, color="666666")),
        ):
            ws.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=end)
            c = ws.cell(row=rr, column=col, value=text)
            c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cc in range(col, end + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = _KPI_FILL
                cell.border = _BORDER
        col = end + 1
    ws.row_dimensions[r].height = 15
    ws.row_dimensions[r + 1].height = 26
    ws.row_dimensions[r + 2].height = 15
    return r + 4


_HEATMAP_WIDTHS = {
    "Application": 34, "Sandbox": 16, "Business Unit": 20, "Attention Score": 9,
    "Priority": 20, "Scan Health": 11, "Scan Age": 9, "Scan Age Bucket": 13,
    "Findings Risk": 13, "Very High Flaws": 11, "High Flaws": 9, "Medium Flaws": 11,
    "Low Flaws": 9, "Total Flaws": 10, "Open Policy Flaws": 12, "SCA": 10,
    "Weighted Flaws": 11, "Flaw Density": 11, "Density Basis": 18,
    "Veracode Score": 11, "Veracode Rating": 11, "Policy Status": 15, "Findings Basis": 26,
    "SCA Components": 12, "Health Issues": 11, "High-Severity Issues": 13,
    "Upload Size (MB)": 13, "Last Scan": 20, "Previous Health": 13, "Change": 9,
    "Trend": 11, "Quadrant": 15,
    "Top Issue": 44,
}


def _write_app_heatmap_sheet(wb: Workbook, apps: Sequence[ApplicationHealth], cfg: dict) -> None:
    ws = wb.create_sheet("App Heatmap")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    rows = build_heatmap_rows(apps, cfg)
    headers = [h for h in (rows[0].keys() if rows else []) if not h.startswith("_")]
    if not rows:
        ws["A1"] = "No applications."
        return

    title_row = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1,
                value="APPLICATION HEATMAP    |    "
                      "Scan health and security findings are separate dimensions. "
                      + DISCLAIMER)
    c.fill = _TITLE_FILL
    c.font = Font(name=_FONT, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    hr = title_row + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.fill = _BAND_FILL
        c.font = Font(name=_FONT, size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = _HEATMAP_WIDTHS.get(h, 13)
    ws.row_dimensions[hr].height = 30

    for ri, row in enumerate(rows, hr + 1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h))
            c.font = Font(name=_FONT, size=9)
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="left" if h in ("Application", "Top Issue", "Business Unit") else "center",
                vertical="center")
            lv = row.get(f"_lv_{h}")
            if lv:
                c.fill = _lvl_fill(lv, cfg)
                if lv in ("RED", "ORANGE"):
                    c.font = Font(name=_FONT, size=9, bold=True, color=cfg["colors"][lv])

    last = hr + len(rows)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = ws.cell(row=hr + 1, column=4)


def _write_issue_heatmap_sheet(wb: Workbook, issue_rows: list[dict],
                               m: TenantMetrics, cfg: dict) -> None:
    ws = wb.create_sheet("Issue Heatmap")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    headers = ["Check #", "Issue", "Category", "Severity", "Apps Affected",
               "% of Tenant", "Prevalence", "Occurrences", "Business Units Affected",
               "Previous Apps Affected", "Trend", "Top Recommendation"]
    widths = {"Check #": 8, "Issue": 46, "Category": 17, "Severity": 10,
              "Apps Affected": 13, "% of Tenant": 12, "Prevalence": 24,
              "Occurrences": 12, "Business Units Affected": 13,
              "Previous Apps Affected": 13, "Trend": 12, "Top Recommendation": 60}

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1,
                value=f"ORGANISATIONAL ISSUE HEATMAP    |    All 31 health checks across "
                      f"{m.total_apps:,} application profiles")
    c.fill = _TITLE_FILL
    c.font = Font(name=_FONT, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    hr = 2
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.fill = _BAND_FILL
        c.font = Font(name=_FONT, size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)
    ws.row_dimensions[hr].height = 30

    for ri, row in enumerate(issue_rows, hr + 1):
        bar = "#" * int(round(30 * min(row["% of Tenant"], 100) / 100.0))
        vals = [row["Check #"], row["Issue"], row["Category"], row["Severity"],
                _ctx(row["Apps Affected"], row["Total Apps"]),
                row["% of Tenant"] / 100.0, bar, row["Occurrences"],
                row["Business Units Affected"], row["Previous Apps Affected"],
                row["Trend"], row["Top Recommendation"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name=_FONT, size=9)
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="left" if headers[ci - 1] in ("Issue", "Top Recommendation", "Category")
                else "center",
                vertical="center", wrap_text=headers[ci - 1] == "Top Recommendation")
            if headers[ci - 1] == "% of Tenant":
                c.number_format = "0.0%"
            if headers[ci - 1] in ("Severity", "Apps Affected"):
                c.fill = _lvl_fill(row["_level"], cfg)
            if headers[ci - 1] == "Prevalence":
                c.font = Font(name="Consolas", size=9, color=cfg["colors"][row["_level"]])
            if headers[ci - 1] == "Trend":
                lv = _trend_risk(row["Trend"])
                if lv != "GRAY":
                    c.fill = _lvl_fill(lv, cfg)

    last = hr + len(issue_rows)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = ws.cell(row=hr + 1, column=3)


# ==========================================================================
# Top-level entry point used by script.py
# ==========================================================================

def generate_dashboard(health_rows: Sequence[dict],
                       issue_records: Sequence[dict],
                       prev_rows: dict | None = None,
                       cfg: dict | None = None,
                       check_categories: dict[int, str] | None = None,
                       prev_issue_counts: dict[int, int] | None = None) -> dict:
    """Normalize, score and aggregate. Returns everything the writers need."""
    cfg = cfg or DASHBOARD_CONFIG
    apps, issues = build_application_health(
        health_rows, issue_records, prev_rows, cfg, check_categories)
    metrics = aggregate_tenant(apps, cfg)
    issue_rows = build_issue_heatmap(issues, metrics.total_apps, cfg, prev_issue_counts)
    tenant_trend = build_tenant_trend(metrics, prev_rows, cfg)
    return {
        "apps": apps,
        "issues": issues,
        "metrics": metrics,
        "issue_rows": issue_rows,
        "tenant_trend": tenant_trend,
        "config": cfg,
    }


def write_dashboard_xlsx(path: str, bundle: dict) -> None:
    """Standalone dashboard workbook (used when the main report is CSV/JSON)."""
    wb = Workbook()
    wb.remove(wb.active)
    write_dashboard_sheets(wb, bundle["apps"], bundle["issues"], bundle["metrics"],
                           bundle["tenant_trend"], bundle["issue_rows"], bundle["config"])
    wb.save(path)
