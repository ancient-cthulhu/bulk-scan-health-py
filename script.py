#!/usr/bin/env python3
"""
Veracode Tenant-Wide Scan Health
Python port of https://github.com/veracode/scan_health (Go v2.47)
Extended for tenant-wide iteration, Excel/CSV/JSON export, trend analysis.

Requirements:
    pip install veracode-api-signing requests openpyxl
"""

from __future__ import annotations

import csv
import json
import re
import time
import logging
import argparse
import threading
import xml.etree.ElementTree as ET
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("scan_health")

# ==========================================================================
# Constants
# ==========================================================================
MAX_FILE_COUNT = 10000
MAX_MODULE_COUNT = 500
MAX_SELECTED_MODULE_COUNT = 100
MAX_FLAW_COUNT = 2500
MAX_TOTAL_MODULE_SIZE = 1_000_000_000
MAX_ANALYSIS_SIZE = 500_000_000
STALE_SCAN_DAYS = 30

REGIONS: dict[str, dict[str, str]] = {
    "commercial": {"base": "https://analysiscenter.veracode.com", "xml": "https://analysiscenter.veracode.com/api/5.0", "rest": "https://api.veracode.com/appsec/v1"},
    "eu":         {"base": "https://analysiscenter.veracode.eu",  "xml": "https://analysiscenter.veracode.eu/api/5.0",  "rest": "https://api.veracode.eu/appsec/v1"},
}

_NS = re.compile(r'\sxmlns="[^"]+"')
_DT_FMTS = ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S %Z", "%Y-%m-%d %H:%M:%S UTC")
_URL_RE = re.compile(r'https?://\S+')

# FancyList pattern lists
JUNK_FILE_PATTERNS = ["!LICENSE*",".*","*.asmx","*.config","*.cs","*.eot","*.gif","*.ico","*.jpeg","*.jpg","*.less","*.manifest","*.map","*.markdown","*.md","*.pdf","*.png","*.properties","*.scss","*.sh","*.svg","*.ttf","*.txt","*.woff","*.xml","AUTHORS","CHANGELOG","CONTRIBUTORS","Dockerfile","LICENSE","Makefile","README","Thumbs.db"]
THIRD_PARTY_PATTERNS = ["7z.dll","7-zip.dll","Google.*.dll","7za.exe","AutoMapper.dll","AutoMapper.*.dll","Azure.*.dll","BouncyCastle.*","Castle.Core.*","Castle.Windsor.*","componentspace.saml2.dll","Dapper.dll","Dapper.*.dll","devexpress.*","entityframework.*","Google.Protobuf.dll","gradle-wrapper.jar","GraphQL.*.dll","itextsharp.dll","log4net.dll","microsoft.*.dll","microsoft.*.pdb","!^_Microsoft.","!^_System.","!^_Azure.","newrelic.*.dll","newtonsoft.json.*","ninject.*.dll","org.eclipse.*.jar","Serilog.dll","syncfusion.*","system.*.dll","Telerik.*.dll","WebGrease.dll","phantomjs.exe","Moq.dll","ComponentSpace.SAML2.dll","^aspnet-codegenerator","sni.dll","AntiXssLibrary.dll","Antlr3.Runtime.dll","FluentValidation.dll"]
SENSITIVE_SECRET_PATTERNS = ["*.asc","*.crt","*.gpg","*.jks","*.key","*.p7b","*.p7s","*.pem","*.pfx","*.pgp","*.p12","*.tfvars","variable.tf",".htpasswd"]
SENSITIVE_BACKUP_PATTERNS = ["*.bac","*.back","*.backup","*.old","*.orig","*.bak"]
SENSITIVE_WORD_PATTERNS = ["*.docx","*.doc","*.docm","*.odt"]
SENSITIVE_SPREADSHEET_PATTERNS = ["*.xlsx","*.xls","*.xlsm","*.ods"]
SENSITIVE_JUPYTER_PATTERNS = ["*.ipynb"]
TEST_FILE_PATTERNS = ["nunit.framework.dll","Moq.dll","^.test.","!Test*","!*Test","!^Test.","!^Tests.","*.unittests.dll","*.unittest.dll","^mock","^unittest","^harness","*.feature","*.js.snap"]
UNWANTED_7Z = ["*.7z"]
UNWANTED_COFFEE = ["*.coffee"]
UNWANTED_SCRIPTS = ["*.sh","*.ps","*.ps1","*.bat"]
UNWANTED_INSTALLERS = ["setup.exe","*setup.exe","*.msi","installer.exe","*installer.exe","*.msix","*.appx","*.msixbundle",".appxbundle"]
UNWANTED_PYD = ["*.pyd"]
UNWANTED_PYC = ["*.pyc"]
UNWANTED_DEPLOY = ["*.deploy"]
UNWANTED_WIBU = ["WibuCmNET.dll"]
SRC_JAVA = ["*.java"]; SRC_CS = ["*.cs"]; SRC_SLN = ["*.sln"]; SRC_CSPROJ = ["*.csproj"]; SRC_C = ["*.c"]; SRC_CPP = ["*.cpp"]; SRC_SWIFT = ["*.swift"]
SCA_SUPPORTED = ["*.dll","*.exe","*.jar","*.apk","*.aab","*.war","*.ear","*.js","*.ts","*.php","*.lock","package-lock.json","npm-shrinkwrap.json","go.sum","vendor.json","*.deps.json","*.py"]
DOTNET_PRECOMPILE_PATTERNS = ["*.cshtml","*.ascx","*.aspx","*.asax"]
GO_WORKSPACE_PATTERNS = ["go.work","go.work.sum"]
EXCESS_MSFT_PATTERNS = ["csc.exe"]
LOOSE_CLASS_FILE = ["*.class"]; LOOSE_CLASS_MODULE = ["class files within*"]
REPO_CANARIES = ["fsmonitor-watchman.sample","FETCH_HEAD"]
GRADLE_WRAPPER = ["gradle-wrapper.jar"]; MINIFIED_JS = ["*.min.js"]

# Excel styling
_HF = PatternFill("solid", fgColor="1F4E79")
_HN = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_DF = Font(name="Arial", size=10)
_BF = Font(bold=True, name="Arial", size=10)
_CA = Alignment(horizontal="center", vertical="center")
_WA = Alignment(horizontal="left", vertical="top", wrap_text=True)
_TH = Side(style="thin", color="CCCCCC")
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_CLR = {"Good": "C6EFCE", "Fair": "FFEB9C", "Poor": "FFC7CE"}
_AGE_CLR = {"<7d": "C6EFCE", "7-30d": "C6EFCE", "30-90d": "FFEB9C", "90d+": "FFC7CE", "N/A": "FFFFFF"}

# ==========================================================================
# Dataclasses
# ==========================================================================

@dataclass
class Issue:
    severity: str
    description: str
    check_num: int = 0
    check_name: str = ""
    check_num: int = 0
    check_name: str = ""

CHECK_CATEGORIES: dict[int, str] = {
    1: "Packaging", 2: "Module Selection", 3: "Flaw Analysis",
    4: "Fatal Errors", 5: "Fatal Errors", 6: "Packaging",
    7: "Packaging", 8: "Packaging", 9: "SCA",
    10: "Module Selection", 11: "Packaging", 12: "Module Quality",
    13: "Module Quality", 14: "Fatal Errors", 15: "Module Selection",
    16: "Security Risk", 17: "Security Risk", 18: "Packaging",
    19: "Packaging", 20: "Packaging", 21: "Packaging",
    22: "Packaging", 23: "Packaging", 24: "Module Selection",
    25: "Module Selection", 26: "Module Selection", 27: "Packaging",
    28: "Packaging", 29: "Module Selection", 30: "Scan Recency",
    31: "Packaging",
}

@dataclass
class FlawSummary:
    total: int = 0; fixed: int = 0; pol_aff: int = 0
    mitigated: int = 0; open_pol: int = 0; open_nopol: int = 0
    # Open (not fixed, not mitigated) flaws by Veracode severity level.
    # 5 = Very High, 4 = High, 3 = Medium, 2/1/0 rolled up as Low.
    # sev_available is False when the severity breakdown could not be derived,
    # so the dashboard can render "unknown" rather than implying zero.
    open_vhigh: int = 0; open_high: int = 0
    open_med: int = 0; open_low: int = 0
    sev_available: bool = False

@dataclass
class ScanResult:
    app_name: str = ""; bu: str = ""; policy: str = ""; app_id: int = 0
    sandbox: str = ""; build_id: str = ""; scan_name: str = ""
    is_latest: bool = True; scan_status: str = "No Scan"
    published: str = ""; days_since: object = "N/A"; duration: str = ""
    engine: str = ""; analysis_size_mb: float = 0.0
    files_uploaded: int = 0; total_modules: int = 0
    selected_modules: int = 0; fatal_errors: int = 0
    flaws: FlawSummary = field(default_factory=FlawSummary)
    health: str = "Good"; high_issues: int = 0; medium_issues: int = 0
    total_issues: int = 0; issues_text: str = "None"; recs_text: str = "None"
    review_url: str = ""; triage_url: str = ""
    selected_names: str = ""; sca_count: int = 0; age_bucket: str = "N/A"
    total_upload_mb: float = 0.0; health_trend: str = ""

    def to_row(self) -> dict:
        return {
            "App Name": self.app_name, "Business Unit": self.bu, "Policy": self.policy,
            "App ID": self.app_id, "Sandbox": self.sandbox,
            "Build ID": self.build_id, "Scan Name": self.scan_name,
            "Is Latest": self.is_latest, "Scan Status": self.scan_status,
            "Published": self.published, "Days Since Scan": self.days_since,
            "Duration": self.duration, "Engine": self.engine,
            "Analysis Size (MB)": self.analysis_size_mb,
            "Total Upload Size (MB)": self.total_upload_mb,
            "Files Uploaded": self.files_uploaded,
            "Total Modules": self.total_modules,
            "Selected Modules": self.selected_modules,
            "Selected Module Names": self.selected_names,
            "Fatal Errors": self.fatal_errors,
            "Total Flaws": self.flaws.total,
            "Open Affecting Policy": self.flaws.open_pol,
            "Mitigated": self.flaws.mitigated, "Fixed": self.flaws.fixed,
            "Policy Affecting": self.flaws.pol_aff,
            "Open Very High Flaws": self.flaws.open_vhigh,
            "Open High Flaws": self.flaws.open_high,
            "Open Medium Flaws": self.flaws.open_med,
            "Open Low Flaws": self.flaws.open_low,
            "Flaw Severity Data": "Available" if self.flaws.sev_available else "Unavailable",
            "SCA Components": self.sca_count,
            "Scan Age Bucket": self.age_bucket,
            "Health": self.health, "Health Trend": self.health_trend,
            "High Issues": self.high_issues, "Medium Issues": self.medium_issues,
            "Total Issues": self.total_issues,
            "Issues": self.issues_text, "Recommendations": self.recs_text,
            "Review Modules URL": self.review_url, "Triage Flaws URL": self.triage_url,
        }

@dataclass
class ModuleRow:
    app_name: str = ""; build_id: str = ""; name: str = ""
    status: str = ""; selected: bool = False; dependency: bool = False
    fatal: bool = False; third_party: bool = False; ignored: bool = False
    platform: str = ""; compiler: str = ""; size: str = ""
    issues: str = ""
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Build ID": self.build_id, "Module": self.name,
                "Status": self.status, "Selected": self.selected, "Dependency": self.dependency,
                "Fatal": self.fatal, "3rd Party": self.third_party, "Ignored": self.ignored,
                "Platform": self.platform, "Compiler": self.compiler, "Size": self.size,
                "Issues": self.issues}

@dataclass
class FileRow:
    app_name: str = ""; build_id: str = ""; name: str = ""
    status: str = ""; md5: str = ""; ignored: bool = False; third_party: bool = False
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Build ID": self.build_id, "File": self.name,
                "Status": self.status, "MD5": self.md5, "Ignored": self.ignored, "3rd Party": self.third_party}

@dataclass
class RecommendationRow:
    app_name: str = ""; build_id: str = ""; severity: str = ""
    recommendation: str = ""; doc_url: str = ""
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Build ID": self.build_id,
                "Severity": self.severity, "Recommendation": self.recommendation,
                "Doc URL": self.doc_url}

@dataclass
class TrendRow:
    app_name: str = ""; sandbox: str = ""
    prev_health: str = ""; curr_health: str = ""; change: str = ""
    prev_flaws: int = 0; curr_flaws: int = 0; flaw_delta: int = 0
    prev_open_pol: int = 0; curr_open_pol: int = 0; open_pol_delta: int = 0
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Sandbox": self.sandbox,
                "Previous Health": self.prev_health, "Current Health": self.curr_health,
                "Health Change": self.change,
                "Previous Total Flaws": self.prev_flaws, "Current Total Flaws": self.curr_flaws,
                "Flaw Delta": self.flaw_delta,
                "Previous Open Policy": self.prev_open_pol, "Current Open Policy": self.curr_open_pol,
                "Open Policy Delta": self.open_pol_delta}

@dataclass
class AppIssues:
    """Carries structured issue data per app for aggregation."""
    app_name: str = ""
    bu: str = ""
    policy: str = ""
    sandbox: str = ""
    health: str = ""
    issues: list[Issue] = field(default_factory=list)
    recs: list[str] = field(default_factory=list)

@dataclass
class AggIssue:
    """Structured issue record for accurate tenant-level aggregation."""
    app_name: str
    bu: str
    sandbox: str
    check_num: int
    check_name: str
    category: str
    severity: str
    description: str
    recommendation: str

# ==========================================================================
# Helpers
# ==========================================================================

def _si(v: object, d: int = 0) -> int:
    try: return int(v)
    except (ValueError, TypeError): return d

def _parse_dt(s: str) -> datetime | None:
    if not s: return None
    s = s.strip()
    for f in _DT_FMTS:
        try:
            dt = datetime.strptime(s, f)
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except ValueError: continue
    return None

def _days_since(s: str) -> int | None:
    dt = _parse_dt(s)
    return (datetime.now(timezone.utc) - dt).days if dt else None

def _dur(sub: str, pub: str) -> str:
    a, b = _parse_dt(sub), _parse_dt(pub)
    if a and b:
        s = int((b - a).total_seconds())
        if s >= 0:
            h, r = divmod(s, 3600); m, sec = divmod(r, 60)
            if h: return f"{h}h {m}m {sec}s"
            return f"{m}m {sec}s" if m else f"{sec}s"
    return ""

def _age_bucket(days: int | None) -> str:
    if days is None: return "N/A"
    if days < 7: return "<7d"
    if days <= 30: return "7-30d"
    if days <= 90: return "30-90d"
    return "90d+"

def _parse_module_size(size_str: str) -> int:
    """Parse prescan module size string like '5MB', '120KB', '2GB' to bytes."""
    s = size_str.strip().upper()
    for suffix, mult in [("GB", 1_000_000_000), ("MB", 1_000_000), ("KB", 1_000)]:
        if s.endswith(suffix):
            try: return int(s[:-len(suffix)]) * mult
            except ValueError: return 0
    return 0

def _extract_url(text: str) -> str:
    m = _URL_RE.search(text)
    return m.group(0).rstrip(".") if m else ""

def _is_java(n: str) -> bool: return n.lower().endswith((".jar",".war",".ear"))
def _is_dotnet(n: str) -> bool: return n.lower().endswith((".dll",".exe"))
def _is_js_module(n: str) -> bool:
    lo = n.lower()
    return lo.startswith("js files within") or lo.startswith("js files extracted from")
def _is_node_module(n: str) -> bool: return "_nodemodule_" in n.lower()
def _has_status(m: dict, s: str) -> bool: return s.lower() in m.get("status","").lower()

def _top5(items: list[str]) -> str:
    if len(items) <= 5: return ", ".join(f'"{i}"' for i in items)
    return ", ".join(f'"{i}"' for i in items[:5]) + f" and {len(items)-5} others"

# ==========================================================================
# FancyList DSL (audited against Go utils/fancy_list.go)
# ==========================================================================

def _fancy_match(filename: str, patterns: list[str]) -> bool:
    """Test a filename against a list of patterns using the FancyList DSL.

    Pattern syntax:
      - Default: case-insensitive equality match (e.g. "readme")
      - Single *: wildcard. "*.exe" matches suffix, "abc.*" matches prefix,
                  "abc.*.xyz" matches prefix+suffix. Only one * allowed.
      - ^ prefix: contains search (e.g. "^test" matches if "test" in filename)
      - ! prefix: force case-sensitive matching for the rest of the pattern.
                  Can combine with * or ^ (e.g. "!^_Microsoft." is case-sensitive
                  contains, "!Test*" is case-sensitive prefix).
    """
    fn = filename.strip()
    for pat in patterns:
        p = pat
        # ! means case-sensitive; otherwise fold to lowercase
        case_sensitive = p.count("!") == 1
        if case_sensitive:
            p = p.replace("!", "")
            f = fn
        else:
            f = fn.lower()
            p = p.lower()
        # ^ means contains
        if "^" in p and p.count("^") == 1:
            p = p.replace("^", "")
            if p in f:
                return True
            continue
        # * wildcard (exactly one allowed)
        if p.count("*") == 1:
            if p.startswith("*"):
                if f.endswith(p[1:]):
                    return True
            elif p.endswith("*"):
                if f.startswith(p[:-1]):
                    return True
            else:
                parts = p.split("*", 1)
                if f.startswith(parts[0]) and f.endswith(parts[1]) and len(f) >= len(parts[0]) + len(parts[1]):
                    return True
        elif f == p:
            return True
    return False

def _fancy_match_files(files: list[dict], patterns: list[str]) -> list[str]:
    found: list[str] = []
    for f in files:
        if f.get("is_ignored") or f.get("is_third_party"): continue
        if _fancy_match(f["name"], patterns) and f["name"] not in found:
            found.append(f["name"])
    return found

def _fancy_match_modules(modules: list[dict], patterns: list[str], selected_only: bool = False) -> list[str]:
    found: list[str] = []
    for m in modules:
        if selected_only and not m.get("is_selected"): continue
        if _fancy_match(m["name"], patterns) and m["name"] not in found:
            found.append(m["name"])
    return found

# ==========================================================================
# API Client
# ==========================================================================

class AuthError(Exception):
    """Raised on 401/403 to signal credential issues."""

class VeracodeClient:
    def __init__(self, region: str = "commercial", timeout: int = 120) -> None:
        self._cfg = REGIONS[region]
        self._timeout = timeout
        self._s = requests.Session()
        self._s.auth = RequestsAuthPluginVeracodeHMAC()
        self._s.headers["User-Agent"] = "veracode-scan-health-py/3.0"
        retry = Retry(total=3, backoff_factor=1.0,
                      status_forcelist=(429, 500, 502, 503, 504),
                      allowed_methods=("GET",))
        self._s.mount("https://", HTTPAdapter(max_retries=retry))

    def close(self) -> None: self._s.close()
    def __enter__(self) -> "VeracodeClient": return self
    def __exit__(self, *a: object) -> None: self.close()

    def _check_auth(self, resp: requests.Response) -> None:
        if resp.status_code in (401, 403):
            body = resp.text[:200] if resp.text else ""
            raise AuthError(
                f"HTTP {resp.status_code} from {resp.url}. "
                f"API credentials may be expired or lack required permissions. "
                f"Response: {body}")

    def _xml(self, ep: str, params: dict | None = None) -> ET.Element:
        r = self._s.get(f"{self._cfg['xml']}/{ep}", params=params, timeout=self._timeout)
        self._check_auth(r)
        r.raise_for_status()
        # Strip ALL xmlns attributes (count=0) for clean XPath
        return ET.fromstring(_NS.sub("", r.text, count=0))

    def _rest(self, path: str, params: dict | None = None) -> dict:
        r = self._s.get(f"{self._cfg['rest']}{path}", params=params, timeout=self._timeout)
        self._check_auth(r)
        r.raise_for_status()
        return r.json()

    @property
    def base(self) -> str: return self._cfg["base"]

    def get_apps(self) -> list[dict]:
        apps: list[dict] = []; page = 0
        while True:
            d = self._rest("/applications", {"page": page, "size": 500})
            emb = d.get("_embedded", {}).get("applications", [])
            if not emb: break
            for a in emb:
                p = a.get("profile", {}); pols = p.get("policies") or []
                apps.append({"app_id": a.get("guid",""), "legacy_id": a.get("id"),
                    "name": p.get("name",""),
                    "bu": (p.get("business_unit") or {}).get("name",""),
                    "policy": pols[0].get("name","") if pols else ""})
            page += 1
            if page >= (d.get("page") or {}).get("total_pages", 1): break
        return apps

    def get_builds(self, aid: int, sbx: str | None = None) -> list[dict]:
        p: dict = {"app_id": aid}
        if sbx: p["sandbox_id"] = sbx
        try: root = self._xml("getbuildlist.do", p)
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getbuildlist.do failed for %s: %s", aid, e); return []
        return [{"id": b.get("build_id"), "ver": b.get("version","")} for b in root.findall(".//build")]

    def get_sandboxes(self, aid: int) -> list[dict]:
        try: root = self._xml("getsandboxlist.do", {"app_id": aid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getsandboxlist.do failed: %s", e); return []
        return [{"id": s.get("sandbox_id",""), "name": s.get("sandbox_name","")} for s in root.findall(".//sandbox")]

    def get_build_info(self, aid: int, bid: str) -> dict | None:
        """Get build status to determine if scan is published."""
        try: root = self._xml("getbuildinfo.do", {"app_id": aid, "build_id": bid})
        except (requests.HTTPError, ET.ParseError): return None
        au = root.find(".//analysis_unit")
        if au is None: return None
        return {"status": au.get("status",""), "published": au.get("published_date","")}

    def get_detailed_report(self, bid: str) -> dict | None:
        try: root = self._xml("detailedreport.do", {"build_id": bid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("detailedreport.do failed for %s: %s", bid, e); return None
        sa = root.find(".//static-analysis")
        if sa is None: return None
        fl = FlawSummary()
        for f in root.findall(".//severity/category/cwe/staticflaws/flaw"):
            fl.total += 1
            apc = f.get("affects_policy_compliance","false") == "true"
            is_fixed = f.get("remediation_status","") == "Fixed"
            is_miti = f.get("mitigation_status","none") not in ("none","rejected")
            if apc: fl.pol_aff += 1
            if is_fixed: fl.fixed += 1
            elif is_miti: fl.mitigated += 1
            elif apc: fl.open_pol += 1
            else: fl.open_nopol += 1
            # Severity breakdown of OPEN flaws only. Same XML, no extra API call.
            if not is_fixed and not is_miti:
                sev = _si(f.get("severity"), -1)
                if sev == 5: fl.open_vhigh += 1
                elif sev == 4: fl.open_high += 1
                elif sev == 3: fl.open_med += 1
                elif sev >= 0: fl.open_low += 1
        # Only claim the breakdown is usable if we actually walked flaw elements.
        fl.sev_available = fl.total > 0
        if fl.total == 0:
            fl.total = _si(root.get("total_flaws","0"))
            fl.open_pol = _si(root.get("flaws_not_mitigated","0"))
            fl.mitigated = fl.total - fl.open_pol
        dr_mods = [{"name": unescape(m.get("name","")), "compiler": unescape(m.get("compiler","")),
            "os": unescape(m.get("os","")), "arch": unescape(m.get("architecture",""))}
            for m in (sa.findall(".//module") or [])]
        sca_node = root.find(".//software_composition_analysis")
        sca_on = sca_node is not None and sca_node.get("sca_service_available","true").lower() != "false"
        sca_comps = [unescape(c.get("file_name","")) for c in root.findall(".//vulnerable_components/component")]
        return {
            "account_id": root.get("account_id",""), "app_id": root.get("app_id",""),
            "sandbox_id": _si(root.get("sandbox_id","0")), "sandbox_name": root.get("sandbox_name",""),
            "analysis_id": root.get("analysis_id",""), "sau_id": root.get("static_analysis_unit_id",""),
            "bu": unescape(root.get("business_unit","")), "app_name": unescape(root.get("app_name","")),
            "scan_name": unescape(sa.get("version","")), "engine": sa.get("engine_version",""),
            "submitted": sa.get("submitted_date",""), "published": sa.get("published_date",""),
            "analysis_size": _si(sa.get("analysis_size_bytes","0")),
            "is_latest": root.get("is_latest_build","true").lower() == "true",
            "flaws": fl, "dr_modules": dr_mods,
            "sca_on": sca_on, "sca_comps": sca_comps,
        }

    def get_files(self, aid: int, bid: str) -> list[dict]:
        try: root = self._xml("getfilelist.do", {"app_id": aid, "build_id": bid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getfilelist.do failed: %s", e); return []
        return [{"name": unescape(f.get("file_name","")), "status": unescape(f.get("file_status","")),
                 "md5": f.get("file_md5",""), "is_ignored": False, "is_third_party": False}
                for f in root.findall(".//file")]

    def get_prescan(self, aid: int, bid: str) -> list[dict]:
        try: root = self._xml("getprescanresults.do", {"app_id": aid, "build_id": bid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getprescanresults.do failed: %s", e); return []
        mods: list[dict] = []
        for m in root.findall(".//module"):
            issues: list[str] = []
            for iss in m.findall(".//issue"):
                d = unescape(iss.get("details",""))
                if d and d not in issues: issues.append(d)
            st = unescape(m.get("status",""))
            if st != "OK":
                for part in st.split(","):
                    p = part.strip()
                    if p and p not in issues: issues.append(p)
            mods.append({"name": unescape(m.get("name","")), "status": st,
                "platform": unescape(m.get("platform","")),
                "size": unescape(m.get("size","")), "md5": m.get("checksum",""),
                "has_fatal": m.get("has_fatal_errors","false").lower() == "true",
                "is_dep": m.get("is_dependency","false").lower() == "true",
                "issues": issues, "is_selected": False, "is_ignored": False, "is_third_party": False})
        return mods

    def get_app_info(self, aid: int) -> dict | None:
        try: root = self._xml("getappinfo.do", {"app_id": aid})
        except (requests.HTTPError, ET.ParseError): return None
        a = root.find(".//application")
        return {"modified": a.get("modified_date","")} if a is not None else None


# ==========================================================================
# Module merging
# ==========================================================================

def _merge_modules(dr_modules: list[dict], prescan_modules: list[dict]) -> list[dict]:
    by_name: dict[str, dict] = {}
    for m in prescan_modules:
        n = m["name"]
        if n not in by_name:
            by_name[n] = {**m, "was_scanned": False, "compiler": "", "os": "", "arch": ""}
        else:
            ex = by_name[n]
            if m["has_fatal"]: ex["has_fatal"] = True
            for iss in m["issues"]:
                if iss not in ex["issues"]: ex["issues"].append(iss)
    for dm in dr_modules:
        n = dm["name"]
        if n in by_name:
            by_name[n]["is_selected"] = True; by_name[n]["was_scanned"] = True
            by_name[n]["compiler"] = dm.get("compiler","")
            by_name[n]["os"] = dm.get("os",""); by_name[n]["arch"] = dm.get("arch","")
        else:
            by_name[n] = {"name": n, "status": "OK", "platform": "", "size": "",
                "md5": "", "has_fatal": False, "is_dep": False, "issues": [],
                "is_selected": True, "was_scanned": True, "is_ignored": False,
                "is_third_party": False, "compiler": dm.get("compiler",""),
                "os": dm.get("os",""), "arch": dm.get("arch","")}
    return list(by_name.values())


# ==========================================================================
# Individual check functions
# ==========================================================================

CheckFunc = Callable[[list[dict], list[dict], FlawSummary, dict, bool, list[str], str], tuple[list[Issue], list[str]]]

def _chk(sev: str, msg: str) -> Issue: return Issue(severity=sev, description=msg)

def check_01_junk_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; ign: list[str] = []
    for idx, f in enumerate(files):
        nm = f["name"]
        if nm.lower().endswith(".pdb") or nm.lower() in (".gitignore","head"):
            files[idx]["is_ignored"] = True; continue
        if _fancy_match(nm, JUNK_FILE_PATTERNS):
            files[idx]["is_ignored"] = True; ign.append(nm)
    for idx, m in enumerate(modules):
        if _fancy_match(m["name"], JUNK_FILE_PATTERNS): modules[idx]["is_ignored"] = True
    if ign:
        msg = f'An unnecessary file was uploaded: "{ign[0]}".' if len(ign)==1 else f'{len(ign)} unnecessary files were uploaded: {_top5(ign)}.'
        issues.append(_chk("medium", msg))
        recs.append("Follow the packaging instructions or use the Veracode auto-packager (https://docs.veracode.com/r/About_auto_packaging) to keep the upload as small as possible to improve upload and scan times.")
    return issues, recs

def check_02_third_party(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, sca_comps: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; sel_tp: list[str] = []
    for idx, f in enumerate(files):
        if _fancy_match(f["name"], THIRD_PARTY_PATTERNS): files[idx]["is_third_party"] = True
    for idx, f in enumerate(files):
        if not f["is_third_party"] and f["name"] in sca_comps: files[idx]["is_third_party"] = True
    for idx, m in enumerate(modules):
        if _fancy_match(m["name"], THIRD_PARTY_PATTERNS):
            modules[idx]["is_third_party"] = True
            if m.get("is_selected"): sel_tp.append(m["name"])
        if not m["is_third_party"] and m["name"] in sca_comps:
            modules[idx]["is_third_party"] = True
    if sel_tp:
        msg = f'A third-party component was selected as an entry point: "{sel_tp[0]}".' if len(sel_tp)==1 else f'{len(sel_tp)} third-party components selected as entry points: {_top5(sel_tp)}.'
        issues.append(_chk("medium", msg)); recs.append("Only select first party components as the entry points for the analysis.")
    return issues, recs

def check_03_flaw_count(files: list[dict], modules: list[dict], flaws: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if flaws.total == 0:
        issues.append(_chk("medium", "No flaws were found in this scan. This is usually due to scan misconfiguration."))
        recs.append("When no flaws have been found this can be an indication that incorrect modules were selected, or the main application was not selected for analysis.")
    elif flaws.total > MAX_FLAW_COUNT:
        issues.append(_chk("medium", "A large number of flaws were reported in this scan."))
        recs.append(f"More than {MAX_FLAW_COUNT} flaws were found which can be an indication that the scan could be misconfigured.")
    return issues, recs

def check_04_fatal_errors(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    # 4a. missing primary debug symbols (.NET)
    fp = [m["name"] for m in modules if m["has_fatal"] and _is_dotnet(m["name"]) and not m["is_ignored"] and not m["is_third_party"] and _has_status(m, "Primary Files Compiled without Debug Symbols")]
    if fp:
        msg = f'{len(fp)} module(s) could not be scanned due to missing debug symbols (PDB): {_top5(fp)}.' if len(fp)>1 else f'Module could not be scanned due to missing debug symbols (PDB): "{fp[0]}".'
        issues.append(_chk("high", msg)); recs.append("Include PDB files for as many components as possible, especially first and second party components.")
    # 4b. no scannable Java binaries
    fj = [m["name"] for m in modules if m["has_fatal"] and _is_java(m["name"]) and _has_status(m, "No Scannable Binaries")]
    if fj:
        msg = f'{len(fj)} Java module(s) contained no compiled Java classes: {_top5(fj)}.' if len(fj)>1 else f'Java module contained no compiled classes: "{fj[0]}".'
        issues.append(_chk("high", msg)); recs.append("Veracode requires Java apps compiled into JAR, WAR or EAR.")
    # 4c. nested JARs
    fn = [m["name"] for m in modules if m["has_fatal"] and _is_java(m["name"]) and _has_status(m, "does not support jar files nested inside")]
    if fn:
        msg = f'{len(fn)} Java module(s) contained nested/shaded JARs: {_top5(fn)}.' if len(fn)>1 else f'Java module contained nested JARs: "{fn[0]}".'
        issues.append(_chk("high", msg)); recs.append("Veracode does not support nested JARs except for Spring Boot.")
    return issues, recs

def check_05_unscannable_java(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    uj = [m["name"] for m in modules if _is_java(m["name"]) and m["has_fatal"]
          and not _has_status(m, "No Scannable Binaries") and not _has_status(m, "does not support jar files nested inside")]
    if uj:
        issues.append(_chk("high", f'{len(uj)} Java module(s) not scannable: {_top5(uj)}.'))
        recs.append("Veracode requires Java apps compiled into JAR, WAR or EAR.")
        recs.append("The Veracode CLI can be used to package Java apps: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_06_unwanted_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    for pats, ftype, rr in [
        (UNWANTED_7Z,"7-zip file",["Veracode does not support 7-zip. Consider zip files instead."]),
        (UNWANTED_COFFEE,"CoffeeScript file",["CoffeeScript is not supported.","Review JS/TS packaging: https://docs.veracode.com/r/compilation_jscript."]),
        (UNWANTED_SCRIPTS,"batch/shell script",["Do not upload batch/shell scripts."]),
        (UNWANTED_INSTALLERS,"installer",["Do not upload installers or setup programs."]),
        (UNWANTED_PYD,"Python-compiled DLL",["Do not upload .pyd files."]),
        (UNWANTED_PYC,"compiled Python file",["Veracode requires Python source code. Do not upload compiled .pyc."]),
        (UNWANTED_DEPLOY,'ClickOnce ".deploy" file',["Veracode does not support ClickOnce deployments."]),
        (UNWANTED_WIBU,"CodeMeter obfuscation file",["Do not use code obfuscation tools other than Dotfuscator Community Edition."]),
    ]:
        ff = _fancy_match_files(files, pats)
        if ff:
            msg = f'{len(ff)} {ftype}(s) uploaded: {_top5(ff)}.' if len(ff)>1 else f'A {ftype} was uploaded: "{ff[0]}".'
            issues.append(_chk("medium", msg))
            for rec in rr: recs.append(rec)
            recs.append("Follow packaging instructions: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_07_nested_archives(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    na = list(dict.fromkeys([f["name"] for f in files if f["status"] == "Archive File Within Another Archive" and not f.get("is_ignored")]))
    if na:
        issues.append(_chk("high", f'{len(na)} nested archive(s) uploaded: {_top5(na)}. Veracode does not process nested archives.'))
        recs.append("Do not upload nested archives.")
    return issues, recs

def check_08_missing_precompiled(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    pcf = _fancy_match_files(files, DOTNET_PRECOMPILE_PATTERNS)
    if pcf:
        issues.append(_chk("high", f'{len(pcf)} .NET view/template file(s) uploaded: {_top5(pcf)}. Precompile ASP.NET views.'))
        recs.append("Precompile ASP.NET projects and upload all generated assemblies.")
    pc_mods = [m["name"] for m in modules if m.get("is_selected") and _is_dotnet(m["name"]) and not m["is_ignored"] and not m["is_third_party"] and any("No precompiled files were found" in iss for iss in m.get("issues",[]))]
    if pc_mods:
        issues.append(_chk("medium", f'{len(pc_mods)} .NET component(s) missing precompiled files: {_top5(pc_mods)}.'))
    return issues, recs

def check_09_missing_sca(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, sca_on: bool, sca_comps: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if sca_on and not sca_comps:
        if any(_fancy_match(f["name"], SCA_SUPPORTED) for f in files):
            issues.append(_chk("medium", "No SCA results for this scan. Possible misconfiguration."))
            recs.append("Follow packaging guidance: https://docs.veracode.com/r/compilation_packaging.")
    return issues, recs

def check_10_unselected_js(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    uj = list(dict.fromkeys([m["name"] for m in modules if _is_js_module(m["name"]) and not _is_node_module(m["name"])
              and "extracted from .map file" not in m["name"].lower()
              and not m["has_fatal"] and not m["is_ignored"] and not m.get("is_selected") and not m["is_third_party"]]))
    if uj:
        issues.append(_chk("medium", f'{len(uj)} JS module(s) not selected: {_top5(uj)}.'))
        recs.append('Select "JS files within ..." modules for JavaScript coverage.')
        recs.append("Under-selection of first party modules affects results quality: https://community.veracode.com/s/article/What-are-Modules-and-how-do-my-results-change-based-on-what-I-select.")
    return issues, recs

def check_11_unexpected_source(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    for pats, ft, rr in [
        (SRC_JAVA,"Java source",["Do not upload Java source. Compile into JAR/WAR/EAR: https://docs.veracode.com/r/compilation_java."]),
        (SRC_CS,"C# source",["Do not upload C# source. Compile with debug symbols: https://docs.veracode.com/r/compilation_net."]),
        (SRC_SLN,".NET solution file",["Do not upload .sln files."]),
        (SRC_CSPROJ,"C# project file",["Do not upload .csproj files."]),
        (SRC_C,"C source",["Do not upload C source. Compile with debug symbols."]),
        (SRC_CPP,"C++ source",["Do not upload C++ source. Compile with debug symbols."]),
        (SRC_SWIFT,"Swift source",["Do not upload Swift source. Compile per iOS packaging guidelines."]),
    ]:
        ff = _fancy_match_files(files, pats)
        if ff:
            issues.append(_chk("high", f'{len(ff)} {ft} file(s) uploaded: {_top5(ff)}. Veracode requires compiled binaries.'))
            for rec in rr: recs.append(rec)
    return issues, recs

def check_12_missing_supporting(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; ms_mods: list[str] = []; ms_count = 0
    for m in modules:
        if not m.get("is_selected") or m["is_ignored"]: continue
        for iss in m.get("issues",[]):
            if iss.startswith("Missing Supporting Files"):
                parts = iss.split(" ")
                if len(parts) > 4:
                    try: ms_count += int(parts[4]); ms_mods.append(m["name"]) if m["name"] not in ms_mods else None
                    except ValueError: pass
    if ms_count:
        issues.append(_chk("medium", f'{len(ms_mods)} module(s) missing {ms_count} supporting file(s): {_top5(ms_mods)}.'))
        recs.append("Resolve missing supporting files on the Review Modules page.")
        recs.append("Ensure all components are present for analysis.")
    return issues, recs

def check_13_missing_debug(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    md = list(dict.fromkeys([m["name"] for m in modules if _is_dotnet(m["name"]) and not m["is_ignored"] and not m["is_third_party"]
          and any("No supporting files or PDB files" in iss for iss in m.get("issues",[]))]))
    if md:
        issues.append(_chk("medium", f'{len(md)} module(s) lack debug symbols (PDB): {_top5(md)}.'))
        recs.append("Include PDB files for first and second party components.")
    return issues, recs

def check_14_unsupported_platform(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    upc = list(dict.fromkeys([m["name"] for m in modules if m["has_fatal"] and not m["is_ignored"] and not m["is_third_party"]
           and (_has_status(m,"(Fatal)Unsupported Platform") or _has_status(m,"(Fatal)Unsupported Compiler"))]))
    if upc:
        issues.append(_chk("high", f'{len(upc)} module(s) have unsupported platform/compiler: {_top5(upc)}.'))
        recs.append("Review packaging docs to ensure compiler is supported.")
    return issues, recs

def check_15_gradle_wrapper(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    gw = _fancy_match_modules(modules, GRADLE_WRAPPER, selected_only=True)
    if gw:
        issues.append(_chk("high", '"gradle-wrapper.jar" selected for analysis. This is a build tool, not the application.'))
        recs.append('Do not upload or select "gradle-wrapper.jar".')
        recs.append("Use Veracode CLI: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_16_sensitive_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    for pats, desc, sr in [
        (SENSITIVE_SECRET_PATTERNS, "potentially sensitive/secret file", ["Do not upload secrets, certificates, or keys."]),
        (SENSITIVE_BACKUP_PATTERNS, "backup/old file", ["Do not upload backup files."]),
        (SENSITIVE_WORD_PATTERNS, "Word document", ["Office documents could contain sensitive information."]),
        (SENSITIVE_SPREADSHEET_PATTERNS, "spreadsheet", ["Office documents could contain sensitive information."]),
        (SENSITIVE_JUPYTER_PATTERNS, "Jupyter notebook", ["Jupyter notebooks could contain sensitive data."]),
    ]:
        ff = _fancy_match_files(files, pats)
        if ff:
            issues.append(_chk("high", f'{len(ff)} {desc}(s) uploaded: {_top5(ff)}.'))
            for rec in sr: recs.append(rec)
            recs.append("Do not upload unnecessary files.")
    return issues, recs

def check_17_repositories(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    repo_f = [f["name"] for f in files if f["name"].lower() in ("fsmonitor-watchman.sample","fetch_head")]
    if repo_f:
        issues.append(_chk("medium", "A git repository was uploaded. Repositories can contain sensitive information."))
        recs.append("Do not upload source code repositories.")
    return issues, recs

def check_18_node_modules(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; nm_found: list[str] = []
    for idx, m in enumerate(modules):
        if _is_node_module(m["name"]):
            if m["name"] not in nm_found: nm_found.append(m["name"])
            modules[idx]["is_third_party"] = True; modules[idx]["is_ignored"] = True
    if nm_found:
        issues.append(_chk("medium", f'{len(nm_found)} "node_modules" folder(s) uploaded. This increases upload size and module count.'))
        recs.append('Do not upload "node_modules" folders.')
        recs.append("Use Veracode CLI: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_19_testing_artefacts(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    sel_test = _fancy_match_modules(modules, TEST_FILE_PATTERNS, selected_only=True)
    if sel_test:
        issues.append(_chk("high", f'{len(sel_test)} testing artefact(s) selected: {_top5(sel_test)}.'))
        recs.append("Do not upload testing artifacts. Do not select them as entry points.")
    upl_test = _fancy_match_files(files, TEST_FILE_PATTERNS)
    if upl_test:
        issues.append(_chk("medium", f'{len(upl_test)} testing artefact(s) uploaded: {_top5(upl_test)}.'))
        recs.append("Do not upload testing artifacts.")
    test_mods = list(dict.fromkeys([m["name"] for m in modules for iss in m.get("issues",[]) if "test/" in iss.lower()]))
    if test_mods:
        issues.append(_chk("medium", f'{len(test_mods)} module(s) contain testing artefacts: {_top5(test_mods)}.'))
    for idx, f in enumerate(files):
        if _fancy_match(f["name"], TEST_FILE_PATTERNS): files[idx]["is_ignored"] = True
    for idx, m in enumerate(modules):
        if _fancy_match(m["name"], TEST_FILE_PATTERNS): modules[idx]["is_ignored"] = True
    return issues, recs

def check_20_too_many_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if len(files) > MAX_FILE_COUNT:
        issues.append(_chk("medium", f"Too many files uploaded ({len(files)}). May cause many modules and long scan times."))
    return issues, recs

def check_21_excess_microsoft(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    ef = _fancy_match_files(files, EXCESS_MSFT_PATTERNS)
    if ef:
        issues.append(_chk("medium", f'{len(ef)} .NET Roslyn/Runtime component(s) uploaded: {_top5(ef)}.'))
        recs.append("Do not include unnecessary Microsoft runtime components.")
    return issues, recs

def check_22_loose_class_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    lcf = _fancy_match_files(files, LOOSE_CLASS_FILE)
    lcm = _fancy_match_modules(modules, LOOSE_CLASS_MODULE)
    if lcf or lcm:
        issues.append(_chk("medium", "Java class files not packaged in JAR/WAR/EAR. Suboptimal compilation."))
        recs.append("Compile Java into JAR/WAR/EAR per packaging instructions.")
        recs.append("Use Veracode CLI: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_23_go_workspace(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    gwf = _fancy_match_files(files, GO_WORKSPACE_PATTERNS)
    if gwf:
        issues.append(_chk("medium", "Go workspaces detected. Multi-module workspaces not supported."))
        recs.append("Follow Go packaging: https://docs.veracode.com/r/compilation_go.")
    return issues, recs

def check_24_unselected_first_party(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    ufp = list(dict.fromkeys([m["name"] for m in modules if not m["is_dep"] and not m["is_ignored"]
           and not m.get("is_selected") and not m["is_third_party"] and not _is_js_module(m["name"]) and not m["has_fatal"]]))
    if ufp:
        issues.append(_chk("medium", f'{len(ufp)} first-party module(s) not selected: {_top5(ufp)}.'))
        recs.append("Under-selection affects quality: https://community.veracode.com/s/article/What-are-Modules-and-how-do-my-results-change-based-on-what-I-select.")
    return issues, recs

def check_25_over_scanning(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    """Detect modules selected that are dependencies of other selected modules."""
    issues: list[Issue] = []; recs: list[str] = []
    selected = {m["name"] for m in modules if m.get("is_selected") and not m["is_ignored"]}
    dep_names = {m["name"] for m in modules if m["is_dep"]}
    over = sorted(selected & dep_names)
    if over:
        issues.append(_chk("medium", f'{len(over)} module(s) selected but are dependencies of other modules: {_top5(over)}. This can lead to duplicate flaw reporting.'))
        recs.append("Only select main entry points, not dependency libraries.")
    return issues, recs

def check_26_dependencies_selected(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    dep_sel = list(dict.fromkeys([m["name"] for m in modules if m.get("is_selected") and not m["is_ignored"] and m["is_dep"]]))
    if dep_sel:
        issues.append(_chk("medium", f'{len(dep_sel)} dependenc{"y" if len(dep_sel)==1 else "ies"} selected as entry point(s): {_top5(dep_sel)}.'))
        recs.append("Only select main entry points: https://community.veracode.com/s/article/What-are-Modules-and-how-do-my-results-change-based-on-what-I-select.")
    return issues, recs

def check_27_duplicate_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    seen: dict[str, list[str]] = {}
    for f in files:
        if f.get("is_ignored") or f.get("is_third_party"): continue
        seen.setdefault(f["name"], []).append(f.get("md5",""))
    diff = {n: h for n, h in seen.items() if len(h) > 1 and len(set(h)) > 1}
    same = {n: len(h) for n, h in seen.items() if len(h) > 1 and len(set(h)) == 1}
    if diff:
        issues.append(_chk("high", f'{len(diff)} duplicate filename(s) with different hashes: {_top5(list(diff.keys()))}. Can cause indeterministic results.'))
    if same:
        issues.append(_chk("medium", f'{len(same)} duplicate file(s) uploaded. Slows scan time.'))
    if diff or same:
        recs.append("De-duplicate modules before upload.")
        recs.append("Upload only one version of each component.")
    return issues, recs

def check_28_minified_js(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    mf = list(dict.fromkeys([f["name"] for f in files if f["name"].lower().endswith(".min.js")]))
    if mf:
        issues.append(_chk("medium", f'{len(mf)} minified JS file(s) uploaded: {_top5(mf)}. Will not be scanned.'))
        recs.append("Submit readable JavaScript source: https://docs.veracode.com/r/compilation_jscript.")
    mm = list(dict.fromkeys([m["name"] for m in modules if _is_js_module(m["name"])
        for iss in m.get("issues",[]) if "because we think it is minified" in iss or "dist/" in iss.lower()]))
    if mm:
        issues.append(_chk("medium", f'{len(mm)} minified JS within modules: {_top5(mm)}.'))
    return issues, recs

def check_29_module_count(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    sel = [m for m in modules if m.get("is_selected")]
    if len(sel) > MAX_SELECTED_MODULE_COUNT:
        issues.append(_chk("medium", f"{len(sel)} modules selected. Suboptimal configuration."))
        recs.append("Select correct modules. Consider splitting application profiles.")
    if len(modules) > MAX_MODULE_COUNT:
        issues.append(_chk("medium", f"{len(modules)} modules identified. Suboptimal upload."))
        recs.append("Follow packaging guidance.")
    return issues, recs

def check_30_regular_scans(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], app_mod: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if app_mod:
        ds = _days_since(app_mod)
        if ds is not None and ds > STALE_SCAN_DAYS:
            issues.append(_chk("medium", f"Application not scanned recently (last activity {ds} days ago)."))
            recs.append("Regular scanning via automation allows faster response to new issues.")
    return issues, recs

def check_31_analysis_size(files: list[dict], modules: list[dict], _f: FlawSummary, scan_meta: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    """Flag when analysis size or total module size exceeds thresholds."""
    issues: list[Issue] = []; recs: list[str] = []
    asz = scan_meta.get("analysis_size", 0)
    if asz > MAX_ANALYSIS_SIZE:
        issues.append(_chk("medium", f"Analysis size ({asz/(1024*1024):.0f} MB) exceeds {MAX_ANALYSIS_SIZE//(1024*1024)} MB threshold."))
        recs.append("Review packaging to exclude unnecessary files and third-party libraries.")
    total_mod_size = sum(_parse_module_size(m.get("size","")) for m in modules)
    if total_mod_size > MAX_TOTAL_MODULE_SIZE:
        issues.append(_chk("medium", f"Total module size ({total_mod_size/(1024*1024):.0f} MB) exceeds {MAX_TOTAL_MODULE_SIZE//(1024*1024)} MB threshold."))
        recs.append("Reduce upload size by following packaging instructions.")
    return issues, recs


CHECK_REGISTRY: list[tuple[int, str, CheckFunc]] = [
    (1, "ignoreJunkFiles", check_01_junk_files),
    (2, "thirdParty", check_02_third_party),
    (3, "flawCount", check_03_flaw_count),
    (4, "fatalErrors", check_04_fatal_errors),
    (5, "unscannableJava", check_05_unscannable_java),
    (6, "detectUnwantedFiles", check_06_unwanted_files),
    (7, "nestedArchives", check_07_nested_archives),
    (8, "missingPrecompiled", check_08_missing_precompiled),
    (9, "missingSCA", check_09_missing_sca),
    (10, "unselectedJS", check_10_unselected_js),
    (11, "unexpectedSource", check_11_unexpected_source),
    (12, "missingSupporting", check_12_missing_supporting),
    (13, "missingDebug", check_13_missing_debug),
    (14, "unsupportedPlatform", check_14_unsupported_platform),
    (15, "gradleWrapper", check_15_gradle_wrapper),
    (16, "sensitiveFiles", check_16_sensitive_files),
    (17, "repositories", check_17_repositories),
    (18, "nodeModules", check_18_node_modules),
    (19, "testingArtefacts", check_19_testing_artefacts),
    (20, "tooManyFiles", check_20_too_many_files),
    (21, "excessMicrosoft", check_21_excess_microsoft),
    (22, "looseClassFiles", check_22_loose_class_files),
    (23, "goWorkspace", check_23_go_workspace),
    (24, "unselectedFirstParty", check_24_unselected_first_party),
    (25, "overScanning", check_25_over_scanning),
    (26, "dependenciesSelected", check_26_dependencies_selected),
    (27, "duplicateFiles", check_27_duplicate_files),
    (28, "minifiedJS", check_28_minified_js),
    (29, "moduleCount", check_29_module_count),
    (30, "regularScans", check_30_regular_scans),
    (31, "analysisSize", check_31_analysis_size),
]

def run_checks(files: list[dict], modules: list[dict], flaws: FlawSummary,
               scan_meta: dict, sca_on: bool, sca_comps: list[str],
               app_mod: str, skip: set[int] | None = None
               ) -> tuple[list[Issue], list[str], dict[int, list[str]]]:
    """Returns (all_issues, all_recs, check_recs) where check_recs maps check# to its recs."""
    all_issues: list[Issue] = []; all_recs: list[str] = []
    check_recs: dict[int, list[str]] = {}
    for num, name, func in CHECK_REGISTRY:
        if skip and num in skip: continue
        try:
            iss, rcs = func(files, modules, flaws, scan_meta, sca_on, sca_comps, app_mod)
            for i in iss:
                i.check_num = num
                i.check_name = name
            all_issues.extend(iss)
            if rcs:
                check_recs[num] = rcs
            for r in rcs:
                if r not in all_recs: all_recs.append(r)
        except Exception as e:
            log.warning("Check #%d (%s) failed: %s", num, name, e)
    return all_issues, all_recs, check_recs


# ==========================================================================
# Orchestration
# ==========================================================================

def _find_latest_published_build(client: VeracodeClient, aid: int, builds: list[dict]) -> dict | None:
    """Return the latest build that has published results, searching from newest to oldest."""
    for b in reversed(builds):
        bi = client.get_build_info(aid, b["id"])
        if bi and bi.get("published"):
            return b
    # Fallback: return the last build even if not published
    return builds[-1] if builds else None


def _process_build(client: VeracodeClient, app: dict, builds: list[dict],
                   legacy_id: int, sandbox_name: str = "",
                   skip_checks: set[int] | None = None,
                   prev_data: dict | None = None) -> tuple[ScanResult, list[ModuleRow], list[FileRow], list[RecommendationRow], list[AggIssue], TrendRow | None]:
    build = _find_latest_published_build(client, legacy_id, builds)
    if build is None:
        sr = _empty_result(app, legacy_id, sandbox_name)
        sr.issues_text = "[HIGH] No published build found"
        sr.health = "Poor"; sr.high_issues = 1; sr.total_issues = 1
        return sr, [], [], [], [], None

    bid = build["id"]
    dr = client.get_detailed_report(bid)
    files = client.get_files(legacy_id, bid)
    prescan = client.get_prescan(legacy_id, bid)
    app_info = client.get_app_info(legacy_id)
    app_mod_date = (app_info or {}).get("modified","") if not sandbox_name else ""

    if dr is None:
        fl = FlawSummary()
        dr = {"account_id":"","app_id":"","sandbox_id":0,"sandbox_name":sandbox_name,
              "analysis_id":"","sau_id":"","bu":"","app_name":app["name"],
              "scan_name":build.get("ver",""),"engine":"",
              "submitted":"","published":"","analysis_size":0,
              "is_latest":True,"flaws":fl,"dr_modules":[],"sca_on":False,"sca_comps":[]}
    else:
        fl = dr["flaws"]

    modules = _merge_modules(dr["dr_modules"], prescan)

    base = client.base
    acct = dr.get("account_id",""); aid_str = dr.get("app_id","")
    an_id = dr.get("analysis_id",""); sau = dr.get("sau_id",""); sbx_id = dr.get("sandbox_id",0)
    rev_url = f"{base}/auth/index.jsp#AnalyzeAppModuleList:{acct}:{aid_str}:{bid}:{an_id}:{sau}::::{sbx_id}" if acct else ""
    tri_url = f"{base}/auth/index.jsp#ReviewResultsStaticFlaws:{acct}:{aid_str}:{bid}:{an_id}:{sau}::::{sbx_id}" if acct else ""

    scan_meta = {"review_url": rev_url, "triage_url": tri_url, "analysis_size": dr.get("analysis_size",0)}
    issues, recs, check_recs = run_checks(files, modules, fl, scan_meta,
                              dr.get("sca_on",False), dr.get("sca_comps",[]),
                              app_mod_date, skip_checks)

    hi = sum(1 for i in issues if i.severity=="high")
    mi = sum(1 for i in issues if i.severity=="medium")
    health = "Good" if not issues else ("Poor" if hi else ("Fair" if mi else "Good"))

    ds = _days_since(dr.get("published",""))
    dur = _dur(dr.get("submitted",""), dr.get("published",""))
    sel = [m for m in modules if m.get("is_selected")]
    sel_names = ", ".join(m["name"] for m in sel)
    total_upload = sum(_parse_module_size(m.get("size","")) for m in modules)
    asz = dr.get("analysis_size",0)

    # Trend
    trend: TrendRow | None = None
    trend_label = ""
    key = (app["name"], sandbox_name or dr.get("sandbox_name",""))
    if prev_data and key in prev_data:
        prev = prev_data[key]
        ph = prev.get("Health",""); pf = _si(prev.get("Total Flaws")); po = _si(prev.get("Open Affecting Policy"))
        if ph and health != ph:
            trend_label = "Improved" if (health == "Good" or (health == "Fair" and ph == "Poor")) else "Degraded"
        elif ph: trend_label = "Unchanged"
        else: trend_label = "New"
        trend = TrendRow(app_name=app["name"], sandbox=sandbox_name or dr.get("sandbox_name",""),
            prev_health=ph, curr_health=health, change=trend_label,
            prev_flaws=pf, curr_flaws=fl.total, flaw_delta=fl.total-pf,
            prev_open_pol=po, curr_open_pol=fl.open_pol, open_pol_delta=fl.open_pol-po)
    elif prev_data:
        trend_label = "New"

    sr = ScanResult(
        app_name=app["name"], bu=app.get("bu",""), policy=app.get("policy",""),
        app_id=legacy_id, sandbox=sandbox_name or dr.get("sandbox_name",""),
        build_id=bid, scan_name=dr.get("scan_name",""),
        is_latest=dr.get("is_latest",True),
        scan_status="Results Ready" if dr.get("published") else "No Results",
        published=dr.get("published",""), days_since=ds if ds is not None else "N/A",
        duration=dur, engine=dr.get("engine",""),
        analysis_size_mb=round(asz/(1024*1024),2) if asz else 0,
        total_upload_mb=round(total_upload/(1024*1024),2) if total_upload else 0,
        files_uploaded=len(files), total_modules=len(modules),
        selected_modules=len(sel), selected_names=sel_names[:500],
        fatal_errors=sum(1 for m in modules if m["has_fatal"]),
        flaws=fl, sca_count=len(dr.get("sca_comps",[])),
        age_bucket=_age_bucket(ds),
        health=health, health_trend=trend_label,
        high_issues=hi, medium_issues=mi, total_issues=len(issues),
        issues_text="; ".join(f"[{i.severity.upper()}] {i.description}" for i in issues) if issues else "None",
        recs_text="; ".join(recs) if recs else "None",
        review_url=rev_url, triage_url=tri_url,
    )

    mod_rows = [ModuleRow(app_name=app["name"], build_id=bid, name=m["name"],
                status=m.get("status",""), selected=m.get("is_selected",False),
                dependency=m["is_dep"], fatal=m["has_fatal"],
                third_party=m["is_third_party"], ignored=m["is_ignored"],
                platform=m.get("platform",""), compiler=m.get("compiler",""),
                size=m.get("size",""), issues="; ".join(m.get("issues",[])))
                for m in modules]

    file_rows = [FileRow(app_name=app["name"], build_id=bid, name=f["name"],
                 status=f["status"], md5=f["md5"],
                 ignored=f["is_ignored"], third_party=f["is_third_party"])
                 for f in files]

    # Build recommendation rows -- match each rec to the highest-severity issue it came from
    rec_rows: list[RecommendationRow] = []
    for r in recs:
        # Find the most severe issue whose check produced this rec
        best_sev = "low"
        for i in issues:
            if best_sev != "high" and i.severity == "high": best_sev = "high"
            elif best_sev == "low" and i.severity == "medium": best_sev = "medium"
        rec_rows.append(RecommendationRow(app_name=app["name"], build_id=bid,
                    severity=best_sev, recommendation=r, doc_url=_extract_url(r)))

    # Build structured aggregation issues (one per issue, with check metadata and matched rec)
    sb_name = sandbox_name or dr.get("sandbox_name", "")
    agg_issues: list[AggIssue] = []
    for i in issues:
        # Get the first recommendation from this specific check
        check_rec_list = check_recs.get(i.check_num, [])
        matched_rec = check_rec_list[0] if check_rec_list else ""
        agg_issues.append(AggIssue(
            app_name=app["name"], bu=app.get("bu", ""), sandbox=sb_name,
            check_num=i.check_num, check_name=i.check_name,
            category=CHECK_CATEGORIES.get(i.check_num, "Other"),
            severity=i.severity, description=i.description,
            recommendation=matched_rec,
        ))

    return sr, mod_rows, file_rows, rec_rows, agg_issues, trend


def _empty_result(app: dict, lid: int, sandbox: str = "") -> ScanResult:
    return ScanResult(app_name=app["name"], bu=app.get("bu",""), policy=app.get("policy",""),
        app_id=lid, sandbox=sandbox, scan_status="No Scan", health="Poor",
        high_issues=1, total_issues=1, issues_text="[HIGH] No policy scan found",
        recs_text="None", age_bucket="N/A")


def _process_app(client: VeracodeClient, app: dict, skip_no: bool, inc_sb: bool,
                 skip_checks: set[int] | None, prev_data: dict | None,
                 resume_keys: set[tuple[str, str]] | None
                 ) -> tuple[list[ScanResult], list[ModuleRow], list[FileRow], list[RecommendationRow], list[AggIssue], list[TrendRow]]:
    lid = app.get("legacy_id")
    if not lid: return [], [], [], [], [], []
    rs: list[ScanResult] = []; ms: list[ModuleRow] = []; fs: list[FileRow] = []
    rrs: list[RecommendationRow] = []; ais: list[AggIssue] = []; ts: list[TrendRow] = []

    key = (app["name"], "")
    if resume_keys and key in resume_keys:
        log.debug("    Skipping (resume): %s", app["name"]); return rs, ms, fs, rrs, ais, ts

    builds = client.get_builds(lid)
    if not builds:
        if skip_no: return rs, ms, fs, rrs, ais, ts
        rs.append(_empty_result(app, lid))
        return rs, ms, fs, rrs, ais, ts

    sr, mr, fr, rr, ai, tr = _process_build(client, app, builds, lid,
                                         skip_checks=skip_checks, prev_data=prev_data)
    rs.append(sr); ms.extend(mr); fs.extend(fr); rrs.extend(rr); ais.extend(ai)
    if tr: ts.append(tr)

    if inc_sb:
        for sb in client.get_sandboxes(lid):
            sb_key = (app["name"], sb["name"])
            if resume_keys and sb_key in resume_keys: continue
            sb_builds = client.get_builds(lid, sb["id"])
            if not sb_builds: continue
            sr2, mr2, fr2, rr2, ai2, tr2 = _process_build(client, app, sb_builds, lid, sb["name"],
                                                       skip_checks=skip_checks, prev_data=prev_data)
            rs.append(sr2); ms.extend(mr2); fs.extend(fr2); rrs.extend(rr2); ais.extend(ai2)
            if tr2: ts.append(tr2)

    return rs, ms, fs, rrs, ais, ts


# ==========================================================================
# Resume
# ==========================================================================

def _load_resume_keys(path: str) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb["Scan Health Summary"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ai = headers.index("App Name") if "App Name" in headers else -1
        si = headers.index("Sandbox") if "Sandbox" in headers else -1
        if ai >= 0:
            for row in ws.iter_rows(min_row=2, values_only=True):
                keys.add((str(row[ai] or ""), str(row[si] or "") if si >= 0 else ""))
        wb.close()
    except Exception as e:
        log.warning("Could not load resume file: %s", e)
    return keys


# ==========================================================================
# Trend
# ==========================================================================

def _load_previous(path: str) -> dict[tuple[str, str], dict]:
    data: dict[tuple[str, str], dict] = {}
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb["Scan Health Summary"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rd = dict(zip(headers, row))
            key = (str(rd.get("App Name","")), str(rd.get("Sandbox","") or ""))
            data[key] = rd
        wb.close()
    except Exception as e:
        log.warning("Could not load previous report: %s", e)
    return data


def _load_previous_agg(path: str) -> dict[int, int]:
    """Read 'Check # -> Apps Affected' from a prior report's Tenant Aggregation
    sheet so the dashboard can trend each health check over time. Optional: an
    older report without that sheet simply yields no baseline."""
    counts: dict[int, int] = {}
    try:
        wb = load_workbook(path, read_only=True)
        if "Tenant Aggregation" not in wb.sheetnames:
            wb.close(); return counts
        ws = wb["Tenant Aggregation"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ci = headers.index("Check #") if "Check #" in headers else -1
        ai = headers.index("Apps Affected") if "Apps Affected" in headers else -1
        if ci >= 0 and ai >= 0:
            for row in ws.iter_rows(min_row=2, values_only=True):
                n = _si(row[ci], -1)
                if n > 0: counts[n] = _si(row[ai])
        wb.close()
    except Exception as e:
        log.debug("Could not load previous aggregation: %s", e)
    return counts


# ==========================================================================
# Output
# ==========================================================================

_CW: dict[str, int] = {"App Name":30,"Issues":80,"Recommendations":80,"Module":30,
    "File":35,"Review Modules URL":50,"Triage Flaws URL":50,"Policy":25,
    "Business Unit":20,"Selected Module Names":40,"Recommendation":60,"Doc URL":40,
    "Check Name":22,"Category":18,"Issue Pattern":60,"Top Recommendation":60,
    "Affected App Names":50,"Business Units":35}

def _hdr(ws: object, n: int) -> None:
    for c in range(1,n+1):
        cl=ws.cell(row=1,column=c); cl.fill=_HF; cl.font=_HN; cl.alignment=_CA; cl.border=_BD

def _sheet(ws: object, rows: list[dict], hcol: str | None = None, age_col: str | None = None) -> None:
    if not rows: return
    hds=list(rows[0].keys()); ws.append(hds); _hdr(ws,len(hds)); ws.freeze_panes="A2"
    for ri,rd in enumerate(rows,2):
        for ci,h in enumerate(hds,1):
            cl=ws.cell(row=ri,column=ci,value=rd.get(h,"")); cl.font=_DF; cl.border=_BD
            if hcol and h==hcol:
                cl.fill=PatternFill("solid",fgColor=_CLR.get(str(rd.get(h,"")),"FFFFFF")); cl.alignment=_CA
            elif age_col and h==age_col:
                cl.fill=PatternFill("solid",fgColor=_AGE_CLR.get(str(rd.get(h,"")),"FFFFFF")); cl.alignment=_CA
            elif h in ("Issues","Recommendations","Recommendation"): cl.alignment=_WA
            else: cl.alignment=_CA
    for ci,h in enumerate(hds,1):
        ws.column_dimensions[get_column_letter(ci)].width=_CW.get(h,14)
    ws.auto_filter.ref=ws.dimensions


def _build_aggregation(agg_issues: list[AggIssue], total_apps: int) -> list[dict]:
    """Build accurate tenant-level aggregation from structured AggIssue records.

    Groups issues by (check_num, check_name) so each row represents one type of
    problem across the tenant. Uses the real check metadata rather than regex
    guesswork on serialized strings.

    Output columns:
      - Check #: the check number for reference
      - Check Name: machine-readable check name
      - Category: Packaging / Module Selection / Fatal Errors / etc.
      - Severity: highest severity observed for this check across all apps
      - Apps Affected: count of unique apps that triggered this check
      - % of Tenant: percentage of total apps
      - Affected App Names: comma-separated, max 10 then "and N others"
      - Business Units Affected: unique BUs with this issue
      - Sample Issue: one representative issue description (shortest, for clarity)
      - Top Recommendation: the recommendation linked to this check
    """
    if not agg_issues:
        return []

    # Group by check number
    by_check: dict[int, list[AggIssue]] = {}
    for ai in agg_issues:
        by_check.setdefault(ai.check_num, []).append(ai)

    rows: list[dict] = []
    for check_num in sorted(by_check.keys()):
        group = by_check[check_num]
        check_name = group[0].check_name
        category = group[0].category

        # Highest severity in this group
        sevs = {ai.severity for ai in group}
        if "high" in sevs:
            sev = "HIGH"
        elif "medium" in sevs:
            sev = "MEDIUM"
        else:
            sev = "LOW"

        # Unique apps
        unique_apps = list(dict.fromkeys(ai.app_name for ai in group))
        app_count = len(unique_apps)
        pct = round(100 * app_count / total_apps, 1) if total_apps else 0

        app_str = ", ".join(unique_apps[:10])
        if len(unique_apps) > 10:
            app_str += f" and {len(unique_apps) - 10} others"

        # Unique business units with counts
        bu_counter: Counter[str] = Counter()
        for ai in group:
            if ai.bu: bu_counter[ai.bu] += 1
        bu_parts = [f"{bu} ({cnt})" for bu, cnt in bu_counter.most_common(10)]
        bu_str = ", ".join(bu_parts)
        if len(bu_counter) > 10:
            bu_str += f" and {len(bu_counter) - 10} others"

        # Total issue occurrences (may exceed app count when one app triggers
        # multiple sub-issues from the same check, e.g. multiple fatal error types)
        occurrence_count = len(group)

        # Representative issue: pick the most common normalized pattern
        pattern_counter: Counter[str] = Counter()
        for ai in group:
            clean = re.sub(r'"[^"]*"', '(name)', ai.description)
            clean = re.sub(r'\b\d+\b', 'N', clean)
            pattern_counter[clean] += 1
        sample = pattern_counter.most_common(1)[0][0] if pattern_counter else ""

        # Top recommendation: pick the first non-empty one
        rec = ""
        for ai in group:
            if ai.recommendation:
                rec = ai.recommendation
                break

        rows.append({
            "Check #": check_num,
            "Check Name": check_name,
            "Category": category,
            "Severity": sev,
            "Apps Affected": app_count,
            "% of Tenant": pct,
            "Total Occurrences": occurrence_count,
            "Affected App Names": app_str,
            "Business Units": bu_str,
            "Issue Pattern": sample,
            "Top Recommendation": rec,
        })

    # Sort by severity (HIGH first), then by app count descending
    sev_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (sev_order.get(r["Severity"], 3), -r["Apps Affected"]))

    return rows


def write_excel(health: list[dict], mods: list[dict], files: list[dict],
                recs: list[dict], trends: list[dict], agg: list[dict], path: str,
                dash_bundle: dict | None = None) -> None:
    wb=Workbook()
    ws=wb.active; ws.title="Scan Health Summary"
    if not health: ws["A1"]="No data."; wb.save(path); return
    _sheet(ws, health, hcol="Health", age_col="Scan Age Bucket")
    if mods: _sheet(wb.create_sheet("Module Details"), mods)
    if files: _sheet(wb.create_sheet("Uploaded Files"), files)
    if recs: _sheet(wb.create_sheet("Recommendations"), recs)
    if trends: _sheet(wb.create_sheet("Trends"), trends)
    if agg: _sheet(wb.create_sheet("Tenant Aggregation"), agg)
    # Overview
    wso=wb.create_sheet("Tenant Overview")
    tot=len(health); good=sum(1 for r in health if r.get("Health")=="Good")
    poor=sum(1 for r in health if r.get("Health")=="Poor")
    fair=sum(1 for r in health if r.get("Health")=="Fair")
    ns=sum(1 for r in health if r.get("Scan Status")=="No Scan")
    ds=[r["Days Since Scan"] for r in health if isinstance(r.get("Days Since Scan"),int)]
    avg=round(sum(ds)/len(ds),1) if ds else 0
    stats=[("Report Generated",datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
           ("Total Applications",tot),("Good Health",good),("Fair",fair),("Poor Health",poor),
           ("No Scan Found",ns),
           ("Total Flaws",sum(_si(r.get("Total Flaws")) for r in health)),
           ("Total Open Affecting Policy",sum(_si(r.get("Open Affecting Policy")) for r in health)),
           ("Avg Days Since Scan",avg)]
    wso.column_dimensions["A"].width=40; wso.column_dimensions["B"].width=20
    wso.append(["Metric","Value"]); _hdr(wso,2)
    for ri,(k,v) in enumerate(stats,2):
        wso.cell(row=ri,column=1,value=k).font=_BF; wso.cell(row=ri,column=2,value=v).font=_DF
        for c in (1,2): wso.cell(row=ri,column=c).border=_BD; wso.cell(row=ri,column=c).alignment=_CA
    # Executive Dashboard / heatmaps are prepended; the seven sheets above are untouched.
    if dash_bundle:
        try:
            write_dashboard_sheets(wb, dash_bundle["apps"], dash_bundle["issues"],
                                         dash_bundle["metrics"], dash_bundle["tenant_trend"],
                                         dash_bundle["issue_rows"], dash_bundle["config"])
        except Exception as e:
            log.warning("Dashboard sheet generation failed, core report unaffected: %s", e)
    wb.save(path); log.info("[+] Report saved: %s", path)


def write_csv(health: list[dict], mods: list[dict], files: list[dict],
              recs: list[dict], trends: list[dict], agg: list[dict], base_path: str) -> None:
    stem = Path(base_path).stem
    parent = Path(base_path).parent
    for name, rows in [("summary",health),("modules",mods),("files",files),
                       ("recommendations",recs),("trends",trends),("aggregation",agg)]:
        if not rows: continue
        p = parent / f"{stem}_{name}.csv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader(); w.writerows(rows)
        log.info("[+] CSV: %s", p)


def write_json(health: list[dict], mods: list[dict], files: list[dict],
               trends: list[dict], path: str) -> None:
    out = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "summary": {"total": len(health),
                     "good": sum(1 for r in health if r.get("Health")=="Good"),
                     "fair": sum(1 for r in health if r.get("Health")=="Fair"),
                     "poor": sum(1 for r in health if r.get("Health")=="Poor")},
        "apps": health, "modules": mods, "files": files,
        "trends": trends or None,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, default=str)
    log.info("[+] JSON: %s", path)


# ==========================================================================
# Executive Dashboard layer
#
# Pure presentation/scoring. Makes NO Veracode API calls: it consumes the
# rows already built by the scanner above.
#   Normalization -> Scoring -> Aggregation -> Visualization -> Export
# ==========================================================================

# ==========================================================================
# Configuration
# ==========================================================================

DASHBOARD_CONFIG: dict[str, Any] = {
    # Scan recency buckets, in days since the published scan.
    "scan_age": {
        "healthy_days": 30,
        "warning_days": 60,
        "critical_days": 90,
    },

    # Count of failing scan-health checks on an application.
    "health_issues": {
        "green_max": 0,     # <= this is GREEN
        "yellow_max": 2,    # <= this is YELLOW, above is RED
    },

    # Count of HIGH-severity failing checks on an application.
    "high_severity_issues": {
        "green_max": 0,
        "yellow_max": 1,
    },

    # Security-findings thresholds. Applied to OPEN flaws (not fixed, not
    # mitigated). "very_high" is Veracode severity 5, "high" is severity 4.
    "flaws": {
        "very_high_yellow": 1,   # >= this -> at least YELLOW
        "very_high_red": 1,      # >= this -> RED
        "high_yellow": 1,
        "high_red": 5,
        "total_open_yellow": 50,
        "total_open_red": 250,
    },

    # Attention Score weights. Any component whose input is unavailable is
    # dropped from BOTH numerator and denominator, so the score stays 0-100
    # and missing data never silently reads as "healthy".
    "attention_score": {
        "scan_health_weight": 30,
        "scan_age_weight": 20,
        "flaw_weight": 25,
        "high_severity_issue_weight": 15,
        "trend_weight": 10,
        "data_quality_weight": 10,
        # Ceiling applied to profiles with no published scan results. Without
        # it, every never-scanned profile scores 100 and crowds genuinely
        # measured high-risk applications out of the Top 10. An unmeasured
        # application is a real gap, but it is not evidence of immediate risk.
        # Set to 100 to disable the cap.
        "never_scanned_cap": 70,
    },

    # Attention Score bands: (min, max, label, risk_level)
    "attention_bands": [
        (0, 19, "Normal", "GREEN"),
        (20, 39, "Monitor", "GREEN"),
        (40, 59, "Attention Required", "YELLOW"),
        (60, 79, "High Priority", "ORANGE"),
        (80, 100, "Immediate Attention", "RED"),
    ],

    # Organisational issue prevalence, as a share of the tenant.
    "issue_prevalence": {
        "widespread_pct": 20.0,
        "common_pct": 5.0,
    },

    "top_n_attention": 10,
    "top_n_issues": 10,

    "colors": {
        "GREEN": "2E7D32",
        "YELLOW": "F9A825",
        "ORANGE": "EF6C00",
        "RED": "C62828",
        "GRAY": "9E9E9E",
    },
    # Softer fills for large table bodies; text stays readable.
    "fills": {
        "GREEN": "D6EAD7",
        "YELLOW": "FDF0CC",
        "ORANGE": "FBDFC8",
        "RED": "F4D3D3",
        "GRAY": "E8E8E8",
    },
}

DISCLAIMER = (
    "Scan health indicates the quality and currency of security analysis. "
    "A healthy scan does not mean an application has no security vulnerabilities."
)

RISK_ORDER = {"RED": 0, "ORANGE": 1, "YELLOW": 2, "GREEN": 3, "GRAY": 4}

# Human-readable labels for the 31 checks, and the highest severity each check
# is capable of emitting. Used so the organisational heatmap can list every
# check, including the ones that nothing triggered.
CHECK_CATALOG: dict[int, tuple[str, str]] = {
    1:  ("Unnecessary / junk files uploaded", "medium"),
    2:  ("Third-party component selected as entry point", "medium"),
    3:  ("Implausible flaw count (zero or excessive)", "medium"),
    4:  ("Fatal scan errors", "high"),
    5:  ("Unscannable Java modules", "high"),
    6:  ("Unwanted file types uploaded", "medium"),
    7:  ("Nested archives", "high"),
    8:  ("Missing precompiled .NET views", "high"),
    9:  ("Missing SCA results", "medium"),
    10: ("JavaScript modules not selected", "medium"),
    11: ("Unexpected source code uploaded", "high"),
    12: ("Missing supporting files", "medium"),
    13: ("Missing debug symbols (PDB)", "medium"),
    14: ("Unsupported platform or compiler", "high"),
    15: ("Gradle wrapper selected for analysis", "high"),
    16: ("Sensitive files uploaded", "high"),
    17: ("Source repository uploaded", "medium"),
    18: ("node_modules uploaded", "medium"),
    19: ("Testing artefacts uploaded or selected", "high"),
    20: ("Excessive uploaded file count", "medium"),
    21: ("Excess Microsoft runtime components", "medium"),
    22: ("Loose Java class files", "medium"),
    23: ("Go workspace detected", "medium"),
    24: ("First-party modules not selected", "medium"),
    25: ("Over-scanning (dependencies of selected modules)", "medium"),
    26: ("Dependencies selected as entry points", "medium"),
    27: ("Duplicate files uploaded", "high"),
    28: ("Minified JavaScript uploaded", "medium"),
    29: ("Excessive module count", "medium"),
    30: ("Application not scanned recently", "medium"),
    31: ("Analysis or upload size over threshold", "medium"),
}



# ==========================================================================
# Small helpers
# ==========================================================================

def _i(v: Any, d: int = 0) -> int:
    try:
        if v is None or v == "":
            return d
        return int(v)
    except (ValueError, TypeError):
        return d


def _f(v: Any, d: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return d
        return float(v)
    except (ValueError, TypeError):
        return d


def _pct(num: float, den: float, nd: int = 1) -> float:
    """Percentage that never divides by zero."""
    if not den:
        return 0.0
    return round(100.0 * num / den, nd)


def _ctx(num: int, den: int, nd: int = 1) -> str:
    """'43 / 250 (17.2%)' -- never hide the denominator."""
    return f"{num:,} / {den:,} ({_pct(num, den, nd)}%)"


def _clamp(v: float, lo: float = 0.0, hi: float = 1.0) -> float:
    return max(lo, min(hi, v))


def _worst(*levels: str) -> str:
    """Worst (most severe) risk level of those supplied, ignoring GRAY unless
    it is the only input."""
    known = [l for l in levels if l and l != "GRAY"]
    if not known:
        return "GRAY" if levels else "GRAY"
    return min(known, key=lambda l: RISK_ORDER.get(l, 99))


# ==========================================================================
# Intermediate model
# ==========================================================================

@dataclass
class HealthIssue:
    """One failing health check on one application profile."""
    app_id: str = ""
    app_name: str = ""
    sandbox: str = ""
    business_unit: str = ""
    check_number: int = 0
    check_name: str = ""
    check_label: str = ""
    category: str = ""
    severity: str = ""          # high | medium | low
    status: str = "FAILED"
    description: str = ""
    recommendation: str = ""


@dataclass
class ApplicationHealth:
    """Normalized, dashboard-ready view of one application profile."""
    app_id: str = ""
    app_name: str = ""
    sandbox: str = ""
    business_unit: str = ""
    policy: str = ""
    build_id: str = ""

    scan_status: str = ""
    health_status: str = "Unknown"      # Good | Fair | Poor | Unknown
    ever_scanned: bool = False

    last_scan_date: str = ""
    scan_age_days: int | None = None
    scan_age_bucket: str = "Never scanned"

    # Security findings. None means "not available", which is NOT zero.
    critical_flaws: int | None = None    # Veracode Very High (severity 5)
    high_flaws: int | None = None
    medium_flaws: int | None = None
    low_flaws: int | None = None
    total_flaws: int = 0
    open_policy_flaws: int = 0
    flaw_severity_available: bool = False

    sca_component_count: int = 0
    sca_status: str = "Unknown"          # Present | Missing | Unknown
    analysis_size_mb: float = 0.0
    total_upload_mb: float = 0.0
    files_uploaded: int = 0
    total_modules: int = 0
    selected_modules: int = 0
    fatal_errors: int = 0

    health_issue_count: int = 0
    high_severity_issue_count: int = 0
    medium_severity_issue_count: int = 0
    failed_checks: list[int] = field(default_factory=list)
    recommendations: list[str] = field(default_factory=list)
    top_issue: str = ""

    previous_health: str = ""
    health_trend: str = "Unknown"        # Improving | Stable | Declining | New | Unknown

    attention_score: int = 0
    attention_band: str = ""
    attention_level: str = "GRAY"
    attention_drivers: list[str] = field(default_factory=list)

    # Two independent dimensions, deliberately never merged.
    scan_health_risk: str = "GRAY"
    security_findings_risk: str = "GRAY"
    quadrant: str = ""

    review_url: str = ""
    triage_url: str = ""

    @property
    def key(self) -> str:
        return f"{self.app_id}|{self.sandbox}"

    @property
    def display_name(self) -> str:
        return f"{self.app_name} [{self.sandbox}]" if self.sandbox else self.app_name


# ==========================================================================
# Normalization
# ==========================================================================

_TREND_MAP = {
    "Improved": "Improving",
    "Degraded": "Declining",
    "Unchanged": "Stable",
    "New": "New",
    "": "Unknown",
}


def _dash_age_bucket(days: int | None, cfg: dict) -> str:
    if days is None:
        return "Never scanned"
    a = cfg["scan_age"]
    if days <= a["healthy_days"]:
        return f"<={a['healthy_days']}d"
    if days <= a["warning_days"]:
        return f"{a['healthy_days'] + 1}-{a['warning_days']}d"
    if days <= a["critical_days"]:
        return f"{a['warning_days'] + 1}-{a['critical_days']}d"
    return f">{a['critical_days']}d"


def _scan_age_risk(days: int | None, cfg: dict) -> str:
    if days is None:
        return "GRAY"
    a = cfg["scan_age"]
    if days <= a["healthy_days"]:
        return "GREEN"
    if days <= a["warning_days"]:
        return "YELLOW"
    if days <= a["critical_days"]:
        return "ORANGE"
    return "RED"


def _health_risk(health: str) -> str:
    return {"Good": "GREEN", "Fair": "YELLOW", "Poor": "RED"}.get(health, "GRAY")


def _issue_count_risk(n: int, bounds: dict) -> str:
    if n <= bounds["green_max"]:
        return "GREEN"
    if n <= bounds["yellow_max"]:
        return "YELLOW"
    return "RED"


def _flaw_risk(app: ApplicationHealth, cfg: dict) -> str:
    """Security-findings risk. GRAY when severity data is unavailable, because
    a flaw total without severities cannot be classified honestly."""
    if not app.ever_scanned:
        return "GRAY"
    if not app.flaw_severity_available:
        return "GRAY"
    t = cfg["flaws"]
    vh = app.critical_flaws or 0
    hi = app.high_flaws or 0
    open_total = app.open_policy_flaws
    if vh >= t["very_high_red"] or hi >= t["high_red"] or open_total >= t["total_open_red"]:
        return "RED"
    if vh >= t["very_high_yellow"] or hi >= t["high_yellow"] or open_total >= t["total_open_yellow"]:
        return "ORANGE" if (vh or hi) else "YELLOW"
    return "GREEN"


def _sca_risk(status: str) -> str:
    return {"Present": "GREEN", "Missing": "RED", "Unknown": "GRAY"}.get(status, "GRAY")


def _trend_risk(trend: str) -> str:
    return {"Improving": "GREEN", "Stable": "GRAY",
            "Declining": "RED", "New": "GRAY", "Unknown": "GRAY"}.get(trend, "GRAY")


def _quadrant(app: ApplicationHealth) -> str:
    """Priority matrix placement. X axis is scan-health risk, Y axis is
    security-findings risk. Built only from indicators the scanner actually
    produces; no business criticality is invented."""
    sh_bad = app.scan_health_risk in ("RED", "ORANGE")
    sf_bad = app.security_findings_risk in ("RED", "ORANGE")
    if app.scan_health_risk == "GRAY" and app.security_findings_risk == "GRAY":
        return "Insufficient Data"
    if sf_bad and sh_bad:
        return "Immediate"
    if sf_bad and not sh_bad:
        return "Review"
    if sh_bad and not sf_bad:
        return "Attention"
    return "Monitor"


def build_application_health(
    health_rows: Sequence[dict],
    issue_records: Sequence[dict] | None = None,
    prev_rows: dict | None = None,
    cfg: dict | None = None,
    check_categories: dict[int, str] | None = None,
) -> tuple[list[ApplicationHealth], list[HealthIssue]]:
    """Normalize scanner output into the dashboard model.

    Matching against the previous report is done on App ID first so that
    renamed applications are still tracked, falling back to name.
    """
    cfg = cfg or DASHBOARD_CONFIG
    cats = check_categories or CHECK_CATEGORIES
    issue_records = issue_records or []
    prev_rows = prev_rows or {}

    # ---- index the previous report by (app_id, sandbox) and (name, sandbox)
    prev_by_id: dict[tuple[str, str], dict] = {}
    prev_by_name: dict[tuple[str, str], dict] = {}
    for row in prev_rows.values():
        sbx = str(row.get("Sandbox") or "")
        aid = str(row.get("App ID") or "")
        if aid:
            prev_by_id[(aid, sbx)] = row
        prev_by_name[(str(row.get("App Name") or ""), sbx)] = row

    # ---- index structured issues by (app_name, sandbox); AggIssue carries no
    #      app_id, so this is the only join key available.
    issues_by_app: dict[tuple[str, str], list[dict]] = {}
    for rec in issue_records:
        k = (str(rec.get("app_name", "")), str(rec.get("sandbox", "") or ""))
        issues_by_app.setdefault(k, []).append(rec)

    apps: list[ApplicationHealth] = []
    all_issues: list[HealthIssue] = []

    for row in health_rows:
        name = str(row.get("App Name", "") or "")
        sbx = str(row.get("Sandbox", "") or "")
        app_id = str(row.get("App ID", "") or "")
        scan_status = str(row.get("Scan Status", "") or "")
        published = str(row.get("Published", "") or "")

        days_raw = row.get("Days Since Scan")
        days = days_raw if isinstance(days_raw, int) else None
        if isinstance(days_raw, str) and days_raw.isdigit():
            days = int(days_raw)

        ever_scanned = bool(published) and scan_status not in ("No Scan", "No Results")

        raw_health = str(row.get("Health", "") or "")
        # The scanner assigns Poor to never-scanned profiles. For the dashboard
        # that is unknown, not measured-and-bad.
        health = raw_health if ever_scanned else "Unknown"

        sev_flag = str(row.get("Flaw Severity Data", "") or "")
        sev_available = ever_scanned and sev_flag == "Available"

        sca_count = _i(row.get("SCA Components"))
        # SCA presence is authoritative from check #9, not from the count of
        # vulnerable components (zero vulnerable components is a good outcome).
        app_issue_recs = issues_by_app.get((name, sbx), [])
        failed_nums = sorted({_i(r.get("check_num")) for r in app_issue_recs})
        if not ever_scanned:
            sca_status = "Unknown"
        elif 9 in failed_nums:
            sca_status = "Missing"
        else:
            sca_status = "Present"

        app = ApplicationHealth(
            app_id=app_id,
            app_name=name,
            sandbox=sbx,
            business_unit=str(row.get("Business Unit", "") or ""),
            policy=str(row.get("Policy", "") or ""),
            build_id=str(row.get("Build ID", "") or ""),
            scan_status=scan_status,
            health_status=health,
            ever_scanned=ever_scanned,
            last_scan_date=published,
            scan_age_days=days,
            scan_age_bucket=_dash_age_bucket(days, cfg),
            total_flaws=_i(row.get("Total Flaws")),
            open_policy_flaws=_i(row.get("Open Affecting Policy")),
            flaw_severity_available=sev_available,
            sca_component_count=sca_count,
            sca_status=sca_status,
            analysis_size_mb=_f(row.get("Analysis Size (MB)")),
            total_upload_mb=_f(row.get("Total Upload Size (MB)")),
            files_uploaded=_i(row.get("Files Uploaded")),
            total_modules=_i(row.get("Total Modules")),
            selected_modules=_i(row.get("Selected Modules")),
            fatal_errors=_i(row.get("Fatal Errors")),
            health_issue_count=_i(row.get("Total Issues")),
            high_severity_issue_count=_i(row.get("High Issues")),
            medium_severity_issue_count=_i(row.get("Medium Issues")),
            failed_checks=failed_nums,
            review_url=str(row.get("Review Modules URL", "") or ""),
            triage_url=str(row.get("Triage Flaws URL", "") or ""),
        )

        if sev_available:
            app.critical_flaws = _i(row.get("Open Very High Flaws"))
            app.high_flaws = _i(row.get("Open High Flaws"))
            app.medium_flaws = _i(row.get("Open Medium Flaws"))
            app.low_flaws = _i(row.get("Open Low Flaws"))

        recs = str(row.get("Recommendations", "") or "")
        app.recommendations = [r.strip() for r in recs.split("; ") if r.strip() and r != "None"]

        # ---- structured issues
        for rec in app_issue_recs:
            num = _i(rec.get("check_num"))
            label = CHECK_CATALOG.get(num, ("Check %d" % num, "medium"))[0]
            hi = HealthIssue(
                app_id=app_id, app_name=name, sandbox=sbx,
                business_unit=app.business_unit,
                check_number=num,
                check_name=str(rec.get("check_name", "") or ""),
                check_label=label,
                category=str(rec.get("category", "") or cats.get(num, "Other")),
                severity=str(rec.get("severity", "") or "medium"),
                description=str(rec.get("description", "") or ""),
                recommendation=str(rec.get("recommendation", "") or ""),
            )
            all_issues.append(hi)

        # Top issue: worst severity, then the check with the highest severity rank.
        highs = [i for i in app_issue_recs if str(i.get("severity")) == "high"]
        pool = highs or app_issue_recs
        if pool:
            n = _i(pool[0].get("check_num"))
            app.top_issue = CHECK_CATALOG.get(n, (str(pool[0].get("description", ""))[:80], ""))[0]
        elif not ever_scanned:
            app.top_issue = "No published scan results"

        # ---- trend, matched on App ID first
        prow = prev_by_id.get((app_id, sbx)) if app_id else None
        if prow is None:
            prow = prev_by_name.get((name, sbx))
        if prev_rows:
            if prow is None:
                app.health_trend = "New"
                app.previous_health = ""
            else:
                app.previous_health = str(prow.get("Health", "") or "")
                app.health_trend = _classify_trend(app.previous_health, health)
        else:
            app.health_trend = "Unknown"

        # ---- risk dimensions
        app.scan_health_risk = _worst(
            _health_risk(app.health_status),
            _issue_count_risk(app.high_severity_issue_count, cfg["high_severity_issues"]),
        ) if ever_scanned else "GRAY"
        app.security_findings_risk = _flaw_risk(app, cfg)
        app.quadrant = _quadrant(app)

        apps.append(app)

    # ---- scoring
    for app in apps:
        score, band, level, drivers = compute_attention_score(app, cfg)
        app.attention_score = score
        app.attention_band = band
        app.attention_level = level
        app.attention_drivers = drivers

    apps.sort(key=lambda a: (-a.attention_score, a.app_name.lower(), a.sandbox.lower()))
    return apps, all_issues


_HEALTH_RANK = {"Good": 0, "Fair": 1, "Poor": 2, "Unknown": 3, "": 3}


def _classify_trend(previous: str, current: str) -> str:
    """Direction of travel between two health classifications.

    Unknown on either side means the direction is genuinely unknown; it is not
    reported as stable.
    """
    if not previous or previous == "Unknown" or current == "Unknown":
        return "Unknown"
    pr, cr = _HEALTH_RANK.get(previous, 3), _HEALTH_RANK.get(current, 3)
    if pr == 3 or cr == 3:
        return "Unknown"
    if cr < pr:
        return "Improving"
    if cr > pr:
        return "Declining"
    return "Stable"


# ==========================================================================
# Attention Score
# ==========================================================================

def compute_attention_score(app: ApplicationHealth,
                            cfg: dict | None = None) -> tuple[int, str, str, list[str]]:
    """Return (score 0-100, band label, risk level, explanation drivers).

    Each component contributes ``weight * risk_fraction``. Components whose
    input is unavailable are excluded from the denominator as well as the
    numerator, so an application is never rewarded for missing data and never
    penalised twice for it (the data-quality component covers that).
    """
    cfg = cfg or DASHBOARD_CONFIG
    w = cfg["attention_score"]
    parts: list[tuple[str, float, float, str]] = []   # (name, weight, fraction, reason)

    # 1. Scan health classification
    frac = {"Good": 0.0, "Fair": 0.5, "Poor": 1.0}.get(app.health_status)
    if frac is None:
        parts.append(("Scan health", w["scan_health_weight"], 1.0,
                      "Scan health is unknown (no published scan results)"))
    else:
        reason = f"Scan health is {app.health_status.upper()}"
        parts.append(("Scan health", w["scan_health_weight"], frac, reason))

    # 2. Scan recency
    a = cfg["scan_age"]
    if app.scan_age_days is None:
        parts.append(("Scan recency", w["scan_age_weight"], 1.0, "Application has never been scanned"))
    else:
        d = app.scan_age_days
        if d <= a["healthy_days"]:
            frac = 0.0
        elif d <= a["warning_days"]:
            frac = 0.4
        elif d <= a["critical_days"]:
            frac = 0.7
        else:
            frac = 1.0
        parts.append(("Scan recency", w["scan_age_weight"], frac, f"Scan is {d} days old"))

    # 3. Security findings
    t = cfg["flaws"]
    if not app.ever_scanned:
        pass  # covered by data quality; no findings dimension exists yet
    elif not app.flaw_severity_available:
        parts.append(("Security findings", w["flaw_weight"], 0.5,
                      "Flaw severity breakdown unavailable for this scan"))
    else:
        vh, hi = app.critical_flaws or 0, app.high_flaws or 0
        frac = 0.0
        bits: list[str] = []
        if vh:
            frac = max(frac, _clamp(0.6 + 0.4 * (vh / max(t["very_high_red"], 1))))
            bits.append(f"{vh} open very-high flaw(s)")
        if hi:
            frac = max(frac, _clamp(0.4 + 0.4 * (hi / max(t["high_red"], 1))))
            bits.append(f"{hi} open high flaw(s)")
        if app.open_policy_flaws >= t["total_open_yellow"]:
            frac = max(frac, _clamp(app.open_policy_flaws / max(t["total_open_red"], 1)))
            bits.append(f"{app.open_policy_flaws} open policy-affecting flaws")
        reason = ", ".join(bits) if bits else "No open very-high or high flaws"
        parts.append(("Security findings", w["flaw_weight"], frac, reason))

    # 4. High-severity scan-health checks
    if app.ever_scanned:
        n = app.high_severity_issue_count
        frac = {0: 0.0, 1: 0.5, 2: 0.75}.get(n, 1.0)
        reason = (f"{n} high-severity scan-health issue(s)" if n
                  else "No high-severity scan-health issues")
        parts.append(("High-severity issues", w["high_severity_issue_weight"], frac, reason))

    # 5. Trend
    if app.health_trend in ("Improving", "Stable", "Declining"):
        frac = {"Improving": 0.0, "Stable": 0.25, "Declining": 1.0}[app.health_trend]
        if app.health_trend == "Declining":
            reason = f"Health declined from {app.previous_health.upper()} to {app.health_status.upper()}"
        elif app.health_trend == "Improving":
            reason = f"Health improved from {app.previous_health.upper()} to {app.health_status.upper()}"
        else:
            reason = f"Health unchanged at {app.health_status.upper()}"
        parts.append(("Trend", w["trend_weight"], frac, reason))

    # 6. Data quality
    gaps: list[str] = []
    if not app.ever_scanned:
        gaps.append("no published scan results")
    else:
        if not app.flaw_severity_available:
            gaps.append("no flaw severity data")
        if app.total_modules == 0:
            gaps.append("no module data")
        if app.sca_status == "Unknown":
            gaps.append("SCA status unknown")
    dq = _clamp(len(gaps) / 2.0) if gaps else 0.0
    if not app.ever_scanned:
        dq = 1.0
    parts.append(("Data quality", w["data_quality_weight"], dq,
                  "Data gaps: " + ", ".join(gaps) if gaps else "Complete data"))

    total_w = sum(p[1] for p in parts)
    if total_w <= 0:
        return 0, "Normal", "GREEN", ["No scoring inputs available"]

    raw = sum(p[1] * p[2] for p in parts)
    score = int(round(100.0 * raw / total_w))
    score = max(0, min(100, score))

    capped = False
    cap = w.get("never_scanned_cap", 100)
    if not app.ever_scanned and score > cap:
        score = cap
        capped = True

    band, level = "Normal", "GREEN"
    for lo, hi, lbl, lvl in cfg["attention_bands"]:
        if lo <= score <= hi:
            band, level = lbl, lvl
            break

    # Drivers: only components that actually pushed the score up, biggest first.
    contribs = sorted(
        [(p[0], p[1] * p[2], p[3]) for p in parts if p[2] > 0],
        key=lambda x: -x[1],
    )
    drivers = [f"{reason} (+{int(round(100.0 * c / total_w))})" for _, c, reason in contribs]
    if capped:
        drivers.append(f"Score capped at {cap}: risk is unmeasured, not confirmed")
    if not drivers:
        drivers = ["No attention drivers; all measured dimensions are healthy"]
    return score, band, level, drivers


# ==========================================================================
# Aggregation
# ==========================================================================

@dataclass
class TenantMetrics:
    generated: str = ""
    total_apps: int = 0
    scanned_apps: int = 0
    never_scanned: int = 0
    good: int = 0
    fair: int = 0
    poor: int = 0
    unknown: int = 0
    stale: int = 0
    fresh: int = 0
    with_high_severity_issues: int = 0
    with_critical_or_high_flaws: int = 0
    flaw_severity_unavailable: int = 0
    total_flaws: int = 0
    open_policy_flaws: int = 0
    critical_flaws: int = 0
    high_flaws: int = 0
    avg_flaws_per_app: float = 0.0
    avg_scan_age: float | None = None
    median_scan_age: float | None = None
    total_sca_components: int = 0
    total_upload_mb: float = 0.0
    total_analysis_mb: float = 0.0
    missing_sca: int = 0
    attention_bands: dict[str, int] = field(default_factory=dict)
    quadrants: dict[str, int] = field(default_factory=dict)
    trends: dict[str, int] = field(default_factory=dict)

    @property
    def scan_compliance_pct(self) -> float:
        return _pct(self.fresh, self.total_apps)

    @property
    def healthy_pct(self) -> float:
        return _pct(self.good, self.total_apps)


def aggregate_tenant(apps: Sequence[ApplicationHealth], cfg: dict | None = None) -> TenantMetrics:
    cfg = cfg or DASHBOARD_CONFIG
    m = TenantMetrics(
        generated=datetime.now().strftime("%Y-%m-%d %H:%M"),
        total_apps=len(apps),
    )
    ages: list[int] = []
    for a in apps:
        if a.ever_scanned:
            m.scanned_apps += 1
        else:
            m.never_scanned += 1
        m.good += a.health_status == "Good"
        m.fair += a.health_status == "Fair"
        m.poor += a.health_status == "Poor"
        m.unknown += a.health_status == "Unknown"

        if a.scan_age_days is None:
            pass
        else:
            ages.append(a.scan_age_days)
            if a.scan_age_days <= cfg["scan_age"]["healthy_days"]:
                m.fresh += 1
            else:
                m.stale += 1

        m.with_high_severity_issues += a.high_severity_issue_count > 0
        if a.flaw_severity_available:
            m.critical_flaws += a.critical_flaws or 0
            m.high_flaws += a.high_flaws or 0
            if (a.critical_flaws or 0) or (a.high_flaws or 0):
                m.with_critical_or_high_flaws += 1
        elif a.ever_scanned:
            m.flaw_severity_unavailable += 1

        m.total_flaws += a.total_flaws
        m.open_policy_flaws += a.open_policy_flaws
        m.total_sca_components += a.sca_component_count
        m.total_upload_mb += a.total_upload_mb
        m.total_analysis_mb += a.analysis_size_mb
        m.missing_sca += a.sca_status == "Missing"

        m.attention_bands[a.attention_level] = m.attention_bands.get(a.attention_level, 0) + 1
        m.quadrants[a.quadrant] = m.quadrants.get(a.quadrant, 0) + 1
        m.trends[a.health_trend] = m.trends.get(a.health_trend, 0) + 1

    m.avg_flaws_per_app = round(m.total_flaws / len(apps), 1) if apps else 0.0
    if ages:
        m.avg_scan_age = round(sum(ages) / len(ages), 1)
        s = sorted(ages)
        mid = len(s) // 2
        m.median_scan_age = float(s[mid]) if len(s) % 2 else round((s[mid - 1] + s[mid]) / 2, 1)
    m.total_upload_mb = round(m.total_upload_mb, 1)
    m.total_analysis_mb = round(m.total_analysis_mb, 1)
    return m


def build_issue_heatmap(issues: Sequence[HealthIssue], total_apps: int,
                        cfg: dict | None = None,
                        prev_issue_counts: dict[int, int] | None = None) -> list[dict]:
    """One row per health check, including checks that nothing triggered.

    Severity-aware: colour combines the check's severity with how widespread it
    is, so 'a few serious issues' stays visually distinct from 'many small
    issues'.
    """
    cfg = cfg or DASHBOARD_CONFIG
    prev_issue_counts = prev_issue_counts or {}
    by_check: dict[int, list[HealthIssue]] = {}
    for i in issues:
        by_check.setdefault(i.check_number, []).append(i)

    p = cfg["issue_prevalence"]
    rows: list[dict] = []
    for num in sorted(CHECK_CATALOG):
        label, max_sev = CHECK_CATALOG[num]
        group = by_check.get(num, [])
        apps_hit = sorted({(g.app_name, g.sandbox) for g in group})
        n = len(apps_hit)
        pct = _pct(n, total_apps)

        if group:
            sev = "HIGH" if any(g.severity == "high" for g in group) else "MEDIUM"
        else:
            sev = max_sev.upper()

        if n == 0:
            level = "GREEN"
        elif sev == "HIGH":
            level = "RED" if pct >= p["common_pct"] else "ORANGE"
        else:
            level = "ORANGE" if pct >= p["widespread_pct"] else "YELLOW"

        prev_n = prev_issue_counts.get(num)
        if prev_n is None:
            trend = "Unknown"
        elif n < prev_n:
            trend = "Improving"
        elif n > prev_n:
            trend = "Declining"
        else:
            trend = "Stable"

        bus = sorted({g.business_unit for g in group if g.business_unit})
        rows.append({
            "Check #": num,
            "Issue": label,
            "Category": group[0].category if group else CHECK_CATEGORIES.get(num, "Other"),
            "Severity": sev,
            "Apps Affected": n,
            "Total Apps": total_apps,
            "% of Tenant": pct,
            "Occurrences": len(group),
            "Business Units Affected": len(bus),
            "Previous Apps Affected": prev_n if prev_n is not None else "N/A",
            "Trend": trend,
            "_level": level,
            "_apps": [f"{a}{' [' + s + ']' if s else ''}" for a, s in apps_hit],
            "Top Recommendation": next((g.recommendation for g in group if g.recommendation), ""),
        })

    sev_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (-r["Apps Affected"], sev_rank.get(r["Severity"], 3)))
    return rows


def build_tenant_trend(current: TenantMetrics,
                       prev_rows: dict | None,
                       cfg: dict | None = None) -> list[dict]:
    """Tenant-level previous-vs-current comparison.

    Direction is semantic: more healthy applications is an improvement, more
    total flaws is a deterioration. The arrow never simply follows the sign of
    the delta.
    """
    if not prev_rows:
        return []
    cfg = cfg or DASHBOARD_CONFIG

    rows = list(prev_rows.values())
    p_total = len(rows)
    p_good = sum(1 for r in rows if r.get("Health") == "Good")
    p_fair = sum(1 for r in rows if r.get("Health") == "Fair")
    p_poor = sum(1 for r in rows if r.get("Health") == "Poor")
    p_flaws = sum(_i(r.get("Total Flaws")) for r in rows)
    p_open = sum(_i(r.get("Open Affecting Policy")) for r in rows)
    p_stale = 0
    for r in rows:
        d = r.get("Days Since Scan")
        if isinstance(d, int) and d > cfg["scan_age"]["healthy_days"]:
            p_stale += 1
        elif isinstance(d, str) and d.isdigit() and int(d) > cfg["scan_age"]["healthy_days"]:
            p_stale += 1

    # (metric, previous, current, higher_is_better)
    spec = [
        ("Total applications", p_total, current.total_apps, None),
        ("Healthy (Good)", p_good, current.good, True),
        ("Fair", p_fair, current.fair, False),
        ("Poor", p_poor, current.poor, False),
        ("Stale scans", p_stale, current.stale, False),
        ("Total flaws", p_flaws, current.total_flaws, False),
        ("Open policy-affecting flaws", p_open, current.open_policy_flaws, False),
    ]

    out: list[dict] = []
    for name, prev, curr, higher_better in spec:
        delta = curr - prev
        if delta == 0:
            arrow, level = "=", "GRAY"
        elif higher_better is None:
            arrow, level = ("+" if delta > 0 else "-"), "GRAY"
        else:
            improved = (delta > 0) == bool(higher_better)
            arrow = "UP" if delta > 0 else "DOWN"
            level = "GREEN" if improved else "RED"
        out.append({
            "Metric": name,
            "Previous": prev,
            "Current": curr,
            "Change": f"{delta:+d}" if delta else "0",
            "Direction": arrow,
            "_level": level,
        })
    return out


def build_heatmap_rows(apps: Sequence[ApplicationHealth],
                       cfg: dict | None = None) -> list[dict]:
    """Application heatmap. '_lv_<Column>' keys carry the colour for each cell
    and are hidden from the rendered table."""
    cfg = cfg or DASHBOARD_CONFIG
    rows: list[dict] = []
    for a in apps:
        crit = a.critical_flaws if a.flaw_severity_available else "N/A"
        high = a.high_flaws if a.flaw_severity_available else "N/A"
        med = a.medium_flaws if a.flaw_severity_available else "N/A"
        low = a.low_flaws if a.flaw_severity_available else "N/A"

        row = {
            "Application": a.app_name,
            "Sandbox": a.sandbox or "(policy)",
            "Business Unit": a.business_unit or "Unknown",
            "Attention Score": a.attention_score,
            "Priority": a.attention_band,
            "Scan Health": a.health_status,
            "Scan Age": a.scan_age_days if a.scan_age_days is not None else "Never",
            "Scan Age Bucket": a.scan_age_bucket,
            "Findings Risk": a.security_findings_risk if a.flaw_severity_available else "Unknown",
            "Very High Flaws": crit,
            "High Flaws": high,
            "Medium Flaws": med,
            "Low Flaws": low,
            "Total Flaws": a.total_flaws if a.ever_scanned else "N/A",
            "Open Policy Flaws": a.open_policy_flaws if a.ever_scanned else "N/A",
            "SCA": a.sca_status,
            "SCA Components": a.sca_component_count,
            "Health Issues": a.health_issue_count if a.ever_scanned else "N/A",
            "High-Severity Issues": a.high_severity_issue_count if a.ever_scanned else "N/A",
            "Upload Size (MB)": a.total_upload_mb,
            "Last Scan": a.last_scan_date or "Never",
            "Previous Health": a.previous_health or "N/A",
            "Change": {"Improving": "UP", "Declining": "DOWN",
                       "Stable": "=", "New": "NEW"}.get(a.health_trend, "N/A"),
            "Trend": a.health_trend,
            "Quadrant": a.quadrant,
            "Top Issue": a.top_issue or "None",
            # hidden colour helpers
            "_lv_Attention Score": a.attention_level,
            "_lv_Priority": a.attention_level,
            "_lv_Scan Health": _health_risk(a.health_status),
            "_lv_Scan Age": _scan_age_risk(a.scan_age_days, cfg),
            "_lv_Scan Age Bucket": _scan_age_risk(a.scan_age_days, cfg),
            "_lv_Findings Risk": a.security_findings_risk,
            "_lv_Very High Flaws": _sev_cell_level(crit, 1, 1),
            "_lv_High Flaws": _sev_cell_level(high, 1, cfg["flaws"]["high_red"]),
            "_lv_SCA": _sca_risk(a.sca_status),
            "_lv_Health Issues": (_issue_count_risk(a.health_issue_count, cfg["health_issues"])
                                  if a.ever_scanned else "GRAY"),
            "_lv_High-Severity Issues": (_issue_count_risk(a.high_severity_issue_count,
                                                           cfg["high_severity_issues"])
                                         if a.ever_scanned else "GRAY"),
            "_lv_Previous Health": _health_risk(a.previous_health) if a.previous_health else "GRAY",
            "_lv_Change": _trend_risk(a.health_trend),
            "_lv_Trend": _trend_risk(a.health_trend),
            "_key": a.key,
        }
        rows.append(row)
    return rows


def _sev_cell_level(v: Any, yellow_at: int, red_at: int) -> str:
    if not isinstance(v, int):
        return "GRAY"
    if v >= red_at:
        return "RED"
    if v >= yellow_at:
        return "ORANGE"
    return "GREEN"


def build_priority_list(apps: Sequence[ApplicationHealth],
                        n: int | None = None,
                        cfg: dict | None = None) -> list[dict]:
    cfg = cfg or DASHBOARD_CONFIG
    n = n or cfg["top_n_attention"]
    out: list[dict] = []
    for rank, a in enumerate(apps[:n], 1):
        ch = (f"{a.critical_flaws} / {a.high_flaws}"
              if a.flaw_severity_available else "Not available")
        out.append({
            "Priority": rank,
            "Application": a.display_name,
            "Business Unit": a.business_unit or "Unknown",
            "Attention Score": a.attention_score,
            "Band": a.attention_band,
            "Scan Health": a.health_status,
            "Very High / High Flaws": ch,
            "Scan Age (days)": a.scan_age_days if a.scan_age_days is not None else "Never scanned",
            "Top Issue": a.top_issue or "None",
            "Trend": a.health_trend,
            "Why": "; ".join(a.attention_drivers[:2]),
            "_level": a.attention_level,
        })
    return out


# ==========================================================================
# Excel rendering
# ==========================================================================

_FONT = "Arial"
_TITLE_FILL = PatternFill("solid", fgColor="14213D")
_BAND_FILL = PatternFill("solid", fgColor="1F4E79")
_SUB_FILL = PatternFill("solid", fgColor="E8EDF3")
_KPI_FILL = PatternFill("solid", fgColor="F5F7FA")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _lvl_fill(level: str, cfg: dict, strong: bool = False) -> PatternFill:
    key = "colors" if strong else "fills"
    return PatternFill("solid", fgColor=cfg[key].get(level, cfg[key]["GRAY"]))


def _lvl_font(level: str, cfg: dict, bold: bool = False, size: int = 10) -> Font:
    return Font(name=_FONT, size=size, bold=bold, color=cfg["colors"].get(level, "000000"))


def _write_title(ws, row: int, text: str, width: int, fill=_BAND_FILL,
                 size: int = 11, color: str = "FFFFFF") -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill
    c.font = Font(name=_FONT, size=size, bold=True, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22 if size <= 12 else 30
    return row + 1


def _write_table(ws, row: int, headers: list[str], data: list[list],
                 levels: list[list[str]] | None, cfg: dict,
                 spans: list[int] | None = None,
                 left_cols: Iterable[int] = (),
                 row_height: int | None = None) -> int:
    """Write a table on the fixed dashboard grid.

    `spans` gives the number of grid columns each logical column occupies, so
    long-text columns can breathe without any block resizing the shared grid.
    Column widths are owned by the sheet, never by individual tables.
    """
    spans = spans or [1] * len(headers)
    left = set(left_cols)

    def _cells(r: int, values: list, styler) -> None:
        col = 1
        for idx, v in enumerate(values):
            span = spans[idx]
            if span > 1:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + span - 1)
            c = ws.cell(row=r, column=col, value=v)
            styler(c, idx)
            for cc in range(col, col + span):
                ws.cell(row=r, column=cc).border = _BORDER
            col += span

    def _hdr_style(c, idx):
        c.fill = _SUB_FILL
        c.font = Font(name=_FONT, size=9, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="left" if idx in left else "center",
                                vertical="center", wrap_text=True, indent=1 if idx in left else 0)

    _cells(row, headers, _hdr_style)
    row += 1

    for ri, rowvals in enumerate(data):
        def _body_style(c, idx, ri=ri):
            c.font = Font(name=_FONT, size=10)
            c.alignment = Alignment(horizontal="left" if idx in left else "center",
                                    vertical="center", wrap_text=True,
                                    indent=1 if idx in left else 0)
            if levels:
                lv = levels[ri][idx]
                if lv:
                    c.fill = _lvl_fill(lv, cfg)
                    if lv in ("RED", "ORANGE"):
                        c.font = Font(name=_FONT, size=10, bold=True, color=cfg["colors"][lv])
        _cells(row + ri, rowvals, _body_style)
        if row_height:
            ws.row_dimensions[row + ri].height = row_height
    return row + len(data)


def write_dashboard_sheets(wb: Workbook,
                           apps: Sequence[ApplicationHealth],
                           issues: Sequence[HealthIssue],
                           metrics: TenantMetrics,
                           tenant_trend: list[dict],
                           issue_rows: list[dict],
                           cfg: dict | None = None) -> None:
    """Insert 'Executive Dashboard', 'App Heatmap' and 'Issue Heatmap' as the
    first three sheets of an existing workbook. Existing sheets are untouched.

    The dashboard sheet uses a mixed block layout, which cannot carry an Excel
    autofilter. The two heatmap sheets carry the filterable/sortable tables.
    """
    cfg = cfg or DASHBOARD_CONFIG
    _write_exec_sheet(wb, apps, metrics, tenant_trend, issue_rows, cfg)
    _write_app_heatmap_sheet(wb, apps, cfg)
    _write_issue_heatmap_sheet(wb, issue_rows, metrics, cfg)
    # Order: dashboard, app heatmap, issue heatmap, then everything pre-existing.
    order = ["Executive Dashboard", "App Heatmap", "Issue Heatmap"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else len(order))


# Fixed 12-column grid for the dashboard sheet. Set once, never overridden.
_EXEC_GRID = 12
_EXEC_WIDTHS = {1: 26}
for _c in range(2, _EXEC_GRID + 1):
    _EXEC_WIDTHS[_c] = 13


def _write_exec_sheet(wb: Workbook, apps: Sequence[ApplicationHealth],
                      m: TenantMetrics, tenant_trend: list[dict],
                      issue_rows: list[dict], cfg: dict) -> None:
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    W = _EXEC_GRID
    for ci, wd in _EXEC_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = wd

    def band(r: int, text: str) -> int:
        return _write_title(ws, r, text, W)

    def note(r: int, text: str) -> int:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=W)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=_FONT, size=9, italic=True, color="555555")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        return r + 1

    r = 1
    # ---- title
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=W)
    c = ws.cell(row=r, column=1, value="VERACODE TENANT SCAN HEALTH")
    c.fill = _TITLE_FILL
    c.font = Font(name=_FONT, size=18, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 26
    ws.row_dimensions[r + 1].height = 16
    r += 2
    r = note(r, f"Generated {m.generated}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=W)
    c = ws.cell(row=r, column=1, value=DISCLAIMER)
    c.font = Font(name=_FONT, size=9, italic=True, color="8A6D00")
    c.fill = PatternFill("solid", fgColor="FDF6E3")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _BORDER
    ws.row_dimensions[r].height = 20
    r += 2

    # ---- KPI cards
    comp_lv = ("GREEN" if m.scan_compliance_pct >= 80
               else "YELLOW" if m.scan_compliance_pct >= 60 else "RED")
    r = _kpi_row(ws, r, [
        ("TOTAL APPS", f"{m.total_apps:,}", "profiles evaluated", "GRAY"),
        ("HEALTHY", f"{m.good:,}", f"{_pct(m.good, m.total_apps)}% of tenant", "GREEN"),
        ("NEEDS ATTENTION", f"{m.fair:,}", f"{_pct(m.fair, m.total_apps)}% of tenant", "YELLOW"),
        ("POOR", f"{m.poor:,}", f"{_pct(m.poor, m.total_apps)}% of tenant", "RED"),
        ("UNKNOWN", f"{m.unknown:,}", f"{_pct(m.unknown, m.total_apps)}% never scanned", "GRAY"),
    ], cfg, span=W)

    r = _kpi_row(ws, r, [
        ("SCAN COMPLIANCE", f"{m.fresh:,} / {m.total_apps:,}",
         f"{m.scan_compliance_pct}% scanned within {cfg['scan_age']['healthy_days']}d", comp_lv),
        ("STALE SCANS", f"{m.stale:,} / {m.total_apps:,}",
         f"{_pct(m.stale, m.total_apps)}% of tenant", "RED" if m.stale else "GREEN"),
        ("HIGH-SEV ISSUES", f"{m.with_high_severity_issues:,} / {m.total_apps:,}",
         f"{_pct(m.with_high_severity_issues, m.total_apps)}% of tenant",
         "RED" if m.with_high_severity_issues else "GREEN"),
        ("OPEN POLICY FLAWS", f"{m.open_policy_flaws:,}",
         f"avg {m.avg_flaws_per_app} flaws per app", "GRAY"),
        ("VERY HIGH / HIGH", f"{m.critical_flaws:,} / {m.high_flaws:,}",
         (f"{m.flaw_severity_unavailable} app(s) lack severity data"
          if m.flaw_severity_unavailable else "open flaws across tenant"),
         "RED" if (m.critical_flaws or m.high_flaws) else "GREEN"),
    ], cfg, span=W)

    # ---- tenant health distribution
    r = band(r, "TENANT SCAN HEALTH DISTRIBUTION")
    for label, n, lv in (("Good", m.good, "GREEN"), ("Fair", m.fair, "YELLOW"),
                         ("Poor", m.poor, "RED"), ("Unknown", m.unknown, "GRAY")):
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name=_FONT, size=10, bold=True, color=cfg["colors"][lv])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=_ctx(n, m.total_apps))
        c.font = Font(name=_FONT, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=W)
        c = ws.cell(row=r, column=4, value="\u2588" * int(round(50 * _pct(n, m.total_apps) / 100.0)))
        c.font = Font(name=_FONT, size=10, color=cfg["colors"][lv])
        c.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    r = note(r, "Unknown covers profiles with no published scan results. "
                "They are never counted as healthy.")
    r += 1

    # ---- attention distribution
    r = band(r, "ATTENTION SCORE DISTRIBUTION")
    meaning = {"RED": "Immediate attention (80-100)", "ORANGE": "High priority (60-79)",
               "YELLOW": "Attention required (40-59)", "GREEN": "Normal or monitor (0-39)"}
    data, levels = [], []
    for lv in ("RED", "ORANGE", "YELLOW", "GREEN"):
        n = m.attention_bands.get(lv, 0)
        data.append([lv, _ctx(n, m.total_apps), meaning[lv]])
        levels.append([lv, "", ""])
    r = _write_table(ws, r, ["Band", "Applications", "Meaning"], data, levels, cfg,
                     spans=[2, 3, 7], left_cols=(2,))
    r += 1

    # ---- priority matrix
    r = band(r, "PRIORITY MATRIX   scan health vs security findings")
    r = note(r, "Two independent dimensions. Built only from indicators the scanner produces; "
                "business criticality is not available from the Veracode data and is not inferred.")
    q = m.quadrants
    grid = [
        ["", "LOW scan-health risk", "HIGH scan-health risk"],
        ["HIGH findings risk", f"REVIEW\n{q.get('Review', 0)} apps", f"IMMEDIATE\n{q.get('Immediate', 0)} apps"],
        ["LOW findings risk", f"MONITOR\n{q.get('Monitor', 0)} apps", f"ATTENTION\n{q.get('Attention', 0)} apps"],
    ]
    glevels = [["", "", ""], ["", "ORANGE", "RED"], ["", "GREEN", "YELLOW"]]
    for ri, gr in enumerate(grid):
        col = 1
        for ci, v in enumerate(gr):
            span = 1 if ci == 0 else 5
            if span > 1:
                ws.merge_cells(start_row=r + ri, start_column=col, end_row=r + ri,
                               end_column=col + span - 1)
            cell = ws.cell(row=r + ri, column=col, value=v)
            cell.font = Font(name=_FONT, size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            lv = glevels[ri][ci]
            for cc in range(col, col + span):
                ws.cell(row=r + ri, column=cc).border = _BORDER
                if lv:
                    ws.cell(row=r + ri, column=cc).fill = _lvl_fill(lv, cfg)
            col += span
        ws.row_dimensions[r + ri].height = 38 if ri else 18
    r += 3
    r = note(r, f"Insufficient data: {q.get('Insufficient Data', 0)} application profile(s).")
    r += 1

    # ---- top organisational issues
    r = band(r, "TOP ORGANISATIONAL ISSUES")
    data, levels = [], []
    for row_ in [x for x in issue_rows if x["Apps Affected"] > 0][:cfg["top_n_issues"]]:
        data.append([row_["Issue"], row_["Severity"],
                     _ctx(row_["Apps Affected"], row_["Total Apps"]),
                     row_["Category"], row_["Trend"]])
        levels.append(["", row_["_level"], row_["_level"], "", _trend_risk(row_["Trend"])])
    if not data:
        data = [["No failing health checks across the tenant", "", "", "", ""]]
        levels = [["GREEN", "", "", "", ""]]
    r = _write_table(ws, r, ["Issue", "Severity", "Apps Affected", "Category", "Trend"],
                     data, levels, cfg, spans=[4, 1, 3, 2, 2], left_cols=(0,))
    r = note(r, "Full list of all 31 health checks is on the 'Issue Heatmap' sheet.")
    r += 1

    # ---- tenant trend
    r = band(r, "TENANT TREND vs PREVIOUS REPORT")
    if tenant_trend:
        arrows = {"UP": "\u2191", "DOWN": "\u2193", "=": "\u2192", "+": "\u2192", "-": "\u2192"}
        data = [[t["Metric"], t["Previous"], t["Current"], t["Change"],
                 arrows.get(t["Direction"], t["Direction"])] for t in tenant_trend]
        levels = [["", "", "", t["_level"], t["_level"]] for t in tenant_trend]
        r = _write_table(ws, r, ["Metric", "Previous", "Current", "Change", "Direction"],
                         data, levels, cfg, spans=[4, 2, 2, 2, 2], left_cols=(0,))
        r = note(r, "Direction is semantic: fewer poor applications is green, more total flaws "
                    "is red. Arrows do not simply follow the sign of the change.")
    else:
        r = note(r, "No previous report supplied. Re-run with --previous-report to enable "
                    "trend analysis.")
    r += 1

    # ---- applications requiring attention
    r = band(r, "APPLICATIONS REQUIRING ATTENTION")
    r = note(r, "Ranked by the transparent Attention Score. This is a scan-health and findings "
                "prioritisation ranking, not a 'most vulnerable applications' list.")
    plist = build_priority_list(apps, cfg=cfg)
    data, levels = [], []
    for p in plist:
        data.append([p["Priority"], p["Application"], p["Attention Score"], p["Scan Health"],
                     p["Very High / High Flaws"], p["Scan Age (days)"], p["Top Issue"],
                     p["Trend"], p["Why"]])
        levels.append(["", "", p["_level"], _health_risk(p["Scan Health"]),
                       "", "", "", _trend_risk(p["Trend"]), ""])
    if not data:
        data, levels = [["", "No applications scored", "", "", "", "", "", "", ""]], None
    r = _write_table(ws, r,
                     ["#", "Application", "Score", "Health", "V.High / High",
                      "Scan age", "Top issue", "Trend", "Why this score"],
                     data, levels, cfg,
                     spans=[1, 2, 1, 1, 1, 1, 2, 1, 2], left_cols=(1, 6, 8),
                     row_height=60)
    r += 1
    r = note(r, "Drill down on the 'App Heatmap' and 'Issue Heatmap' sheets, both filterable "
                "and sortable. Raw evidence remains on the original detail sheets.")
    ws.freeze_panes = "A5"


def _kpi_row(ws, r: int, kpis: list[tuple[str, str, str, str]], cfg: dict, span: int) -> int:
    """Render KPI cards spread evenly across the full `span` of grid columns."""
    n = len(kpis)
    base, rem = divmod(span, n)
    col = 1
    for idx, (label, value, sub, lv) in enumerate(kpis):
        width = base + (1 if idx < rem else 0)
        end = col + width - 1
        for rr, text, font in (
            (r, label, Font(name=_FONT, size=9, bold=True, color="555555")),
            (r + 1, value, Font(name=_FONT, size=18, bold=True, color=cfg["colors"][lv])),
            (r + 2, sub, Font(name=_FONT, size=8, color="666666")),
        ):
            ws.merge_cells(start_row=rr, start_column=col, end_row=rr, end_column=end)
            c = ws.cell(row=rr, column=col, value=text)
            c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cc in range(col, end + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = _KPI_FILL
                cell.border = _BORDER
        col = end + 1
    ws.row_dimensions[r].height = 15
    ws.row_dimensions[r + 1].height = 26
    ws.row_dimensions[r + 2].height = 15
    return r + 4


_HEATMAP_WIDTHS = {
    "Application": 34, "Sandbox": 16, "Business Unit": 20, "Attention Score": 9,
    "Priority": 20, "Scan Health": 11, "Scan Age": 9, "Scan Age Bucket": 13,
    "Findings Risk": 13, "Very High Flaws": 11, "High Flaws": 9, "Medium Flaws": 11,
    "Low Flaws": 9, "Total Flaws": 10, "Open Policy Flaws": 12, "SCA": 10,
    "SCA Components": 12, "Health Issues": 11, "High-Severity Issues": 13,
    "Upload Size (MB)": 13, "Last Scan": 20, "Previous Health": 13, "Change": 9,
    "Trend": 11, "Quadrant": 15,
    "Top Issue": 44,
}


def _write_app_heatmap_sheet(wb: Workbook, apps: Sequence[ApplicationHealth], cfg: dict) -> None:
    ws = wb.create_sheet("App Heatmap")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    rows = build_heatmap_rows(apps, cfg)
    headers = [h for h in (rows[0].keys() if rows else []) if not h.startswith("_")]
    if not rows:
        ws["A1"] = "No applications."
        return

    title_row = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1,
                value="APPLICATION HEATMAP    |    "
                      "Scan health and security findings are separate dimensions. "
                      + DISCLAIMER)
    c.fill = _TITLE_FILL
    c.font = Font(name=_FONT, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    hr = title_row + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.fill = _BAND_FILL
        c.font = Font(name=_FONT, size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = _HEATMAP_WIDTHS.get(h, 13)
    ws.row_dimensions[hr].height = 30

    for ri, row in enumerate(rows, hr + 1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h))
            c.font = Font(name=_FONT, size=9)
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="left" if h in ("Application", "Top Issue", "Business Unit") else "center",
                vertical="center")
            lv = row.get(f"_lv_{h}")
            if lv:
                c.fill = _lvl_fill(lv, cfg)
                if lv in ("RED", "ORANGE"):
                    c.font = Font(name=_FONT, size=9, bold=True, color=cfg["colors"][lv])

    last = hr + len(rows)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = ws.cell(row=hr + 1, column=4)


def _write_issue_heatmap_sheet(wb: Workbook, issue_rows: list[dict],
                               m: TenantMetrics, cfg: dict) -> None:
    ws = wb.create_sheet("Issue Heatmap")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:2"
    headers = ["Check #", "Issue", "Category", "Severity", "Apps Affected",
               "% of Tenant", "Prevalence", "Occurrences", "Business Units Affected",
               "Previous Apps Affected", "Trend", "Top Recommendation"]
    widths = {"Check #": 8, "Issue": 46, "Category": 17, "Severity": 10,
              "Apps Affected": 13, "% of Tenant": 12, "Prevalence": 24,
              "Occurrences": 12, "Business Units Affected": 13,
              "Previous Apps Affected": 13, "Trend": 12, "Top Recommendation": 60}

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1,
                value=f"ORGANISATIONAL ISSUE HEATMAP    |    All 31 health checks across "
                      f"{m.total_apps:,} application profiles")
    c.fill = _TITLE_FILL
    c.font = Font(name=_FONT, size=11, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    hr = 2
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.fill = _BAND_FILL
        c.font = Font(name=_FONT, size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)
    ws.row_dimensions[hr].height = 30

    for ri, row in enumerate(issue_rows, hr + 1):
        bar = "#" * int(round(30 * min(row["% of Tenant"], 100) / 100.0))
        vals = [row["Check #"], row["Issue"], row["Category"], row["Severity"],
                _ctx(row["Apps Affected"], row["Total Apps"]),
                row["% of Tenant"] / 100.0, bar, row["Occurrences"],
                row["Business Units Affected"], row["Previous Apps Affected"],
                row["Trend"], row["Top Recommendation"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name=_FONT, size=9)
            c.border = _BORDER
            c.alignment = Alignment(
                horizontal="left" if headers[ci - 1] in ("Issue", "Top Recommendation", "Category")
                else "center",
                vertical="center", wrap_text=headers[ci - 1] == "Top Recommendation")
            if headers[ci - 1] == "% of Tenant":
                c.number_format = "0.0%"
            if headers[ci - 1] in ("Severity", "Apps Affected"):
                c.fill = _lvl_fill(row["_level"], cfg)
            if headers[ci - 1] == "Prevalence":
                c.font = Font(name="Consolas", size=9, color=cfg["colors"][row["_level"]])
            if headers[ci - 1] == "Trend":
                lv = _trend_risk(row["Trend"])
                if lv != "GRAY":
                    c.fill = _lvl_fill(lv, cfg)

    last = hr + len(issue_rows)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = ws.cell(row=hr + 1, column=3)


# ==========================================================================
# Top-level entry point used by script.py
# ==========================================================================

def generate_dashboard(health_rows: Sequence[dict],
                       issue_records: Sequence[dict],
                       prev_rows: dict | None = None,
                       cfg: dict | None = None,
                       check_categories: dict[int, str] | None = None,
                       prev_issue_counts: dict[int, int] | None = None) -> dict:
    """Normalize, score and aggregate. Returns everything the writers need."""
    cfg = cfg or DASHBOARD_CONFIG
    apps, issues = build_application_health(
        health_rows, issue_records, prev_rows, cfg, check_categories)
    metrics = aggregate_tenant(apps, cfg)
    issue_rows = build_issue_heatmap(issues, metrics.total_apps, cfg, prev_issue_counts)
    tenant_trend = build_tenant_trend(metrics, prev_rows, cfg)
    return {
        "apps": apps,
        "issues": issues,
        "metrics": metrics,
        "issue_rows": issue_rows,
        "tenant_trend": tenant_trend,
        "config": cfg,
    }


def write_dashboard_xlsx(path: str, bundle: dict) -> None:
    """Standalone dashboard workbook (used when the main report is CSV/JSON)."""
    wb = Workbook()
    wb.remove(wb.active)
    write_dashboard_sheets(wb, bundle["apps"], bundle["issues"], bundle["metrics"],
                           bundle["tenant_trend"], bundle["issue_rows"], bundle["config"])
    wb.save(path)


# ==========================================================================
# Main
# ==========================================================================

def _print_summary(health: list[dict]) -> None:
    tot = len(health)
    good = sum(1 for r in health if r.get("Health")=="Good")
    fair = sum(1 for r in health if r.get("Health")=="Fair")
    poor = sum(1 for r in health if r.get("Health")=="Poor")
    print(f"\nTenant Summary: {tot} apps - Good: {good}, Fair: {fair}, Poor: {poor}")
    counter: Counter[str] = Counter()
    for r in health:
        txt = r.get("Issues","")
        if txt == "None": continue
        for part in txt.split("; "):
            clean = re.sub(r'\[(?:HIGH|MEDIUM|LOW)\]\s*', '', part).strip()
            clean = re.sub(r'"[^"]*"', '(name)', clean)
            if clean: counter[clean] += 1
    if counter:
        print("Top issues:")
        for pattern, count in counter.most_common(3):
            print(f"  [{count}x] {pattern[:100]}")


def main() -> None:
    p = argparse.ArgumentParser(description="Veracode Tenant-Wide Scan Health v3.0")
    p.add_argument("--output", default=f"scan_health_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    p.add_argument("--output-format", choices=["xlsx","csv","json"], default="xlsx")
    p.add_argument("--max-apps", type=int, default=0)
    p.add_argument("--delay", type=float, default=0.5)
    p.add_argument("--skip-no-scan", action="store_true")
    p.add_argument("--include-sandboxes", action="store_true")
    p.add_argument("--region", choices=["commercial","eu"], default="commercial")
    p.add_argument("--app-name-filter", default=None, help="Regex to filter app names")
    p.add_argument("--parallel", type=int, default=1, help="Concurrent workers (default 1)")
    p.add_argument("--resume", default=None, help="Path to prior partial xlsx to skip processed apps")
    p.add_argument("--previous-report", default=None, help="Path to prior xlsx for trend analysis")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--log-level", choices=["DEBUG","INFO","WARNING"], default="INFO")
    p.add_argument("--timeout", type=int, default=120)
    p.add_argument("--skip-checks", default=None, help="Comma-separated check numbers to skip")
    p.add_argument("--dashboard", action="store_true",
                   help="Generate the executive dashboard (heatmaps, KPIs, attention scores)")
    p.add_argument("--dashboard-output", default=None,
                   help="Dashboard workbook path. Only used when --output-format is csv or "
                        "json; with xlsx the dashboard sheets are added to the main report")
    args = p.parse_args()

    logging.basicConfig(level=getattr(logging, args.log_level),
                        format="%(asctime)s %(levelname)-5s %(message)s", datefmt="%H:%M:%S",
                        force=True)

    skip_checks: set[int] = set()
    if args.skip_checks:
        skip_checks = {int(x.strip()) for x in args.skip_checks.split(",")}
        log.info("[*] Skipping checks: %s", sorted(skip_checks))

    resume_keys: set[tuple[str, str]] | None = None
    if args.resume:
        resume_keys = _load_resume_keys(args.resume)
        log.info("[*] Resume: %d apps already processed", len(resume_keys))

    prev_data: dict | None = None
    if args.previous_report:
        prev_data = _load_previous(args.previous_report)
        log.info("[*] Previous report: %d apps loaded for trend analysis", len(prev_data))

    name_filter = re.compile(args.app_name_filter) if args.app_name_filter else None

    try:
        with VeracodeClient(args.region, timeout=args.timeout) as client:
            log.info("[*] Region: %s", args.region)
            apps = client.get_apps()
            log.info("[*] Found %d apps", len(apps))

            if name_filter:
                apps = [a for a in apps if name_filter.search(a["name"])]
                log.info("[*] Filtered to %d apps", len(apps))
            if args.max_apps:
                apps = apps[:args.max_apps]

            if args.dry_run:
                print(f"Would process {len(apps)} apps:")
                for a in apps:
                    print(f"  {a['name']} (id={a.get('legacy_id')})")
                return

            all_sr: list[ScanResult] = []; all_mr: list[ModuleRow] = []
            all_fr: list[FileRow] = []; all_rr: list[RecommendationRow] = []
            all_ai: list[AggIssue] = []; all_tr: list[TrendRow] = []
            lock = threading.Lock()
            delay_lock = threading.Lock()

            def _do_app(idx_app: tuple[int, dict]) -> None:
                idx, app = idx_app
                log.info("[%d/%d] %s (id=%s)", idx, len(apps), app["name"], app.get("legacy_id"))
                try:
                    sr, mr, fr, rr, ai, tr = _process_app(
                        client, app, args.skip_no_scan, args.include_sandboxes,
                        skip_checks, prev_data, resume_keys)
                    with lock:
                        all_sr.extend(sr); all_mr.extend(mr); all_fr.extend(fr)
                        all_rr.extend(rr); all_ai.extend(ai); all_tr.extend(tr)
                    for s in sr:
                        sb = f' [{s.sandbox}]' if s.sandbox else ""
                        log.info("    %s | Issues: %d%s", s.health, s.total_issues, sb)
                except AuthError as e:
                    log.error("Authentication failed: %s", e)
                    raise
                except Exception as e:
                    log.warning("    [!] Failed: %s", e)
                    with lock:
                        all_sr.append(_empty_result(app, app.get("legacy_id",0)))
                if args.delay > 0:
                    with delay_lock:
                        time.sleep(args.delay)

            if args.parallel > 1:
                with ThreadPoolExecutor(max_workers=args.parallel) as pool:
                    futs = {pool.submit(_do_app, (i, a)): a for i, a in enumerate(apps, 1)}
                    for fut in as_completed(futs):
                        try: fut.result()
                        except AuthError: raise
                        except Exception as e: log.warning("Worker error: %s", e)
            else:
                for i, app in enumerate(apps, 1):
                    _do_app((i, app))

            # Build output rows
            h_rows = [s.to_row() for s in all_sr]
            m_rows = [m.to_row() for m in all_mr]
            f_rows = [f.to_row() for f in all_fr]
            r_rows = [r.to_row() for r in all_rr]
            t_rows = [t.to_row() for t in all_tr]
            agg = _build_aggregation(all_ai, len(h_rows))

            log.info("\n[*] Writing %d health / %d module / %d file / %d rec / %d trend rows...",
                     len(h_rows), len(m_rows), len(f_rows), len(r_rows), len(t_rows))

            # Dashboard layer: consumes the rows already built above. No extra API calls.
            bundle = None
            if args.dashboard:
                if _dash is None:
                    log.warning("[!] --dashboard requested but dashboard.py is not importable")
                else:
                    bundle = generate_dashboard(
                        health_rows=h_rows, issue_records=[asdict(a) for a in all_ai],
                        prev_rows=prev_data, check_categories=CHECK_CATEGORIES,
                        prev_issue_counts=(_load_previous_agg(args.previous_report)
                                           if args.previous_report else None))
                    log.info("[*] Dashboard: %d profiles scored, %d structured issues",
                             len(bundle["apps"]), len(bundle["issues"]))

            fmt = args.output_format
            if fmt == "xlsx":
                write_excel(h_rows, m_rows, f_rows, r_rows, t_rows, agg, args.output,
                            dash_bundle=bundle)
            elif fmt == "csv":
                write_csv(h_rows, m_rows, f_rows, r_rows, t_rows, agg, args.output)
            elif fmt == "json":
                write_json(h_rows, m_rows, f_rows, t_rows, args.output)

            # For csv/json the dashboard has nowhere to live, so it gets its own workbook.
            if bundle is not None and fmt != "xlsx":
                xp = args.dashboard_output or f"{Path(args.output).with_suffix('')}_dashboard.xlsx"
                write_dashboard_xlsx(xp, bundle)
                log.info("[+] Dashboard workbook: %s", xp)

            _print_summary(h_rows)

    except AuthError as e:
        log.error("FATAL: %s", e)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
